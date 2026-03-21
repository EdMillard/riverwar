"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from pathlib import Path
import colorado.allb as all_b
import colorado.lb as lb
import pandas as pd
from sheet.sheet import Sheet
import openpyxl
from sheet import sheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from typing import List, Dict, Union
import csv

class LBReservoirCUL(Sheet):
    def __init__(self):
        self.path:Path = Path('data/USBR_Lower_Colorado_CUL/Reservoir')
        headers = [
            lb.LAKE_MEAD_CUL,
            lb.LC_RESERVOIR_TOTAL_CUL,
            lb.LAKE_MOHAVE_CUL, lb.LAKE_HAVASU_CUL,
            lb.SENATOR_WASH_CUL, lb.DIVERSION_DAMS_CUL
        ]
        super().__init__(headers, start_year=1971, end_year=2024)
        self.years: List[int] = list(range(self.start_year, self.end_year+1))

        # lower_basin_reservoir_cul_from_excel(self.path, self.years)
        # generate_cul_totals(self.path)


    def load_df(self, df_compact : pd.DataFrame) -> None:
        divisor = 1
        path = self.path

        df = sheet.read_csv(path / 'lake_mead.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.LAKE_MEAD_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'lake_mohave.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.LAKE_MOHAVE_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'lake_havasu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.LAKE_HAVASU_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'senator_wash.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.SENATOR_WASH_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'diversion_dams.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.DIVERSION_DAMS_CUL, divisor=divisor)

    def build_sheet(self) -> None:
        ws: Worksheet = self.ws

        sheet.add_borders_to_column(ws, 1, 1, ws.max_row, end_col=ws.max_column, which='outer')

        sheet.formula_average(ws, self.df, self.years)
        sheet.add_borders_to_column(ws, 1, ws.max_row-1, ws.max_row-1, end_col=ws.max_column, which='outer')

        lc_reservoir = [lb.LAKE_MOHAVE_CUL, lb.LAKE_HAVASU_CUL, lb.LAKE_MEAD_CUL, lb.DIVERSION_DAMS_CUL]
        sheet.formula_add(ws, self.df, lb.LC_RESERVOIR_TOTAL_CUL, lc_reservoir)
        self.set_bg(lb.LAKE_MEAD_CUL, to=lb.DIVERSION_DAMS_CUL, color=all_b.USBR_LB_CUL_RESERVOIR_BG)

        sheet.clear_range(ws, ws.max_row, ws.max_row, 1, ws.max_column)

        self.format_header()

        self.set_column_width(lb.LC_RESERVOIR_TOTAL_CUL, 6, to=lb.DIVERSION_DAMS_CUL)

def get_reservoir_node(reservoir_name: str,  years: List[int], nodes: Dict) -> pd.DataFrame:
    node_name = f"{reservoir_name}"
    df: Union[pd.DataFrame, None] = nodes.get(node_name, None)
    if df is None:
        df = sheet.create_month_year_df(years)
        nodes[node_name] = df
    else:
        pass

    return df

def lower_basin_reservoir_cul_from_excel(out_path: Path, years):
    wb: Workbook = openpyxl.load_workbook('data/Colorado_River/1971-2024 Lower Colorado River System CUL Data.xlsx', data_only=True)
    ws: Worksheet = wb['Mainstream_Reservoirs']

    header_row: int = 1

    sheet.ensure_directory(out_path)

    headers: List[str] = []
    for column_index in range(ws.min_column, 20):
        header: str = ws.cell(row=header_row, column=column_index).value
        headers.append(header)

    nodes: dict[str, pd.DataFrame] = {}
    df:  Union[pd.DataFrame, None] = None
    year: int = 0
    reservoir_name: str | None = None
    for row in ws.iter_rows(min_row=2):
        column_index = 0
        for cell in row:
            if column_index < len(headers):
                header = headers[column_index]
                if header == 'STATE_NAME':
                    pass
                elif header == 'RESERVOIR_NAME':
                    reservoir_name = cell.value.lower()
                elif header == 'YEAR':
                    year = cell.value
                    df = get_reservoir_node(reservoir_name,  years, nodes)
                elif header == 'Total':
                    df.loc[df['Year'] == year, header] += int(cell.value)
                elif header is None:
                    pass
                else:
                    month = header
                    try:
                        if cell.value is not None:
                            df.loc[df['Year'] == year, month] += int(cell.value)
                        else:
                            pass
                    except KeyError:
                        print(header)
                    except TypeError:
                        print(header)
            else:
                pass
            column_index += 1

    for key, df in nodes.items():
        exclude_cols = ['Year']
        all_rows_are_zero = df.drop(columns=exclude_cols, errors='ignore').eq(0).all(axis=1).all()
        if not all_rows_are_zero:
            out_csv_path = out_path / f'{key}.csv'
            df.to_csv(out_csv_path, index=False, quoting=csv.QUOTE_NONE,  escapechar='\\', sep=' ')

def generate_cul_totals(path:Path):
    sheet.generate_cul_river_total(path, '', exclude='lake_mead', out_file='lc_reservor_cul.csv')