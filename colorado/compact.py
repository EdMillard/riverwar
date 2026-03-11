"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from openpyxl.worksheet.worksheet import Worksheet
from source import usbr_rise
from source.water_year_info import WaterYearInfo
from datetime import datetime
import colorado.lb as lb
import colorado.ub as ub
import colorado.allb as all_b
import pandas as pd
from sheet import sheet
from sheet.sheet import Sheet
from sheet.sheet import cl, cn

class Compact(Sheet):
    def __init__(self):
        headers = [ub.NATURAL_LEES_FERRY, lb.NATURAL_IMPERIAL,
                        lb.CU_NV, lb.CU_AZ, lb.CU_CA, lb.III_A_LB, lb.MEXICO, lb.MEAD_EVAPORATION, lb.HOOVER_RELEASE,
                        lb.H_M, lb.DIFF_7_5, lb.HOOVER_USGS, lb.DIAMOND_CREEK, lb.MEAD,
                        ub.POWELL,ub.POWELL_EVAPORATION, ub.GLEN_CANYON, ub.LEES_FERRY_USGS,
                        ub.INFLOW, ub.INFLOW_UNREGULATED,
                        ub.III_A_UB, ub.CU_CO, ub.CU_UT, ub.CU_WY,
                        ub.CU_NM, ub.CU_AZ,
                        lb.GC_INFLOW, lb.MEAD_ELEVATION, ub.POWELL_ELEVATION]
        super().__init__(headers)

    def load_df(self, df_compact : pd.DataFrame) -> None:
        usbr_lake_powell_elevation_af = 508
        sheet.usbr_last_value(self.df, usbr_lake_powell_elevation_af, 1965, self.end_year, title=ub.POWELL_ELEVATION, divisor=1)

        sheet.upper_basin_cu_from_excel(self.df)
        sheet.natural_flow_from_excel(self.df)
        sheet.lf_natural_flow_from_excel(self.df)

        Compact.lower_basin_annual_reports(self.df)

        self.df[lb.III_A_LB] = [f'=SUM(D{row}:F{row})' for row in range(2, len(self.df) + 2)]
        self.df[lb.III_A_LB] = self.df[lb.III_A_LB].astype(str)

        df_mead_evap = sheet.read_csv('data/Colorado_River/mead_evap.csv', sep=',')
        sheet.merge_annual_column(self.df, df_mead_evap, lb.MEAD_EVAPORATION, inp_column_name='Evaporation_AcreFeet')

        df_hoover = sheet.read_csv('data/USBR_Reports/releases/usbr_releases_hoover_dam.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df_hoover, lb.HOOVER_RELEASE)

        sheet.usgs_annuals(self.df, '09421500', self.start_year, self.end_year, title=lb.HOOVER_USGS)

        usbr_lake_mead_storage_af = 6124
        sheet.usbr_last_value(self.df, usbr_lake_mead_storage_af, self.start_year, self.end_year, title=lb.MEAD, month=10)

        usbr_lake_powell_storage_af = 509
        sheet.usbr_last_value(self.df, usbr_lake_powell_storage_af, self.start_year, self.end_year,  title=ub.POWELL)

        usbr_lake_mead_elevation_ft = 6123
        sheet.usbr_last_value(self.df, usbr_lake_mead_elevation_ft, self.start_year, self.end_year, title=lb.MEAD_ELEVATION, divisor=1)

        usbr_lake_powell_elevation_af = 508
        sheet.usbr_last_value(self.df, usbr_lake_powell_elevation_af, 1965, self.end_year, title=ub.POWELL_ELEVATION, divisor=1)

        sheet.usgs_annuals(self.df, '09404200', 2007, self.end_year,  title=lb.DIAMOND_CREEK, offset=1)
        sheet.usgs_annuals(self.df, '09380000', self.start_year, self.end_year, title=ub.LEES_FERRY_USGS)

        usbr_lake_powell_evap_af = 510
        sheet.usbr_annuals(self.df, usbr_lake_powell_evap_af, self.start_year, self.end_year,  title=ub.POWELL_EVAPORATION)

        usbr_lake_powell_release_total_af = 4354
        sheet.usbr_annuals(self.df, usbr_lake_powell_release_total_af, self.start_year, self.end_year, title=ub.GLEN_CANYON)

        usbr_lake_powell_regulated_inflow_af = 4288
        sheet.usbr_annuals(self.df, usbr_lake_powell_regulated_inflow_af, self.start_year, self.end_year, title=ub.INFLOW)

        usbr_lake_powell_unregulated_inflow_af = 4301
        sheet.usbr_annuals(self.df, usbr_lake_powell_unregulated_inflow_af, self.start_year, self.end_year, title=ub.INFLOW_UNREGULATED)

        self.df[lb.GC_INFLOW][43:len(self.df)] = [f'=N{row}-S{row}' for row in range(45, len(self.df)+2)]

    def build_sheet(self) -> None:
        ws: Worksheet = self.ws

        dest_col_h_m = cn(ws, lb.H_M)
        dest_col_diff = cn(ws, lb.DIFF_7_5)
        for row in range(2, len(self.df) + 2):   # adjust range
            formula = f"=J{row} - H{row}"
            ws.cell(row=row, column=dest_col_h_m).value = formula
            formula = f"=K{row}-7.5"
            ws.cell(row=row, column=dest_col_diff).value = formula

        self.set_bg(lb.MEXICO, color=all_b.USBR_AR_FLOW)
        self.set_bg(lb.HOOVER_RELEASE, color=all_b.USBR_AR_FLOW)

        self.set_column_negative_red(lb.DIFF_7_5)

        sheet.add_borders_to_column(ws, 1, 1, ws.max_row, which='vertical')
        sheet.add_borders_to_column(ws, 3, 1, ws.max_row, which='right')
        sheet.add_borders_to_column(ws, 7, 1, ws.max_row, which='vertical')

        self.set_bg(ub.NATURAL_LEES_FERRY, lb.NATURAL_IMPERIAL, color=all_b.USBR_NATURAL_BG)
        self.set_bg(lb.CU_NV, lb.III_A_LB, color=all_b.USBR_AR_CU_BG)

        # USGS
        self.set_bg(lb.HOOVER_USGS, lb.DIAMOND_CREEK, color=all_b.USGS_BG)

        lower_basin_end_col = 15
        # Reservoirs
        self.set_bg(lb.MEAD, color=all_b.LIGHT_BLUE_BG)
        self.set_bg(ub.POWELL, color=all_b.LIGHT_BLUE_BG)
        self.set_bg(ub.POWELL_EVAPORATION, color=all_b.EVAPORATION_BG)

        # sheet.add_borders_to_column(ws, lower_basin_end_col, 1, ws.max_row, which='right', border_style='medium')
        # sheet.add_borders_to_column(ws, 22, 1, ws.max_row, which='vertical')

        self.set_bg(lb.MEAD_EVAPORATION, color=all_b.EVAPORATION_BG)

        # USGS
        self.set_bg(ub.GLEN_CANYON_RELEASE, ub.INFLOW_UNREGULATED, color=all_b.USGS_BG)
        self.set_bg(lb.GC_INFLOW, color=all_b.USGS_BG)

        # Upper Basin CUL
        self.set_bg(ub.III_A_UB, ub.CU_AZ, color=all_b.USBR_UB_CUL_BG)

        # Elevations
        self.set_bg(lb.MEAD_ELEVATION, ub.POWELL_ELEVATION, color=all_b.USBR_RISE_ELEVATION_BG)

        self.format_header()

        current_str = datetime.now().strftime("%m/%d/%Y %I:%M%p")
        ws.cell(row=1, column=len(self.headers)+2).value = current_str

        ws.sheet_properties.fullCalcOnLoad = True

    @staticmethod
    def lower_basin_annual_reports(df: pd.DataFrame):
        df_ca = sheet.read_csv('data/USBR_Reports/ca/usbr_ca_total_consumptive_use.csv', sep='\s+')
        sheet.merge_annual_column(df, df_ca, lb.CU_CA)

        df_az = sheet.read_csv('data/USBR_Reports/az/usbr_az_total_consumptive_use.csv', sep='\s+')
        sheet.merge_annual_column(df, df_az, lb.CU_AZ)

        df_nv = sheet.read_csv('data/USBR_Reports/nv/usbr_nv_total_consumptive_use.csv', sep='\s+')
        sheet.merge_annual_column(df, df_nv,lb.CU_NV)

        df_mx = sheet.read_csv('data/USBR_Reports/mx/usbr_mx_satisfaction_of_treaty.csv', sep='\s+')
        sheet.merge_annual_column(df, df_mx, lb.MEXICO)

    @staticmethod
    def lake_powell(clicked_str):
        annuals_total = None
        # https://pubs.usgs.gov/sir/2022/5017/sir20225017.pdf
        # https://www.sciencebase.gov/catalog/item/614ccd07d34e0df5fb9868e2

        # 510: Daily Lake/Reservoir Evaporation - af (acre-feet per day; this is the primary evaporation loss series for Lake Powell)
        # 4301: Daily Lake/Reservoir Inflow - Unregulated - af (daily total unregulated inflow volume in acre-feet; key for natural supply calculations, forecasts, and storage change)
        # 512: Daily Lake/Reservoir Inflow - Unregulated - cfs (same as above but in cubic feet per second; often used interchangeably or converted to af)
        # 4288: Daily Lake/Reservoir Inflow - af (total regulated/observed inflow; contrasts with unregulated)
        # 511: Daily Lake/Reservoir Inflow - cfs (total inflow in cfs)
        # 4354: Daily Lake/Reservoir Release - Total - af (total dam releases, as we discussed earlier)
        # 507: Daily Lake/Reservoir Release - Power plant - cfs
        # 508: Daily Lake/Reservoir Elevation - ft (pool elevation)

        if clicked_str == 'Lake Powell Elevation':
            ts = pd.Timestamp('2025-11-01 00:00:00')
            water_year_info = WaterYearInfo.get_water_year(ts, month=10)
            usbr_lake_powell_elevation_af = 508
            usbr_rise.load(usbr_lake_powell_elevation_af, water_year_info=water_year_info)
        return annuals_total