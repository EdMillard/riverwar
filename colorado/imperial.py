"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
import colorado.lb as lb
import colorado.ub as ub
import colorado.allb as allb
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from sheet import sheet
from sheet.sheet import Sheet

class Imperial(Sheet):
    def __init__(self):
        super().__init__()
        self.start_year = 1964
        self.end_year = 2026

        self.headers = [lb.SALTON_ELEVATION, lb.SALTON_INFLOW, lb.ALAMO_RIVER, lb.NEW_RIVER, lb.WHITEWATER,
                            lb.IMPERIAL_TOTAL_CU,
                            lb.IMPERIAL_DIVERSION, lb.IMPERIAL_CU, lb.IMPERIAL_RETURN,
                            lb.COACHELLA_DIVERSION, lb.COACHELLA_CU, lb.COACHELLA_RETURN]
        self.df = sheet.create_df(self.start_year, self.end_year, self.headers)

    def load_df(self, df_main : pd.DataFrame):
        sheet.usgs_annuals(self.df, '10254730', self.start_year, self.end_year, title=lb.ALAMO_RIVER)          # 10254580 Alamo at border
        sheet.usgs_annuals(self.df, '10255550', self.start_year, self.end_year, title=lb.NEW_RIVER)    # 10254970 New at border
        sheet.usgs_annuals(self.df, '10259540', self.start_year, self.end_year, title=lb.WHITEWATER)

        self.df[lb.SALTON_INFLOW] = [f'=SUM(D{row}:F{row})' for row in range(2, len(self.df) + 2)]

        sheet.usgs_value(self.df, '10254005', 1988, self.end_year, title=lb.SALTON_ELEVATION, parameterCd='62614', statCd='00003')

        self.lower_basin_annual_reports(self.df)

    def build_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> Worksheet:

        ws, df_data = sheet.export_to_excel(self.df, writer, sheet_name)

        imperial_total_cu_col = sheet.get_column_number(ws, lb.IMPERIAL_TOTAL_CU)

        imperial_diversion_col = sheet.get_column_letter_insensitive(ws, lb.IMPERIAL_DIVERSION)
        imperial_cu_col = sheet.get_column_letter_insensitive(ws, lb.IMPERIAL_CU)
        imperial_return_col = sheet.get_column_number(ws, lb.IMPERIAL_RETURN)

        coachella_diversion_col = sheet.get_column_letter_insensitive(ws, lb.COACHELLA_DIVERSION)
        coachella_cu_col = sheet.get_column_letter_insensitive(ws, lb.COACHELLA_CU)
        coachella_return_col = sheet.get_column_number(ws, lb.COACHELLA_RETURN)
        for row in range(2, len(self.df) + 2):
            formula = f"={imperial_diversion_col}{row} - {imperial_cu_col}{row}"
            ws.cell(row=row, column=imperial_return_col).value = formula

            formula = f"={coachella_diversion_col}{row} - {coachella_cu_col}{row}"
            ws.cell(row=row, column=coachella_return_col).value = formula

            formula = f"={imperial_cu_col}{row} + {coachella_cu_col}{row}"
            ws.cell(row=row, column=imperial_total_cu_col).value = formula

        sheet.color_column(ws, 2, 2, ws.max_row, bg_color=allb.LIGHT_PURPLE_BG)

        for col in range(3, 7):
            sheet.color_column(ws, col, 2, ws.max_row, bg_color=allb.LIGHT_YELLOW_BG)

        sheet.color_column(ws, 7, 2, ws.max_row, bg_color=allb.LIGHT_RED_BG)
        sheet.color_column(ws, 9, 2, ws.max_row, bg_color=allb.LIGHT_RED_BG)
        sheet.color_column(ws, 12, 2, ws.max_row, bg_color=allb.LIGHT_RED_BG)

        sheet.format_header(ws, df_data)

        return ws

    @staticmethod
    def lower_basin_annual_reports(df: pd.DataFrame):
        df_len = len(df) + 2

        df_div = sheet.read_csv('data/USBR_Reports/ca/usbr_ca_imperial_irrigation_diversion.csv', sep='\s+')
        sheet.merge_annual_column(df, df_div, lb.IMPERIAL_DIVERSION)

        df_cu = sheet.read_csv('data/USBR_Reports/ca/usbr_ca_imperial_irrigation_consumptive_use.csv', sep='\s+')
        sheet.merge_annual_column(df, df_cu, lb.IMPERIAL_CU)

        df_div = sheet.read_csv('data/USBR_Reports/ca/usbr_ca_coachella_diversion.csv', sep='\s+')
        sheet.merge_annual_column(df, df_div, lb.COACHELLA_DIVERSION)

        df_cu = sheet.read_csv('data/USBR_Reports/ca/usbr_ca_coachella_consumptive_use.csv', sep='\s+')
        sheet.merge_annual_column(df, df_cu, lb.COACHELLA_CU)

