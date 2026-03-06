"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
import colorado.lb as lb
import colorado.ub as ub
import colorado.allb as allb
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from sheet import sheet
from sheet.sheet import Sheet

IIIC = "iii(c)"
EQUALS = " = "
MINUS = " - ("
PLUS = " + "
CLOSE = ")"
UPPER_CU = 'UPPER CU'
LOWER_CU = 'LOWER CU'

class III_C(Sheet):
    def __init__(self):
        super().__init__()
        self.start_year = 1964
        self.end_year = 2026

        self.headers = [IIIC, EQUALS, lb.BORDER_NATURAL, MINUS, LOWER_CU, PLUS, UPPER_CU, CLOSE, lb.MX_TREATY,
                              lb.HOOVER_RELEASE, ub.GLEN_CANYON_RELEASE]
        self.df = sheet.create_df(self.start_year, self.end_year, self.headers)

    def load_df(self, df_main : pd.DataFrame):
        df_len = len(self.df) + 2
        self.df[IIIC] = [f"=D{row}-( F{row} + H{row})" for row in range(2, df_len)]
        self.df[EQUALS] = [f"=" for _ in range(2, df_len)]

        sheet.natural_flow_from_excel(self.df)

        self.df.loc[57:, lb.BORDER_NATURAL] = df_main[ub.LEES_FERRY_NATURAL].values[57:]
        self.df[MINUS] = [f"- (" for _ in range(2, df_len)]
        self.df[LOWER_CU] = [f"='Colorado River'!G{row} + 'Colorado River'!I{row}" for row in range(2, df_len)]
        self.df[LOWER_CU] = self.df[LOWER_CU].astype(str)
        self.df[PLUS] = [f"+" for _ in range(2, df_len)]
        self.df[UPPER_CU] = [f"='Colorado River'!V{row}" for row in range(2, df_len)]
        self.df[UPPER_CU] = self.df[UPPER_CU].astype(str)

        self.df[lb.MX_TREATY] = [f"='Colorado River'!H{row}" for row in range(2, df_len)]
        #self.df[HOOVER_RELEASE] = [f"='Colorado River'!J{row}" for row in range(2, df_len)]

        df_hoover = sheet.read_csv('data/USBR_Reports/releases/usbr_releases_hoover_dam.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df_hoover, lb.HOOVER_RELEASE)

        df_glen_canyon = sheet.read_csv('data/USBR_Reports/releases/usbr_releases_glen_canyon_dam.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df_glen_canyon, ub.GLEN_CANYON_RELEASE)

        self.df[CLOSE] = [f")" for _ in range(2, df_len)]

    def build_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> Worksheet:
        ws, df_data = sheet.export_to_excel(self.df, writer, sheet_name)

        df_len = len(self.df) + 2
        for col in range(10, 12):
            sheet.color_column(ws, col, 2, ws.max_row, bg_color=lb.LOWER_BASIN_AR_FLOW)

        sheet.set_column_alignment(ws, 3, 2, df_len, horizontal='center')
        sheet.set_column_alignment(ws, 5, 2, df_len, horizontal='center')
        sheet.set_column_alignment(ws, 7, 2, df_len, horizontal='center')
        sheet.set_column_alignment(ws, 9, 2, df_len, horizontal='center')

        sheet.color_column(ws, 4, 2, ws.max_row, bg_color=allb.LIGHT_GREEN_BG)
        sheet.color_column(ws, 6, 2, ws.max_row, bg_color=allb.LIGHT_RED_BG)
        sheet.color_column(ws, 8, 2, ws.max_row, bg_color=allb.LIGHT_ORANGE_BG)

        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['E'].width = 3
        ws.column_dimensions['G'].width = 3

        sheet.set_column_negative_red(ws, 2, 2, df_len)
        sheet.format_header(ws, df_data)

        return ws