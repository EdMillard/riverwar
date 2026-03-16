"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from pathlib import Path
import colorado.lb as lb
import colorado.allb as all_b
import pandas as pd
from sheet.sheet import Sheet
import openpyxl
from sheet import sheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from typing import List, Dict
import csv

class LB_CUL(Sheet):
    def __init__(self):
        self.path:Path = Path('data/USBR_Lower_Colorado_CUL/Tributary')
        headers = [
            lb.LB_TRIBUTARY_CUL,
            lb.AZ_TRIBUTARY_CUL, lb.NV_TRIBUTARY_CUL, lb.UT_TRIBUTARY_CUL,  lb.NM_TRIBUTARY_CUL,

            lb.AZ_GILA_DOME_USGS,
            lb.GILA_CUL, lb.AZ_GILA_CUL, lb.NM_GILA_CUL,
            lb.AZ_LITTLE_COLORADO_CAMERON_USGS, lb.LITTLE_COLORADO_CUL, lb.AZ_LITTLE_COLORADO_CUL, lb.NM_LITTLE_COLORADO_CUL,
            lb.AZ_VIRGIN_LITTLEFIELD_USGS, lb.VIRGIN_CUL, lb.AZ_VIRGIN_CUL, lb.NV_VIRGIN_CUL, lb.UT_VIRGIN_CUL,
            lb.AZ_BILL_WILLIAMS_USGS, lb.AZ_BILL_WILLIAMS_CUL,
            lb.NV_MUDDY_MOAPA_USGS, lb.NV_MUDDY_CUL,
            lb.AZ_TRIB_BELOW_LAKE_MEAD_CUL, lb.NV_TRIB_ABOVE_LAKE_MEAD_CUL, lb.UT_TRIB_ABOVE_LAKE_MEAD_CUL
        ]
        super().__init__(headers, start_year=1971, end_year=2024)
        self.years: List[int] = list(range(self.start_year, self.end_year+1))

        # generate_cul_totals(self.path)
        # lower_basin_cu_from_excel(self.path, years)

    def load_df(self, df_compact : pd.DataFrame) -> None:
        divisor = 1
        path = self.path



        # AZ Gila
        sheet.usgs_annuals(self.df, '09520500', self.start_year, self.end_year, title=lb.AZ_GILA_DOME_USGS, divisor=1)
        df = sheet.read_csv(path / 'az_gila_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_CUL, divisor=divisor)
        # self.load_df_az_gila()

        # AZ Little Colorado
        # Little Colorado River Abv Mouth NR Desert View, AZ
        # sheet.usgs_annuals(self.df, '09402300', 1990, self.end_year, title=lb.AZ_LITTLE_COLORADO_MOUTH_USGS, divisor=1)

        # Little Colorado River Near Cameron, AZ
        sheet.usgs_annuals(self.df, '09402000', 1981, self.end_year, title=lb.AZ_LITTLE_COLORADO_CAMERON_USGS, divisor=1)
        df = sheet.read_csv(path / 'az_little_colorado_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_CUL, divisor=divisor)
        # self.load_df_az_little_colorado()

        # AZ Virgin
        #  USGS 09415000: Virgin River at Littlefield, AZ
        sheet.usgs_annuals(self.df, '09415000', self.start_year, self.end_year, title=lb.AZ_VIRGIN_LITTLEFIELD_USGS, divisor=1)
        df = sheet.read_csv(path / 'az_virgin_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_VIRGIN_CUL, divisor=divisor)
        # self.load_df_az_virgin()

        # AZ Bill Williams
        # USGS 09426620: Bill Williams River Near Parker, AZ
        sheet.usgs_annuals(self.df, '09426620', 1988, self.end_year, title=lb.AZ_BILL_WILLIAMS_USGS, divisor=1)
        df = sheet.read_csv(path / 'az_bill_williams_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_BILL_WILLIAMS_CUL, divisor=divisor)

        # AZ Trib
        df = sheet.read_csv(path / 'az_trib_above_lake_mead_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_TRIB_BELOW_LAKE_MEAD_CUL, divisor=divisor)

        # NV Virgin
        df = sheet.read_csv(path / 'nv_virgin_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_VIRGIN_CUL, divisor=divisor)

        # NV Muddy
        # USGS 09419515: Muddy River above Lake Mead near Overton, NV.
        # USGS 09416000 Muddy River near Moapa, NV.
        sheet.usgs_annuals(self.df, '09416000', self.start_year, self.end_year, title=lb.NV_MUDDY_MOAPA_USGS, divisor=1)
        df = sheet.read_csv(path / 'nv_muddy_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_MUDDY_CUL, divisor=divisor)

        # NV Trib
        df = sheet.read_csv(path / 'nv_trib_above_lake_mead_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_TRIB_ABOVE_LAKE_MEAD_CUL, divisor=divisor)

        # NM Gila
        df = sheet.read_csv(path / 'nm_gila_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NM_GILA_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'nm_little_colorado_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NM_LITTLE_COLORADO_CUL, divisor=divisor)

        # UT Virgin
        df = sheet.read_csv(path / 'ut_virgin_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.UT_VIRGIN_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'ut_trib_above_lake_mead_total_cu.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.UT_TRIB_ABOVE_LAKE_MEAD_CUL, divisor=divisor)

    @staticmethod
    def set_col_formula(ws:Worksheet, df:pd.DataFrame, formula:str, column_name:str, start_row=1) -> None:
        col_idx = df.columns.get_loc(column_name) + 1  # 1-based
        for i in range(start_row, len(df)+1):
            excel_row = i + 1
            f = formula.replace("[row]", str(excel_row))
            ws.cell(row=excel_row, column=col_idx, value=f)

    def build_sheet(self) -> None:
        ws: Worksheet = self.ws

        sheet.add_borders_to_column(ws, 1, 1, ws.max_row, end_col=ws.max_column, which='outer')

        sheet.formula_average(ws, self.df, self.years)
        sheet.add_borders_to_column(ws, 1, ws.max_row-1, ws.max_row-1, end_col=ws.max_column, which='outer')

        self.set_bg(lb.LB_TRIBUTARY_CUL, to=lb.NM_TRIBUTARY_CUL, color=all_b.USBR_LB_CUL_BG)

        # Sum LB Rivers
        columns = [lb.AZ_TRIBUTARY_CUL, lb.NV_TRIBUTARY_CUL, lb.UT_TRIBUTARY_CUL, lb.NM_TRIBUTARY_CUL]
        sheet.formula_add(ws, self.df, lb.LB_TRIBUTARY_CUL, columns)

        # Sum AZ Rivers
        columns = [lb.AZ_GILA_CUL, lb.AZ_LITTLE_COLORADO_CUL, lb.AZ_VIRGIN_CUL, lb.AZ_BILL_WILLIAMS_CUL, lb.AZ_TRIB_BELOW_LAKE_MEAD_CUL]
        sheet.formula_add(ws, self.df, lb.AZ_TRIBUTARY_CUL, columns)

        # Sum NV Rivers
        columns = [lb.NV_VIRGIN_CUL, lb.NV_MUDDY_CUL, lb.NV_TRIB_ABOVE_LAKE_MEAD_CUL]
        sheet.formula_add(ws, self.df, lb.NV_TRIBUTARY_CUL, columns)

        # Sum UT Rivers
        columns = [lb.UT_VIRGIN_CUL, lb.UT_TRIB_ABOVE_LAKE_MEAD_CUL]
        sheet.formula_add(ws, self.df, lb.UT_TRIBUTARY_CUL, columns)

        # Sum NM Rivers
        columns = [lb.NM_GILA_CUL, lb.NM_LITTLE_COLORADO_CUL]
        sheet.formula_add(ws, self.df, lb.NM_TRIBUTARY_CUL, columns)

        # AZ Gila
        self.set_bg(lb.AZ_GILA_DOME_USGS, color=all_b.USGS_BG)
        self.set_bg(lb.GILA_CUL,  color=all_b.USBR_LB_CUL_BG)
        columns = [lb.AZ_GILA_CUL, lb.NM_GILA_CUL]
        sheet.formula_add(ws, self.df, lb.GILA_CUL, columns)
        self.set_bg(lb.AZ_GILA_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.AZ_GILA_CUL, lb.AZ_GILA_IRRIGATION_CUL, lb.AZ_GILA_WITHIN_SYSTEM_CUL)
        # self.set_bg(lb.AZ_GILA_IRRIGATION_CUL, lb.AZ_GILA_WITHIN_SYSTEM_CUL, color=all_b.USBR_LB_CULL_BG, end_row=last_row)

        # AZ Little Colorado
        self.set_bg(lb.AZ_LITTLE_COLORADO_CAMERON_USGS, color=all_b.USGS_BG)
        self.set_bg(lb.LITTLE_COLORADO_CUL,  color=all_b.USBR_LB_CUL_BG)
        columns = [lb.AZ_LITTLE_COLORADO_CUL, lb.NM_LITTLE_COLORADO_CUL]
        sheet.formula_add(ws, self.df, lb.LITTLE_COLORADO_CUL, columns)
        self.set_bg(lb.AZ_LITTLE_COLORADO_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.AZ_LITTLE_COLORADO_CUL, lb.AZ_LITTLE_COLORADO_IRRIGATION_CUL,
        #                  lb.AZ_LITTLE_COLORADO_WITHIN_SYSTEM_CUL)
        # self.set_bg(lb.AZ_LITTLE_COLORADO_IRRIGATION_CUL, lb.AZ_LITTLE_COLORADO_WITHIN_SYSTEM_CUL,
        #             color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # AZ Virgin
        self.set_bg(lb.AZ_VIRGIN_LITTLEFIELD_USGS, color=all_b.USGS_BG)
        self.set_bg(lb.VIRGIN_CUL,  color=all_b.USBR_LB_CUL_BG)
        columns = [lb.AZ_VIRGIN_CUL, lb.NV_VIRGIN_CUL, lb.UT_VIRGIN_CUL]
        sheet.formula_add(ws, self.df, lb.VIRGIN_CUL, columns)
        self.set_bg(lb.AZ_VIRGIN_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.AZ_VIRGIN_CUL, lb.AZ_VIRGIN_IRRIGATION_CUL, lb.AZ_VIRGIN_M_AND_I_CUL)
        # self.set_bg(lb.AZ_VIRGIN_IRRIGATION_CUL, lb.AZ_VIRGIN_M_AND_I_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # AZ Bill Williams
        self.set_bg(lb.AZ_BILL_WILLIAMS_USGS, color=all_b.USGS_BG)
        self.set_bg(lb.AZ_BILL_WILLIAMS_CUL,  color=all_b.USBR_LB_CUL_BG)

        # AZ Trib Below
        self.set_bg(lb.AZ_TRIB_BELOW_LAKE_MEAD_CUL,  color=all_b.USBR_LB_CUL_BG)
        # self.set_bg(lb.AZ_TRIB_BELOW_LAKE_MEAD_IRRIGATION_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # NV Trib Above
        self.set_bg(lb.NV_TRIB_ABOVE_LAKE_MEAD_CUL,  color=all_b.USBR_LB_CUL_BG)
        # self.set_bg(lb.NV_TRIB_ABOVE_LAKE_MEAD_M_AND_I_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # NV Virgin
        self.set_bg(lb.NV_VIRGIN_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.NV_VIRGIN_CUL, lb.NV_VIRGIN_IRRIGATION_CUL, lb.NV_VIRGIN_M_AND_I_CUL)
        # self.set_bg(lb.NV_VIRGIN_IRRIGATION_CUL, lb.NV_VIRGIN_M_AND_I_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # NV Muddy
        self.set_bg(lb.NV_MUDDY_MOAPA_USGS, color=all_b.USGS_BG)
        self.set_bg(lb.NV_MUDDY_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.NV_MUDDY_CUL, lb.NV_MUDDY_IRRIGATION_CUL, lb.NV_MUDDY_M_AND_I_CUL)
        # self.set_bg(lb.NV_MUDDY_IRRIGATION_CUL, lb.NV_MUDDY_M_AND_I_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # NM Gila
        self.set_bg(lb.NM_GILA_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.NM_GILA_CUL, lb.NM_GILA_IRRIGATION_CUL, lb.NM_GILA_M_AND_I_CUL)
        # self.set_bg(lb.NM_GILA_IRRIGATION_CUL, lb.NM_GILA_M_AND_I_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # NM Gila
        self.set_bg(lb.NM_LITTLE_COLORADO_CUL,  color=all_b.USBR_LB_CUL_BG)

        # UT Virgin
        self.set_bg(lb.UT_VIRGIN_CUL,  color=all_b.USBR_LB_CUL_BG)
        # sheet.formula_sum(ws, self.df, lb.UT_VIRGIN_IRRIGATION_CUL, lb.UT_VIRGIN_M_AND_I_CUL, lb.NM_GILA_M_AND_I_CUL)
        # self.set_bg(lb.UT_VIRGIN_CUL, lb.UT_VIRGIN_M_AND_I_CUL, color=all_b.USBR_LB_CUL_BG, end_row=last_row)

        # UT Trib Above
        self.set_bg(lb.UT_TRIB_ABOVE_LAKE_MEAD_CUL,  color=all_b.USBR_LB_CUL_BG)

        sheet.clear_range(ws, ws.max_row, ws.max_row, 1, ws.max_column)

        self.format_header()

        self.set_column_width(lb.LB_TRIBUTARY_CUL, 6, to=lb.UT_TRIB_ABOVE_LAKE_MEAD_CUL)

    def load_df_az_gila(self):
        divisor = 1
        path = self.path

        df = sheet.read_csv(path / 'az_gila_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_M_AND_I_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_mineral_resources.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_MINERALS_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_livestock.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_LIVESTOCK_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_stockpond.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_STOCK_POND_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_measured_reservoirs.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_RESERVOIR_MEASURED_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_unmeasured_reservoirs.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_RESERVOIR_UNMEASURED_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_tep.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_TEP_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_gila_within_system.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_GILA_WITHIN_SYSTEM_CUL, divisor=divisor)

    def load_df_az_little_colorado(self):
        divisor = 1
        path = self.path

        df = sheet.read_csv(path / 'az_little_colorado_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_M_AND_I_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_mineral_resources.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_MINERALS_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_livestock.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_LIVESTOCK_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_stockpond.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_STOCK_POND_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_measured_reservoirs.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_RESERVOIR_MEASURED_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_unmeasured_reservoirs.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_RESERVOIR_UNMEASURED_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_tep.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_TEP_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_little_colorado_within_system.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_LITTLE_COLORADO_WITHIN_SYSTEM_CUL, divisor=divisor)

    def load_df_az_virgin(self):
        divisor = 1
        path = self.path

        df = sheet.read_csv(path / 'az_virgin_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_VIRGIN_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'az_virgin_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.AZ_VIRGIN_M_AND_I_CUL, divisor=divisor)

    def load_df_nv_virgin(self):
        divisor = 1
        path = self.path
        df = sheet.read_csv(path / 'nv_virgin_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_VIRGIN_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'nv_virgin_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_VIRGIN_M_AND_I_CUL, divisor=divisor)

    def load_df_nv_muddy(self):
        divisor = 1
        path = self.path
        df = sheet.read_csv(path / 'nv_muddy_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_MUDDY_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'nv_muddy_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_MUDDY_M_AND_I_CUL, divisor=divisor)

    def load_df_nv_above_lake_mead(self):
        divisor = 1
        path = self.path
        df = sheet.read_csv(path / 'nv_trib_above_lake_mead_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NV_TRIB_ABOVE_LAKE_MEAD_M_AND_I_CUL, divisor=divisor)

    def load_df_nm_gila(self):
        divisor = 1
        path = self.path
        df = sheet.read_csv(path / 'nm_gila_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NM_GILA_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'nm_gila_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.NM_GILA_M_AND_I_CUL, divisor=divisor)

    def load_df_ut_trib(self):
        divisor = 1
        path = self.path
        df = sheet.read_csv(path / 'ut_virgin_irrigation.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.UT_VIRGIN_IRRIGATION_CUL, divisor=divisor)

        df = sheet.read_csv(path / 'ut_virgin_m_i_other.csv', sep='\s+')
        sheet.merge_annual_column(self.df, df, lb.UT_VIRGIN_M_AND_I_CUL, divisor=divisor)

states = {
    '04': 'az',
    '06': 'ca',
    '32': 'nv',
    '35': 'nm',
    '49': 'ut',
}

def get_node(node_code: str, area_reference: Dict[str, Dict], calc_type:str, years: List[int], nodes: Dict) -> pd.DataFrame:
    parts = node_code.split('_')
    if len(parts) != 3:
        print(f'Invalid node code: {node_code}')
        return pd.DataFrame()
    state_code = parts[0]
    # down_stream_node = parts[1]
    node: str = parts[2]
    area = area_reference.get(node, None)
    if area:
        area_name:str | None = area.get('REPORTING_AREA', None)
        if area_name:
            state = states.get(state_code, None)
            node_name = f"{state}_{area_name.lower()}_{calc_type}"
            df: Union[pd.DataFrame, None] = nodes.get(node_name, None)
            if df is None:
                df = sheet.create_month_year_df(years)
                nodes[node_name] = df
        else:
            print('Reporting area not found')
            return pd.DataFrame()
    else:
        print(f'node code not found: {area}')
        return pd.DataFrame()

    return df

def lower_basin_cu_from_excel(out_path: Path, years):
    wb: Workbook = openpyxl.load_workbook('data/Colorado_River/1971-2024 Lower Colorado River System CUL Data.xlsx', data_only=True)
    area_reference = sheet.worksheet_to_dict_of_dicts(wb['Area_Reference'], 1, 2, 'HU8_CODE')
    ws: Worksheet = wb['Tributary']

    header_row: int = 1

    sheet.ensure_directory(out_path)

    headers: List[str] = []
    for column_index in range(ws.min_column, 20):
        header: str = ws.cell(row=header_row, column=column_index).value
        headers.append(header)

    nodes: dict[str, pd.DataFrame] = {}
    df:  Union[pd.DataFrame, None] = None
    year: int = 0
    calc_type: str | None = None
    node_code: str | None = None
    for row in ws.iter_rows(min_row=2):
        column_index = 0
        for cell in row:
            if column_index < len(headers):
                header = headers[column_index]
                if header == 'STATE_CODE':
                    pass
                elif header == 'NODE_CODE':
                    node_code = cell.value
                elif header == 'SOURCE':
                    pass
                elif header == 'CALC_TYPE':
                    calc_type = cell.value.lower()
                elif header == 'Year':
                    year = cell.value
                    df = get_node(node_code, area_reference, calc_type, years, nodes)
                elif header == 'Total':
                    df.loc[df['Year'] == year, header] += int(cell.value)
                elif header is None:
                    pass
                else:
                    month = header
                    try:
                        df.loc[df['Year'] == year, month] += int(cell.value)
                    except KeyError:
                        print(header)
            else:
                pass
            column_index += 1

    for key, df in nodes.items():
        exclude_cols = ['Year']
        all_rows_are_zero = df.drop(columns=exclude_cols, errors='ignore').eq(0).all(axis=1).all()
        if not all_rows_are_zero:
            out_csv_path = out_path / f'{key}.csv'
            df.to_csv(out_csv_path, index=False, quoting=csv.QUOTE_NONE,  escapechar='\\', sep=' ')

import pandas as pd
from pathlib import Path
from typing import List, Union

def sum_csv_files_by_year(
    csv_files: List[Union[str, Path]],
    output_path: Union[str, Path],
    year_column: str = 'Year',
    encoding: str = 'utf-8'
) -> pd.DataFrame:
    """
    Sums numeric columns across CSV files, aligning rows by the 'Year' column (treated as int).
    Keeps 'Year' as a regular column in the output (does not sum it).
    """
    if not csv_files:
        raise ValueError("No CSV files provided")

    csv_paths = [Path(f) for f in csv_files]
    output_path = Path(output_path)

    # ── Load first file ───────────────────────────────────────────────────────
    df_total = pd.read_csv(
        csv_paths[0],
        encoding=encoding,
        # dtype=str,                   # read everything as string first → safer
        sep=r'\s+',
        low_memory=False
    )

    print(f"\nBase file: {csv_paths[0].name}")
    print("Columns found:", list(df_total.columns))
    print("First few rows:\n", df_total.head(3))

    # Try to find the year column (robust matching)
    year_col = None
    for col in df_total.columns:
        if str(col).strip().lower() in ['year', 'yr', year_column.lower()]:
            year_col = col
            break

    if year_col is None:
        raise KeyError(
            f"Could not find a 'Year' column in {csv_paths[0].name}.\n"
            f"Available columns: {list(df_total.columns)}"
        )

    # Convert year to integer (after stripping any weird whitespace)
    df_total[year_col] = pd.to_numeric(df_total[year_col], errors='coerce').astype('Int64')
    df_total = df_total.dropna(subset=[year_col])  # drop rows where year is invalid

    # Set as index (but we'll bring it back later)
    df_total = df_total.set_index(year_col).sort_index()

    print(f"Using '{year_col}' as year column. Unique years: {sorted(df_total.index.unique())}")

    # ── Process remaining files ───────────────────────────────────────────────
    for path in csv_paths[1:]:
        df = pd.read_csv(
            path,
            encoding=encoding,
            # dtype=str,
            sep=r'\s+',
            low_memory=False
        )

        print(f"\nProcessing: {path.name}")
        print("Columns:", list(df.columns))

        # Find year column in this file
        this_year_col = None
        for col in df.columns:
            if str(col).strip().lower() in ['year', 'yr', year_column.lower()]:
                this_year_col = col
                break

        if this_year_col is None:
            print(f"→ Skipping {path.name} — no 'Year' column found")
            continue

        df[this_year_col] = pd.to_numeric(df[this_year_col], errors='coerce').astype('Int64')
        df = df.dropna(subset=[this_year_col])
        df = df.set_index(this_year_col).sort_index()

        # Only add numeric columns that exist in both
        number_total_cols = df_total.select_dtypes(include='number')
        number_cols = df.select_dtypes(include='number')
        numeric_cols = number_cols.columns.intersection(
            number_total_cols.columns
        )

        if len(numeric_cols) == 0:
            print(f"→ No common numeric columns — skipping addition")
            continue

        # Align and add (missing years get 0)
        df_total[numeric_cols] = df_total[numeric_cols].add(
            df[numeric_cols].reindex(df_total.index, fill_value=0),
            fill_value=0
        )

        print(f"→ Added {len(numeric_cols)} numeric columns from {path.name}")

    # Bring Year back as column
    df_total = df_total.reset_index(names=year_column)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_total.to_csv(output_path, index=False, sep=' ', encoding=encoding)

    print(f"\nDone. Saved to: {output_path}")
    print(f"Final shape: {df_total.shape}")
    print(f"Years: {sorted(df_total[year_column].unique())}")

    return df_total

def generate_cul_totals(path:Path):
    generate_cul_river_total(path, 'az_gila')
    generate_cul_river_total(path, 'az_little_colorado')
    generate_cul_river_total(path, 'az_virgin')
    generate_cul_river_total(path, 'az_bill_williams')
    generate_cul_river_total(path, 'az_trib_above_lake_mead')

    generate_cul_river_total(path, 'nv_muddy')
    generate_cul_river_total(path, 'nv_trib_above_lake_mead')
    generate_cul_river_total(path, 'nv_virgin')

    generate_cul_river_total(path, 'nm_gila')
    generate_cul_river_total(path, 'nm_little_colorado')

    generate_cul_river_total(path, 'ut_trib_above_lake_mead')
    generate_cul_river_total(path, 'ut_virgin')

    generate_cul_river_total(path, '*gila_total_cu', out_file='gila_total_cu.csv')
    generate_cul_river_total(path, '*virgin_total_cu', out_file='virgin_total_cu.csv')
    generate_cul_river_total(path, '*little_colorado_total_cu', out_file='little_colorado_total_cu.csv')

    generate_cul_river_total(path, 'nv_*total_cu.csv', out_file='nv_tributary_total_cu.csv')
    generate_cul_river_total(path, 'ut_*total_cu.csv', out_file='ut_tributary_total_cu.csv')
    generate_cul_river_total(path, 'nm_*total_cu.csv', out_file='nm_tributary_total_cu.csv')
    generate_cul_river_total(path, 'az_*total_cu.csv', out_file='az_tributary_total_cu.csv')

    generate_cul_river_total(path, '*tributary_total_cu.csv', out_file='lb_tributary_total_cu.csv')

def generate_cul_river_total(path:Path, river:str, out_file:str | None=None):
    if out_file is None:
        out_path = path / f"{river}_total_cu.csv"
    else:
        out_path = path / out_file

    remove_file(out_path)
    files = find_files(Path(path), f"{river}*")
    sum_csv_files_by_year(files, out_path, year_column='Year')

def remove_file(file_path: str | Path) -> bool:
    """
    Safely delete a file at the given path.

    Returns:
        True  → file was successfully deleted
        False → file did not exist (no error raised)

    Raises:
        IsADirectoryError → if the path points to a directory
        PermissionError   → if access denied
        OSError           → other OS-level errors
    """
    path = Path(file_path)

    try:
        path.unlink(missing_ok=True)  # missing_ok=True → no error if file doesn't exist
        print(f"Removed: {path}")
        return True
    except IsADirectoryError:
        print(f"Error: {path} is a directory, not a file.")
        return False
    except PermissionError:
        print(f"Error: Permission denied for {path}")
        return False
    except OSError as e:
        print(f"Error removing {path}: {e}")
        return False

def find_files(directory: str | Path, pattern: str) -> list[str]:
    """
    Find files using pathlib.glob (Python 3.5+)
    """
    base_dir = Path(directory).resolve()
    return sorted([
        str(path)
        for path in base_dir.glob(pattern)
        if path.is_file()
    ])