"""
Copyright (c) 2025 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from data_sets.data_set import DataSet
from api.registry import Registry
from pathlib import Path
import openpyxl
from api import df_utils
import colorado.lb as lb
import colorado.ub as ub
import pandas as pd
from sheet.sheet import max_used_column, read_year_value_pairs
from typing import Optional

class NaturalFlowDataSet(DataSet):
    def __init__(self, name:str, month:int=10):
        super().__init__(name, month=month)
        self.df = self.from_csv('natural_flow')

    def load(self) -> Optional[pd.DataFrame]:
        start_year = 1906
        start_row = 4
        # self.start_year = 1964
        # self.start_row = 62
        end_year = 2024
        end_row = 122

        df: pd.DataFrame = df_utils.create_df(start_year, end_year, [ub.NATURAL_LEES_FERRY])

        NaturalFlowDataSet.lf_natural_flow_from_excel(df, start_row=start_row, wy='WY', column_name=ub.NATURAL_LEES_FERRY)
        df_utils.moving_average(df, ub.NATURAL_LEES_FERRY, 'Supply 10 yr avg')

        # start_year = 1906
        start_row = 7
        # self.start_year = 1964
        # self.start_row = 65
        # end_year = 2020
        end_row = 121

        NaturalFlowDataSet.natural_flow_from_excel(df, start_row=start_row, end_row=end_row, wy='WY')

        return df


    @staticmethod
    def lf_natural_flow_from_excel(df: pd.DataFrame, start_row:int=4, end_row=122, wy='WY', column_name:str=ub.NATURAL_LEES_FERRY):
        wb = openpyxl.load_workbook('data/Colorado_River/LFnatFlow1906-2024.2024.9.12.xlsx', data_only=True)
        if wy == 'WY':
            ws = wb['Water Year']
        else:
            ws = wb['Calendar Year']
        # ws = wb['AnnualCYTotalNaturalFlow']
        # ws = wb['TotalNaturalFlow']   # Monthly
        data_start_row = start_row
        data_end_row = end_row
        year_column_index = 1

        for column_index in range(ws.min_column, max_used_column(ws) + 1):
            # gage = ws.cell(row=gage_row, column=column_index).value
            # header = ws.cell(row=header_row, column=column_index).value
            # units = ws.cell(row=unit_row, column=column_index).value

            if column_index == 2: # Lees Ferry
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index,
                                                      data_start_row, data_end_row, divisor=1)
                df.loc[0: 0 + len(values) - 1, column_name] = values

    @staticmethod
    def natural_flow_from_excel(df:pd.DataFrame, start_row:int=7, end_row=121, wy:str='WY'):
        wb = openpyxl.load_workbook('data/Colorado_River/NaturalFlows1906-2020_20221215.xlsx', data_only=True)
        if wy == 'WY':
            ws = wb['AnnualWYTotalNaturalFlow']
        else:
            ws = wb['AnnualCYTotalNaturalFlow']
        # ws = wb['TotalNaturalFlow']   # Monthly
        gage_row = 3
        unit_row = 6
        data_start_row = start_row
        data_end_row = end_row
        year_column_index = 3
        gage_to_name = {}
        name_to_gage = {}
        for column_index in range(ws.min_column,  max_used_column(ws) + 1):
            gage = ws.cell(row=gage_row, column=column_index).value
            gage_name =  ws.cell(row=gage_row+1, column=column_index).value
            # header = ws.cell(row=header_row, column=column_index).value
            units = ws.cell(row=unit_row, column=column_index).value

            if units == 'Water Year':
                year_column_index = column_index

            if gage is None:
                pass
            elif gage == 'Corresponding USGS gauge number':
                pass
            else:
                name_to_gage[gage_name] = gage
                gage_to_name[gage] = gage_name
                if gage == '09429490': # Imperial
                    pairs, values = read_year_value_pairs(ws, year_column_index, column_index,
                                                          data_start_row, data_end_row, divisor=1)
                    df.loc[0: 0 + len(values) - 1, lb.NATURAL_IMPERIAL] = values
                else:
                    pairs, values = read_year_value_pairs(ws, year_column_index, column_index,
                                                          data_start_row, data_end_row, divisor=1)
                    df.loc[0: 0 + len(values) - 1, gage_name] = values

        NaturalFlowDataSet.print_as_dict(gage_to_name)
        NaturalFlowDataSet.print_as_dict(name_to_gage)


    @staticmethod
    def print_as_dict(d: dict, indent: int = 4):
        """Print dictionary in valid Python syntax with nice formatting"""
        import pprint
        pp = pprint.PrettyPrinter(
            indent=indent,
            width=100,
            sort_dicts=False,
            compact=False
        )
        print(pp.pformat(d))