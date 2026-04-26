"""
Copyright (c) 2025 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from data_sets.data_set import DataSet
import openpyxl
from api import df_utils
import colorado.ub as ub
import pandas as pd
from sheet.sheet import max_used_column, read_year_value_pairs
from typing import Optional

class UpperBasinCULDataSet(DataSet):
    def __init__(self, name:str, month:int=10):
        super().__init__(name, month=month)
        self.df = self.from_csv(name)

    def load(self) -> Optional[pd.DataFrame]:
        start_year = 1971
        end_year = 2024

        df: pd.DataFrame = df_utils.create_df(start_year, end_year, [])

        UpperBasinCULDataSet.upper_basin_cul_from_excel(df, row_offset=0, divisor=1)

        df_utils.add_column_sum(df,
                                [ub.POWELL_EVAPORATION, ub.FLAMING_GORGE_EVAPORATION_WY,
                                 ub.BLUE_MESA_EVAPORATION_WY, ub.MORROW_EVAPORATION_WY],
                                ub.UB_RESERVOIR_EVAP)
        return df

    @staticmethod
    def upper_basin_cul_from_excel(df: pd.DataFrame, row_offset: int = 7, path: str = '', divisor=1_000_000):
        if not path:
            # path = 'data/Colorado_River/v24.5_UB_CU_WY_Annual.xlsx'
            path = 'data/Colorado_River/v24.5_CUL_ResultsCU_CY.xlsx'
        wb = openpyxl.load_workbook(path, data_only=True)
        # ws = wb['WY_Pivot']
        ws = wb['CY Pivot']
        header_row = 2
        unit_row = 3
        data_start_row = 4  # 1971
        data_end_row = 57  # 2024, 2025 is partial not usable
        year_column_index = 1
        for column_index in range(ws.min_column, max_used_column(ws) + 1):
            header = ws.cell(row=header_row, column=column_index).value
            units = ws.cell(row=unit_row, column=column_index).value

            if units == 'Calendar Year':
                year_column_index = column_index

            if header == 'Total Result' or header == 'Grand Total':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.UB_TOTAL] = values
            if header == 'Colorado':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.CU_CO] = values
            elif header == 'Utah':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.CU_UT] = values
            elif header == 'Wyoming':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.CU_WY] = values
            elif header == 'NewMexico':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.CU_NM] = values
            elif header == 'Arizona':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.AZ_CU] = values
            elif header == 'Lake Powell':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.POWELL_EVAPORATION] = values
            elif header == 'Flaming Gorge':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.FLAMING_GORGE_EVAPORATION_WY] = values
            elif header == 'Blue Mesa':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.BLUE_MESA_EVAPORATION_WY] = values
            elif header == 'Navajo':  # Not in USBR spreadsheet for some reason
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.NAVAJO_EVAPORATION_WY] = values
            elif header == 'Morrow Point':
                pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row,
                                                      divisor=divisor)
                df.loc[row_offset: row_offset + len(values) - 1, ub.MORROW_EVAPORATION_WY] = values
            elif header is None:
                pass
            else:
                pass
