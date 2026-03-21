"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from pathlib import Path
from copy import copy
from datetime import date
from io import StringIO
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from source.usgs_gage import USGSGage, daily_to_water_year, usgs_csv_summary
from source.water_year_info import WaterYearInfo
from typing import Optional
from graph.water import WaterGraph
from source import usbr_rise
import numpy as np
from typing import List, Tuple, Any, Dict, Union
import colorado.lb as lb
import colorado.ub as ub
from abc import ABC, abstractmethod
import re
import numexpr as ne
import pandas as pd
from pycel import ExcelCompiler
import string

cn = lambda ws, name: get_column_number(ws, name)
cl = lambda ws, name: get_column_letter_insensitive(ws, name)

class Sheet(ABC):
    def __init__(self, name:str, headers: List[str], start_year:int=1964, end_year:int=2026):
        self.name = name
        self.start_parameters = 0
        self.end_parameters = 0
        self.start_year: int = start_year
        self.end_year: int = end_year
        self.headers: List[str] = headers
        self.df: pd.DataFrame = create_df(self.start_year, self.end_year, headers)
        self.ws: Optional[Worksheet] = None

    def export(self, writer: pd.ExcelWriter, df_main:pd.DataFrame, number_format:str='0.00') -> Worksheet:
        try:
            self.load_df(df_main)
        except Exception as e:
            print(f"export load_df {self.name} error: {e}")

        try:
            self.ws, self.df = export_to_excel(self.df, writer, self.name, number_format=number_format)
        except Exception as e:
            print(f"export load_df {self.name} error: {e}")

        try:
            self.build_sheet()
        except Exception as e:
            print(f"export build_sheet {self.name} error: {e}")

        return self.ws

    @abstractmethod
    def load_df(self, df_compact : pd.DataFrame) -> None:
        pass

    @abstractmethod
    def build_sheet(self) -> None:
        pass

    def set_bg(self, name:str, to:str=None, color:str="ffffff", start_row=2, end_row=None) -> None:
        name_col_num = cn(self.ws, name)
        if end_row is None:
            end_row = self.ws.max_row
        if name_col_num:
            if to is None:
                color_column(self.ws, name_col_num, start_row, end_row, bg_color=color)
            else:
                for col in range(name_col_num, cn(self.ws, to)+1):
                    color_column(self.ws, col, start_row, end_row, bg_color=color)
        else:
            print(f'set_bg name not in sheet: {name}')

    def set_column_width(self, name:str, width:int, to:str=None) -> None:
        name_col_num = cn(self.ws, name)
        if name_col_num:
            if to is None:
                self.ws.column_dimensions[cl(self.ws, name)].width = width
            else:
                for col in range(name_col_num, cn(self.ws, to)+1):
                    letter = get_column_letter(col)
                    self.ws.column_dimensions[letter].width = width
                    # color_column(self.ws, col, 2, self.ws.max_row, bg_color=color)

    def set_column_alignment(self, name:str, horizontal:str) -> None:
        set_column_alignment(self.ws, cn(self.ws, name), 2, len(self.df) + 2, horizontal=horizontal)

    def set_column_negative_red(self, name: str, negative_color: str='Red', positive_color: str='Color1') -> None:
        set_column_negative_red(self.ws, cn(self.ws, name), 2, len(self.df) + 2,
                                negative_color=negative_color, positive_color=positive_color)

    # Variable name ahd units Row centered
    #
    def format_header(self):
        start_row = 1
        end_row = 2
        start_col = 1
        end_col = len(self.df.columns) + 1
        self.ws.row_dimensions[1].height = 25
        for col in range(1, end_col):
            self.ws.column_dimensions[get_column_letter(col)].width = 5
            for row in range(start_row, end_row):
                cell = self.ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

        # Colorado.add_borders_to_column(ws, start_col, start_row, end_row-1, end_col=end_col-1, which='none')
        add_borders_to_column(self.ws, start_col, start_row, end_row - 1, end_col=end_col - 1, which='outer')

        self.set_column_width('Year', 4)

def read_csv(filename, sep='\s+', comment_char='#'):
    cleaned_lines = []
    with open(filename, 'r') as f:
        for line in f:
            # Remove everything from comment_char to end of line
            line = line.split(comment_char, 1)[0].rstrip()
            # Skip completely empty lines after cleaning
            if line.strip():
                cleaned_lines.append(line)

    # Now feed the cleaned text to pandas
    clean_text = '\n'.join(cleaned_lines)
    try:
        df = pd.read_csv(StringIO(clean_text), sep=sep, comment=comment_char)
    except Exception as e:
        print(f"Read csv {filename} error: {e}")
        df = pd.DataFrame()

    return df

sheets:List[Sheet] = []

def export_to_excel(df:pd.DataFrame, writer: pd.ExcelWriter, sheet_name:str, number_format:str='0.00'):
    df_data = pd.DataFrame(df)

    # df_data = MainFrame.insert_units_row(df_data, units)

    df_data.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    ws.freeze_panes = ws['B2']
    # Date column format
    #
    # for row in range(1, len(df_data) + 2):
    #    cell = ws.cell(row=row, column=1)
    #   cell.number_format = 'yyyy'

    format_sheet(ws, number_format=number_format)

    return ws, df_data

def max_used_column(ws: Worksheet):
    """
    Return the 1-based column index of the *right-most* cell that contains
    a value (int, float, str, datetime, bool, …).
    Formatted-but-empty cells are ignored.
    """
    max_col = 0
    # iter_rows() yields every cell that openpyxl knows about.
    # Empty cells are still present if they have a style, but .value is None.
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_col = max(max_col, cell.column)
    return max_col

def read_year_value_pairs(
        ws: Worksheet,
        year_col: int,
        value_col: int,
        start_row: int,
        end_row: int,
        divisor=1_000_000
) -> Tuple[List[Any], List[Any]]:
    """
    Reads year and value pairs from an openpyxl worksheet.

    Parameters:
    -----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet object to read from
    year_col : int
        Column index for the year (1 = A, 2 = B, etc.)
    value_col : int
        Column index for the value (1 = A, 2 = B, etc.)
    start_row : int
        First row to read (inclusive, 1-based)
    end_row : int
        Last row to read (inclusive, 1-based)

    Returns:
    --------
    List[Tuple[year, value]]
        List of (year, value) tuples
        - year will be int if possible, else str
        - value will be a float/int if numeric, else str

    Example:
    --------
    wb = load_workbook("data.xlsx")
    ws = wb["Sheet1"]
    pairs = read_year_value_pairs(ws, year_col=1, value_col=3, start_row=2, end_row=50)
    """
    pairs = []
    values = []

    for row in range(start_row, end_row + 1):
        # Get cells (1-based indexing in openpyxl)
        year_cell = ws.cell(row=row, column=year_col)
        value_cell = ws.cell(row=row, column=value_col)

        year = year_cell.value
        value = value_cell.value

        # Try to convert year to int (common for year columns)
        if year is not None:
            try:
                year = int(year)
            except (ValueError, TypeError):
                pass  # keep as is (str or None)

        # Try to convert value to float/int
        if value is not None:
            try:
                if isinstance(value, (int, float)):
                    pass  # already numeric
                elif str(value).replace(".", "").replace("-", "").isdigit():
                    value = float(value) if "." in str(value) else int(value)
                else:
                    value = str(value).strip()
            except (ValueError, TypeError):
                value = str(value).strip() if value is not None else None

        # Only append if we have a year (skip completely empty rows)
        if year is not None:
            pairs.append((year, value / divisor))
            values.append(value / divisor)

    return pairs, values

def create_df(min_year: int, max_year: int, headers: List[str]):
    years = list(range(min_year, max_year + 1))

    df = pd.DataFrame(index=range(len(years)), columns=['Year'] + headers)
    df['Year'] = years
    df.iloc[:, 1:] = pd.NA

    return df

def create_daily_df(
        start_date: date | str | tuple[int, int, int],
        end_date: date | str | tuple[int, int, int],
        headers: list[str],
        include_end_date: bool = True
) -> pd.DataFrame:
    """
    Creates a DataFrame with one row per day between start_date and end_date (inclusive by default),
    with a proper datetime64[ns] index (date only) and the requested columns filled with pd.NA.

    Parameters:
        start_date: date, 'YYYY-MM-DD' string, or (year, month, day) tuple
        end_date:   date, 'YYYY-MM-DD' string, or (year, month, day) tuple
        headers: list of column names (excluding the date)
        include_end_date: whether to include the end_date row (default True)

    Returns:
        DataFrame with daily date index
    """
    # Normalize inputs to datetime
    if isinstance(start_date, (tuple, list)):
        start_date = date(*start_date)
    if isinstance(end_date, (tuple, list)):
        end_date = date(*end_date)

    start = pd.to_datetime(start_date)
    end = pd.to_datetime(end_date)

    # Generate daily date range
    dates = pd.date_range(
        start=start,
        end=end,
        freq='D',
        inclusive='both' if include_end_date else 'left'
    )

    # Create empty DataFrame
    df = pd.DataFrame(
        index=dates,
        columns=headers,
        dtype='object'
    )

    df[:] = pd.NA

    # Optional: make index name nicer
    df.index.name = 'date'

    return df

def fill_df_from_structured_array(
    df: pd.DataFrame,
    arr: np.ndarray,
    value_column_name: str = None,
    date_column_name: str = None,
    method: str = 'setitem'   # or 'loc' / 'at'
) -> pd.DataFrame:
    """
    Fills values from a structured array [('dt', '<M8[s]'), ('val', '<f4')]
    into the DataFrame, matching on date (truncating time to date-only).

    Assumes:
    - If df.index is datetime-like → uses index for matching
    - If df has a date column → uses that column (specify date_column_name)

    Parameters:
        df: DataFrame with daily dates (index or column)
        arr: structured ndarray with fields 'dt' and 'val'
        value_column_name: optional - name of column in df to fill
                           (if None, assumes there's only one data column)
        date_column_name: required only if date is not the index
        method: 'setitem' (fastest), 'loc', or 'at'

    Returns:
        Modified DataFrame (in place)
    """
    if not len(arr):
        return df

    # Extract dates and values
    dates = arr['dt'].astype('datetime64[D]')   # truncate to day
    values = arr['val']

    # Determine where dates live
    if isinstance(df.index, (pd.DatetimeIndex, pd.RangeIndex)):
        date_source = 'index'
        df_dates = df.index.floor('D') if df.index.dtype == 'datetime64[ns]' else df.index
    elif date_column_name is not None and date_column_name in df.columns:
        date_source = 'column'
        df_dates = pd.to_datetime(df[date_column_name]).dt.floor('D')
    else:
        raise ValueError("Could not find date information. "
                         "Use date_column_name= or make sure index is datetime-like.")

    # Choose target column
    if value_column_name is None:
        data_cols = [c for c in df.columns if c not in (date_column_name or [])]
        if len(data_cols) != 1:
            raise ValueError("value_column_name required when >1 non-date column exists")
        value_column_name = data_cols[0]

    if value_column_name not in df.columns:
        raise ValueError(f"Column '{value_column_name}' not found in DataFrame")

    # Fastest method: vectorized boolean indexing + numpy assignment
    if method == 'setitem':
        for dt, val in zip(dates, values):
            mask = df_dates == dt
            if mask.any():
                df.loc[mask, value_column_name] = val

    elif method == 'loc':
        # Alternative (sometimes clearer)
        for dt, val in zip(dates, values):
            if date_source == 'index':
                if dt in df.index:
                    df.loc[dt, value_column_name] = val
            else:
                mask = df[date_column_name].dt.date == dt.astype('datetime64[D]').astype(date)
                if mask.any():
                    df.loc[mask, value_column_name] = val

    elif method == 'at':
        # Fastest for single-cell writes when index is unique
        if date_source != 'index':
            raise ValueError("'at' method only supported when date is index")
        for dt, val in zip(dates, values):
            if dt in df.index:
                df.at[dt, value_column_name] = val

    else:
        raise ValueError("method must be 'setitem', 'loc' or 'at'")

    return df

def lf_natural_flow_from_excel(df: pd.DataFrame):
    wb = openpyxl.load_workbook('data/Colorado_River/LFnatFlow1906-2024.2024.9.12.xlsx', data_only=True)
    ws = wb['Calendar Year']
    # ws = wb['AnnualCYTotalNaturalFlow']
    # ws = wb['TotalNaturalFlow']   # Monthly
    data_start_row = 62 # 1964
    data_end_row = 122  # 2020
    year_column_index = 1
    for column_index in range(ws.min_column, max_used_column(ws) + 1):
        # gage = ws.cell(row=gage_row, column=column_index).value
        # header = ws.cell(row=header_row, column=column_index).value
        # units = ws.cell(row=unit_row, column=column_index).value

        if column_index == 2: # Lees Ferry
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[0: 0 + len(values) - 1, ub.NATURAL_LEES_FERRY] = values

def natural_flow_from_excel(df:pd.DataFrame):
    wb = openpyxl.load_workbook('data/Colorado_River/NaturalFlows1906-2020_20221215.xlsx', data_only=True)
    ws = wb['AnnualCYTotalNaturalFlow']
    # ws = wb['AnnualCYTotalNaturalFlow']
    # ws = wb['TotalNaturalFlow']   # Monthly
    gage_row = 3
    unit_row = 6
    data_start_row = 65 # 1964
    data_end_row = 121  # 2020
    year_column_index = 3
    for column_index in range(ws.min_column,  max_used_column(ws) + 1):
        gage = ws.cell(row=gage_row, column=column_index).value
        # header = ws.cell(row=header_row, column=column_index).value
        units = ws.cell(row=unit_row, column=column_index).value

        if units == 'Water Year':
            year_column_index = column_index

        if gage == '09429490': # Imperial
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[0: 0 + len(values) - 1, lb.NATURAL_IMPERIAL] = values

def worksheet_to_dict_of_dicts(
        ws: Worksheet,
        header_row: int,
        data_start_row: int,
        key_name: str
) -> Dict[Any, Dict[str, Any]]:
    """
    Convert an openpyxl worksheet into a dictionary of dictionaries.

    The outer dict uses values from the column named `key_name` (in header_row)
    as keys. Each inner dict contains all other columns using header names as keys.

    Skips rows where the key cell is empty/None.

    Args:
        ws: openpyxl Worksheet object
        header_row: Row number (1-based) containing column headers
        data_start_row: First row of actual data (1-based)
        key_name: Exact header name to use as the primary key

    Returns:
        Dict where keys are values from the key column, values are dicts of {header: cell_value}

    Example:
        {
            "Arizona": {"Year": 2024, "Jan": 1200.5, "Feb": 1150.0, ...},
            "California": {"Year": 2024, "Jan": 4500.0, ...},
            ...
        }
    """
    if not isinstance(ws, Worksheet):
        raise TypeError("ws must be an openpyxl Worksheet object")

    # Get header row values (1-based row)
    headers = []
    header_cells = ws[header_row]
    for cell in header_cells:
        if cell.value is None:
            headers.append(f"Column{cell.column_letter}")  # fallback for empty headers
        else:
            # Convert to string and strip (clean common issues)
            headers.append(str(cell.value).strip())

    # Find the index (0-based) of our key column
    try:
        key_index = headers.index(key_name)
    except ValueError:
        raise ValueError(
            f"Key name '{key_name}' not found in headers on row {header_row}. "
            f"Available headers: {headers}"
        )

    result = {}

    # Process data rows
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_cells = ws[row_idx]
        if not row_cells:  # empty row
            continue

        # Get the key value
        key_cell = row_cells[key_index]
        key_value = key_cell.value

        # Skip rows where key is empty/None
        if key_value is None or (isinstance(key_value, str) and not key_value.strip()):
            continue

        # Make inner dict for this row
        row_dict = {}
        for col_idx, header in enumerate(headers):
            if header == key_name:
                continue  # skip the key itself in the inner dict (optional – remove if you want it included)
            cell = row_cells[col_idx]
            row_dict[header] = cell.value

        # Store using the key
        # If duplicate keys exist, last one wins (you can change to raise error or collect list)
        result[key_value] = row_dict

    return result

def ensure_directory(path: str | Path) -> Path:
    """Create directory (and parents) if it doesn't exist"""
    directory = Path(path)
    directory.mkdir(parents=True, exist_ok=True)
    return directory

def create_month_year_df(years: List[int]) -> pd.DataFrame:
    return pd.DataFrame({
        'Year': years,
        'January': 0,
        'February': 0,
        'March': 0,
        'April': 0,
        'May': 0,
        'June': 0,
        'July': 0,
        'August': 0,
        'September': 0,
        'October': 0,
        'November': 0,
        'December': 0,
        'Total': 0
    })

def upper_basin_cul_from_excel(df:pd.DataFrame):
    wb = openpyxl.load_workbook('data/Colorado_River/V24.5_CUL_ResultsCU_CY.xlsx', data_only=True)
    ws = wb['CY Pivot']
    header_row = 2
    unit_row = 3
    data_start_row = 4 # 1971
    data_end_row = 57  # 2024, 2025 is partial not usable
    year_column_index = 1
    for column_index in range(ws.min_column, max_used_column(ws) + 1):
        header = ws.cell(row=header_row, column=column_index).value
        units = ws.cell(row=unit_row, column=column_index).value

        if units == 'Calendar Year':
            year_column_index = column_index

        if header == 'Grand Total':
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[7: 7 + len(values) - 1, ub.III_A_UB] = values
        if header == 'Colorado':
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[7: 7 + len(values) - 1, ub.CU_CO] = values
        elif header == 'Utah':
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[7: 7 + len(values) - 1, ub.CU_UT] = values
        elif header == 'Wyoming':
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[7: 7 + len(values) - 1, ub.CU_WY] = values
        elif header == 'NewMexico':
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[7: 7 + len(values) - 1, ub.CU_NM] = values
        elif header == 'Arizona':
            pairs, values = read_year_value_pairs(ws, year_column_index, column_index, data_start_row, data_end_row)
            df.loc[7: 7 + len(values) - 1, ub.AZ_CU] = values


def usgs_annuals(df, gage_id, start_year, end_year, title='', parameter_cd='00060', stat_cd='00003',
                 month=1, offset=0, divisor=1_000_000):
    annuals = []
    values = []
    if month != 1:
        start_year -= 1
    for year in range(start_year, end_year+1):
        ts = pd.Timestamp(f'{year}-{month}-01 00:00:00')
        water_year_info = WaterYearInfo.get_water_year(ts, month=month)
        gage = USGSGage(gage_id, water_year_info)
        try:
            daily_af = gage.daily_discharge(water_year_info=water_year_info, alias=title, parameterCd=parameter_cd,
                                        statCd=stat_cd)
            annual_af = daily_to_water_year(daily_af)
            # result = (annual_af[0]['dt'], annual_af[0]['val'])
            # annuals.append(result)
            if len(annual_af) == 1:
                values.append(annual_af[0][1] / divisor)
                annuals.append(annual_af[0])
            elif annual_af is None:
                print(f'No years returned {gage_id} {year}')
            elif len(annual_af) == 0:
                print(f'usgs_annuals no years returned  {gage_id} {year}')
                values.append(0)
            else:
                print(f'usgs_annuals multiple years returned  {gage_id} {year} {annual_af}')
        except Exception as e:
            print(f"usgs_annuals {gage_id}  error: {e}")

    # if title:
    #    print(title)
    #    for annual in annuals:
    #        print(f'{annual[0]} {annual[1] / divisor:10.2f} ')

    if title:
        if len(values) != len(df[title]):
            insert_values_from_year(df, title, start_year, values, offset=offset)
        else:
            df[title] = values

    return values

def usgs_value(df, gage_id, start_year, end_year, title='', parameterCd='00060', statCd='00003', month=1):
    years = []
    annuals = []
    values = []
    if month != 1:
        start_year -= 1
    for year in range(start_year, end_year+1):
        ts = pd.Timestamp(f'{year}-{month}-01 00:00:00')
        water_year_info = WaterYearInfo.get_water_year(ts, month=month)
        gage = USGSGage(gage_id, water_year_info)
        daily_feet = gage.daily_discharge(water_year_info=water_year_info, alias=title, parameterCd=parameterCd,
                                        statCd=statCd)
        # total = daily_release_ft['val'].sum()
        feet = daily_feet[-1]
        years.append(year + 1)
        values.append(feet[1])
        annuals.append(feet)

    # if title:
    #     print(title)
    #    for annual in annuals:
    #       print(f'{annual[0]} {annual[1]:10.2f} ')

    if title:
        insert_values_from_year(df, title, start_year, values)

    return values

def usbr_get_last_value(gage_id, year, cfs_to_af=False, month=1)-> float:
    if month != 1:
        year -= 1
    ts = pd.Timestamp(f'{year}-{month}-01 00:00:00')
    water_year_info = WaterYearInfo.get_water_year(ts, month=month)
    if cfs_to_af:
        info, daily_cfs = usbr_rise.load(gage_id, water_year_info=water_year_info)
        daily_af = WaterGraph.convert_cfs_to_af_per_day(daily_cfs)
    else:
        info, daily_af = usbr_rise.load(gage_id, water_year_info=water_year_info)

    af = daily_af[-1]
    return af

def usbr_last_value(df, gage_id, start_year, end_year, title='', cfs_to_af=False, month=1, divisor=1_000_000):
    years = []
    annuals = []
    values = []
    if month != 1:
        start_year -= 1
    for year in range(start_year, end_year+1):
        af:float = usbr_get_last_value(gage_id, year, cfs_to_af=cfs_to_af, month=month)
        years.append(year + 1)
        values.append(af[1] / divisor)
        annuals.append(af)

    # if title:
    #    print(title)
    #    for annual in annuals:
    #        print(f'{annual[0]} {annual[1] / divisor:10.2f} ')

    if title:
        year = df['Year'][0]
        if np.isnan(year):
            df['Year'] = years
        # print(len(df[title]), len(values))
        if len(df[title]) == len(values):
            df[title] = values
        else:
            # print(f'usbr_last_value {title} {len(df[title])} {len(values)}')
            insert_values_from_year(df, title, start_year, values)
    return values

def usbr_annuals(df, gage_id, start_year, end_year, title='', cfs_to_af=False, month=1, divisor=1_000_000, offset=0):
    annuals = []
    values = []
    start_year = start_year
    if month != 1:
        start_year -= 1
    for year in range(start_year, end_year+1):
        ts = pd.Timestamp(f'{year}-{month}-01 00:00:00')
        water_year_info = WaterYearInfo.get_water_year(ts, month=month)
        if cfs_to_af:
            info, daily_cfs = usbr_rise.load(gage_id, water_year_info=water_year_info)
            daily_af = WaterGraph.convert_cfs_to_af_per_day(daily_cfs)
        else:
            info, daily_af = usbr_rise.load(gage_id, water_year_info=water_year_info)

        # total = daily_release_ft['val'].sum()
        try:
            annual_af = WaterGraph.daily_to_water_year(daily_af, water_year_month=month)
            if len(annual_af) == 1:
                values.append(annual_af[0][1] / divisor)
                annuals.append(annual_af[0])
            elif annual_af is None:
                print(f'usbr_annuals No years returned {gage_id} {year}')
            elif len(annual_af) == 0:
                print(f'usbr_annuals no years returned  {gage_id} {year}')
                values.append(0)
            else:
                print(f'Multiple years returned {gage_id} {year}')
        except Exception as e:
            print(f"usbr_annuals {gage_id}  error: {e}")

    # if title:
    #    print(title)
    #    for annual in annuals:
    #        print(f'{annual[0]} {annual[1] / divisor:10.2f} ')

    if title:
        if len(values) != len(df[title]):
            insert_values_from_year(df, title, start_year, values, offset=offset)
        else:
            df[title] = values

    return values

def format_sheet(ws: Worksheet, number_format:str='0.00'):
    # Set font for everything
    red_font = Font(name='Arial Narrow', size=10, color="700000")
    green_font = Font(name='Arial Narrow', size=10, color="007000")
    blue_font = Font(name='Arial Narrow', size=10, color="000070")
    black_font = Font(name='Arial Narrow', size=10, color="000000")  # RRGGBB hex
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = black_font
            # cell.fill = red_fill

    set_number_format(ws, 2, ws.max_row, 2, ws.max_column, number_format=number_format)

def set_number_format(ws: Worksheet, start_row:int, end_row:int, start_col:int, end_col:int, number_format:str='0.00'):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.number_format = number_format

def force_numeric(df:pd.DataFrame, numeric_cols_to_convert:List[str]):
    for col in numeric_cols_to_convert:
        if col in df.columns:
            # Coerce invalid values → NaN
            df[col] = pd.to_numeric(df[col], errors='coerce')

def set_col_formula(ws:Worksheet, df:pd.DataFrame, formula:str, column_name:str, start_row=1) -> None:
    col_idx = df.columns.get_loc(column_name) + 1  # 1-based
    for i in range(start_row, len(df)+1):
        excel_row = i + 1
        f = formula.replace("[row]", str(excel_row))
        ws.cell(row=excel_row, column=col_idx, value=f)

def formula_add(ws: Worksheet, df: pd.DataFrame, target_column: str, column_names: List[str]):
    df[target_column] = df[column_names].sum(axis=1)
    formula = f'='
    for i, column_name in enumerate(column_names):
        letter = cl(ws, column_name)
        if i == 0:
            formula += f'{letter}[row]'
        else:
            formula += f'+{letter}[row]'

    set_col_formula(ws, df, formula, target_column)


def replace_header_names_with_column_letters(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        formula: str,
        header_row: int = 1
) -> str:
    """
    Replaces header names in single quotes (e.g. 'III(a) Lower') with column letters (A, B, ...)
    by looking them up in the specified header row of the worksheet.

    Example input formula:
        "=('III(a) Lower' - 2.6) + ('III(a) Upper' - 4.4)"

    Returns something like:
        "=(C2 - 2.6) + (E2 - 4.4)"   (assuming row 2 is the data row)
    """
    # Create a mapping: header name → column letter
    header_to_col = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value is not None:
            # Convert to string and strip whitespace for safer matching
            header_name = str(cell_value).strip()
            if header_name:  # skip empty headers
                header_to_col[header_name] = get_column_letter(col_idx)

    # Find all header names inside single quotes
    # Pattern matches anything inside '…' (non-greedy)
    def replacer(match):
        name = match.group(1).strip()  # remove any extra spaces inside quotes
        col_letter = header_to_col.get(name)
        if col_letter is None:
            print(f"Warning: Header '{name}' not found in row {header_row}")
            return match.group(0)  # keep original if not found
        # Replace with just the column letter (Excel will understand C2, E2, etc. implicitly)
        new_string = col_letter + '[row]'
        return new_string

    # Replace 'Header Name' → C  (or whatever letter)
    # We use re.sub with a callback function
    modified = re.sub(r"'([^']+)'", replacer, formula)

    return modified

def formula(ws: Worksheet, df: pd.DataFrame, target_column: str, formula: str, start_row:int=1):
    modified = replace_header_names_with_column_letters(ws, formula)
    set_col_formula(ws, df, modified, target_column, start_row)

def formula_subtract(ws: Worksheet, df: pd.DataFrame, target_column: str, minuend: str, subtrahend: str, start_row:int=1):
    numeric_cols_to_convert = [minuend, subtrahend]
    force_numeric(df, numeric_cols_to_convert)
    df[target_column] = df.eval(f'`{minuend}` - `{subtrahend}`')

    formula = f'={cl(ws, minuend)}[row]-{cl(ws, subtrahend)}[row]'
    set_col_formula(ws, df, formula, target_column, start_row)

def formula_delta(ws: Worksheet, df: pd.DataFrame, column_name: str, delta_column: str, start_row:int=1):
    formula = f'={cl(ws, delta_column)}[row]-{cl(ws, delta_column)}[row-1]'
    col_idx = df.columns.get_loc(column_name) + 1  # 1-based
    for i in range(start_row+1, len(df)+1):
        excel_row = i + 1
        f = formula.replace("[row]", str(excel_row))
        f = f.replace("[row-1]", str(excel_row-1))
        ws.cell(row=excel_row, column=col_idx, value=f)

def formula_subtract_constant(ws: Worksheet, df: pd.DataFrame, target_column: str, minuend: str, subtrahend: str):
    numeric_cols_to_convert = [minuend]
    force_numeric(df, numeric_cols_to_convert)
    df[target_column] = df.eval(f'`{minuend}` - {subtrahend}')

    formula = f'={cl(ws, minuend)}[row]-{subtrahend}'
    set_col_formula(ws, df, formula, target_column)

def formula_sum(ws: Worksheet, df: pd.DataFrame, target_column: str, first_column: str, last_column:str):
    df[target_column] = df.loc[:, first_column:last_column].sum(axis=1)

    col1 = cl(ws, first_column)
    col2 = cl(ws, last_column)
    formula = f'=SUM({col1}[row]:{col2}[row])'

    set_col_formula(ws, df, formula, target_column)

def set_row_formula(
        ws: Worksheet,
        df: pd.DataFrame,
        formula_template: str,
        target_row: int,
        start_col: str | int | None = None,
        end_col: str | int | None = None,
        start_data_row: int = 2,
        header: str = ''
) -> None:
    """
    Puts the same formula pattern across a range of columns in one specific row.

    Parameters:
    - ws: openpyxl Worksheet
    - df: pandas DataFrame (used only to know column order/names)
    - formula_template: str with placeholder e.g. '=AVERAGE({col_letter}{start_data_row}:{col_letter}[row-1])'
    - target_row: the Excel row number where formulas should be written
    - start_col / end_col: optional column limits (can be named, letter or 1-based index)
    - start_data_row: usually 2 (first data row after header)

    Example:
        formula = '=AVERAGE(B{start}:B[row-1])'
        set_row_formula(ws, df, formula, target_row=15, start_data_row=2)
        → puts in row 15: =AVERAGE(B2:B14), =AVERAGE(C2:C14), etc.
    """
    from openpyxl.utils import get_column_letter, column_index_from_string

    # Determine column range
    all_cols = list(df.columns)
    n_cols = len(all_cols)

    # Resolve start_col and end_col to 0-based indices in df
    if start_col is None:
        start_idx = 0
    elif isinstance(start_col, str):
        if start_col in all_cols:
            start_idx = all_cols.index(start_col)
        else:
            start_idx = column_index_from_string(start_col) - 1
    else:
        start_idx = int(start_col) - 1  # assume 1-based → 0-based

    if end_col is None:
        end_idx = n_cols - 1
    elif isinstance(end_col, str):
        if end_col in all_cols:
            end_idx = all_cols.index(end_col)
        else:
            end_idx = column_index_from_string(end_col) - 1
    else:
        end_idx = int(end_col) - 1

    start_idx = max(0, start_idx)
    end_idx = min(n_cols - 1, end_idx)

    for col_idx in range(start_idx, end_idx + 1):  # 0-based in df
        col_letter = get_column_letter(col_idx + 1)  # 1-based for Excel
        excel_col = col_idx + 1

        # Replace placeholders
        f = formula_template
        f = f.replace("[row]", str(target_row))
        f = f.replace("{col_letter}", col_letter)
        f = f.replace("{col}", col_letter)
        f = f.replace("[col]", col_letter)

        # Common patterns for dynamic row ranges
        f = f.replace("{start}", str(start_data_row))
        f = f.replace("[start]", str(start_data_row))
        f = f.replace("{start_row}", str(start_data_row))
        f = f.replace("[start_row]", str(start_data_row))

        # Optional: [row-1], [row+1], etc.
        import re
        def offset_replacer(m):
            try:
                offset = int(m.group(1))
                return str(target_row + offset)
            except:
                return m.group(0)

        f = re.sub(r'\[row([+-]\d+)]', offset_replacer, f)

        # Write formula to cell
        ws.cell(row=target_row, column=excel_col, value=f)
    ws.cell(row=target_row, column=start_col-1, value=header)

def formula_average(ws: Worksheet, df:pd.DataFrame, years: List[int], number_format:str='#,##0;-#,##0'):
    target_row = len(years) + 2
    formula = '=AVERAGE({col_letter}2:{col_letter}[row-1])'
    set_row_formula(ws, df, formula, target_row=target_row, start_data_row=2,
                           start_col=2, end_col=ws.max_column, header='Avg')
    set_number_format(ws, target_row, target_row + 1, 2, ws.max_column, number_format=number_format)
    set_font(ws, start_row=target_row, end_row=target_row + 1, start_col=2, end_col=ws.max_column)
    set_font(ws, start_row=target_row, end_row=target_row + 1, start_col=1, end_col=1)

def set_font(
        ws: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        font_name: str = 'Arial Narrow',
        font_size: int = 10,
        font_color: str = '000000',  # black by default (RRGGBB hex without #)
        bg_color: str | None = None  # optional background fill (RRGGBB)
):
    """
    Apply font (and optional background fill) to a rectangular range of cells.

    Args:
        ws: openpyxl Worksheet
        start_row, end_row: 1-based row numbers (inclusive)
        start_col, end_col: 1-based column numbers (inclusive)
        font_name: e.g. 'Arial Narrow', 'Calibri'
        font_size: point size
        font_color: hex color string without # (e.g. 'FF0000' = red)
        bg_color: optional hex background color (e.g. 'FFFFCC'), or None
    """
    # Create font object
    font = Font(name=font_name, size=font_size, color=font_color)

    # Create fill object only if bg_color is provided
    fill = None
    if bg_color:
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

    # Iterate over the range using 1-based indices (openpyxl uses 1-based)
    for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=False  # we need cell objects, not values
    ):
        for cell in row:
            cell.font = font
            if fill is not None:
                cell.fill = fill

def insert_units_row(df, variable_name_to_units):
    variable_names = df.columns.tolist()
    units = []
    for variable_name in variable_names:
        unit = variable_name_to_units.get(variable_name)
        if unit:
            units.append(unit)
        else:
            units.append('')
    df.loc[-1] = units  # Add units row at index -1
    df.index = df.index + 1  # Shift index
    df = df.sort_index()  # Sort to move units row to position 0
    return df.reset_index(drop=True)


def color_column(
        ws: Worksheet,
        column: str | int,  # e.g. "C" or 3
        start_row: int,
        end_row: int,
        fg_color: str = "FFFFFF",  # foreground / fill color (hex without #)
        bg_color: str = None,  # optional separate background (usually same as fg)
        text_color: str = "000000",  # font color (usually black)
) -> None:
    """
    Apply fill (background) and font color to a specific column over a row range.

    Args:
        ws: openpyxl worksheet object
        column: column letter ("A", "C") or column number (1-based)
        start_row: first row to color (inclusive)
        end_row: last row to color (inclusive)
        fg_color: fill color hex (e.g. "FFFF00" = yellow) - without #
        bg_color: if you want different background vs foreground (rare)
        text_color: font color hex (e.g. "FF0000" = red)

    Example:
        color_column(ws, "D", 5, 25, fg_color="FFD966", text_color="000000")
    """
    # Normalize column to letter
    if isinstance(column, int):
        col_letter = get_column_letter(column)
    else:
        col_letter = column.upper()

    # Decide fill color (most people mean background when they say "background color")
    fill_color = bg_color if bg_color else fg_color

    # Create fill style
    fill = PatternFill(
        start_color=fill_color,
        end_color=fill_color,
        fill_type="solid"
    )

    # Create font style
    font = Font(name='Arial Narrow', size=10, color=text_color)

    # Apply to each cell in the range
    for row in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{row}"]
        cell.fill = fill
        cell.font = font

def merge_annual_column(df:pd.DataFrame, annual_df:pd.DataFrame,
                        column_name:str, inp_column_name='Total', divisor=1_000_000):
    if divisor != 1:
        try:
            annual_df[inp_column_name] = annual_df[inp_column_name] / divisor
        except Exception as e:
            print(f"merge_annual_column {column_name} {inp_column_name} error: {e}")
    df1 = df.merge(
        annual_df[['Year', inp_column_name]],
        on='Year',
        how='left'  # keeps all rows from df1, fills NaN where no match
    )
    try:
        df[column_name] = df1[inp_column_name].combine_first(df1[column_name])
    except Exception as e:
        print(f"merge_annual_column {inp_column_name} {column_name} error: {e}")

def insert_values_from_year(
        df: pd.DataFrame,
        target_column: str,
        start_year: int,
        values: list,
        offset = 0
) -> pd.DataFrame:
    """
    Inserts a list of values into an existing column of a DataFrame,
    starting from the row where 'Year' matches start_year and continuing
    downward for each subsequent value.

    Parameters:
    - df: pandas DataFrame with a 'Year' column
    - target_column: name of the existing column to update
    - start_year: integer year when insertion begins
    - values: list of values to insert (len(values) determines how many rows are updated)

    Returns:
    - The modified DataFrame (in-place changes are also applied)

    Raises:
    - ValueError if target_column doesn't exist or 'Year' column missing
    - LookupError if start_year not found
    """
    if 'Year' not in df.columns:
        raise ValueError("DataFrame must have a 'Year' column")
    if target_column not in df.columns:
        raise ValueError(f"Target column '{target_column}' does not exist in DataFrame")

    # Ensure Year is integer for reliable matching
    df['Year'] = pd.to_numeric(df['Year'], errors='coerce').astype('Int64')

    # Find the starting row index
    matching_rows = df[df['Year'] == start_year]
    if matching_rows.empty:
        raise LookupError(f"Year {start_year} not found in DataFrame")

    start_idx = matching_rows.index[0] + offset

    # Check if there are enough rows after start_idx
    required_rows = len(values)
    available_rows = len(df) - start_idx
    if required_rows > available_rows:
        print(f"Warning: Only {available_rows} rows available after year {start_year}. "
              f"Inserting {available_rows} values and truncating the rest.")
        values = values[:available_rows]

    # Insert the values
    df.loc[start_idx: start_idx + len(values) - 1, target_column] = values

    return df

def add_borders_to_column(
        ws,
        start_col: int,
        start_row: int,
        end_row: int,
        end_col: int = None,
        border_style: str = "thin",  # thin, medium, thick, dashed, dotted, double, etc.
        which: str = "all",  # "all", "outer", "horizontal", "vertical", "top", "bottom", "left", "right"
        color: str = "000000"  # hex color without #
) -> None:
    """
    Apply borders to every cell in a specified column over a range of rows.

    Args:
        ws: openpyxl worksheet object
        start_col: 1-based column number
        start_row: first row (inclusive)
        end_row: last row (inclusive)
        end_col: last col (inclusive)
        border_style: 'thin', 'medium', 'thick', 'dashed', 'dotted', 'double', ...
        which: which borders to apply
               - "all": full box around each cell
               - "outer": only top of first + bottom of last + left + right
               - "horizontal": top + bottom only
               - "vertical": left + right only
               - or single side: "top", "bottom", "left", "right"
        color: border color (hex without #)

    Examples:
        add_borders_to_column(ws, "D", 5, 25, border_style="medium", which="all")
        add_borders_to_column(ws, 2, 2, 100, "thin", "horizontal", "808080")  # gray horizontal lines
    """
    if end_col is None:
        end_col = start_col
    # Define side style
    side = Side(border_style=border_style, color=color)

    # Prepare border objects based on 'which'
    if which == "all":
        full_border = Border(left=side, right=side, top=side, bottom=side)
        cell_border = full_border
    elif which == "none":
        cell_border = Border(left=None, right=None, top=None, bottom=None)
    elif which == "outer":
        # Only outer perimeter of the whole range
        for col in range(start_col, end_col + 1):
            for row in range(start_row, end_row + 1):
                col_letter = get_column_letter(col)
                cell = ws[f"{col_letter}{row}"]
                border = Border()
                if row == start_row:
                    border.top = side
                if row == end_row:
                    border.bottom = side
                if col == start_col:
                    border.left = side
                if col == end_col:
                    border.right = side
                cell.border = border
        return  # special case — skip the loop below
    elif which == "horizontal":
        cell_border = Border(top=side, bottom=side)
    elif which == "vertical":
        cell_border = Border(left=side, right=side)
    elif which == "top":
        cell_border = Border(top=side)
    elif which == "bottom":
        cell_border = Border(bottom=side)
    elif which == "left":
        cell_border = Border(left=side)
    elif which == "right":
        cell_border = Border(right=side)
    else:
        raise ValueError(
            f"Invalid 'which' value: {which}. Use 'all', 'outer', 'horizontal', 'vertical', 'top', 'bottom', 'left', 'right'")

    # Apply to each cell in range
    for col in range(start_col, end_col + 1):
        for row in range(start_row, end_row + 1):
            col_letter = get_column_letter(col)
            cell = ws[f"{col_letter}{row}"]
            cell.border = cell_border

def set_column_alignment(
    ws: Worksheet,
    column: str | int,
    start_row: int,
    end_row: int,
    horizontal: str = 'center',
    vertical: str = 'center',
    wrap_text: bool = False,
    shrink_to_fit: bool = False,
    indent: int = 0
) -> None:
    """
    Set alignment for all cells in a given column between start_row and end_row (inclusive).

    Parameters:
    -----------
    ws : Worksheet
        The openpyxl worksheet to modify
    column : str or int
        Column letter ('A', 'B', ...) or column number (1, 2, ...)
    start_row : int
        First row to apply alignment to (1-based)
    end_row : int
        Last row to apply alignment to (1-based, inclusive)
    horizontal : str, optional
        'left', 'center', 'right', 'general', 'justify', 'distributed'
        Default: 'center'
    vertical : str, optional
        'top', 'center', 'bottom', 'justify', 'distributed'
        Default: 'center'
    wrap_text : bool, optional
        Whether to wrap text in the cells
        Default: False
    shrink_to_fit : bool, optional
        Whether text should shrink to fit the cell
        Default: False
    indent : int, optional
        Indent level (0 or more)
        Default: 0

    Example:
    --------
    set_column_alignment(ws, 'B', 5, 100, horizontal='right', vertical='center', wrap_text=True)
    set_column_alignment(ws, 3, 2, 50, horizontal='center', vertical='top')  # column C
    """
    # Normalize column to letter
    if isinstance(column, int):
        col_letter = get_column_letter(column)
    else:
        col_letter = str(column).upper()

    # Create the alignment object once
    align = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text,
        shrink_to_fit=shrink_to_fit,
        indent=indent
    )

    # Fastest way: iterate only over the specific column slice
    for row in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{row}"]
        cell.alignment = align

def set_column_negative_red(
        ws: Worksheet,
        column: str | int,
        start_row: int,
        end_row: int,
        base_format: str = '#,##0.00',
        use_parentheses: bool = False,
        negative_color: str = 'Red',
        positive_color: str = 'Color1'
) -> None:
    """
    Apply number format so negative values show in red **with a minus sign**.

    Options:
    - base_format:       Format for positive numbers (e.g. '#,##0.00', '0', '$#,##0.00')
    - use_parentheses:   If True → negatives appear as (123.45) in red (accounting style)
                         If False → negatives appear as -123.45 in red (default)
    """
    if isinstance(column, int):
        col_letter = get_column_letter(column)
    else:
        col_letter = str(column).upper().strip()

    if use_parentheses:
        # Accounting style: positives normal, negatives red + parentheses
        fmt = f"[{positive_color}]({base_format});[{negative_color}]({base_format})"
    else:
        # Standard: explicit minus sign in red
        fmt = f"[{positive_color}]{base_format};[{negative_color}]-{base_format}"

    for row in range(start_row, end_row + 1):
        cell = ws[f"{col_letter}{row}"]
        cell.number_format = fmt

def get_column_letter_insensitive(ws, header_text):
    target = header_text.lower().strip()
    for cell in ws[1]:
        if cell.value and str(cell.value).lower().strip() == target:
            return cell.column_letter
    return None

def get_column_number(
    ws,
    header: str,
    case_sensitive: bool = True
) -> Optional[int]:
    """
    Returns the 1-based column number (A=1) for the given header in row 1.
    Returns None if not found.
    """
    target = header if case_sensitive else header.lower()

    for col_idx, cell in enumerate(ws[1], start=1):
        cell_value = cell.value
        if cell_value is None:
            continue
        compare_value = cell_value if case_sensitive else str(cell_value).lower()
        if compare_value == target:
            return col_idx

    return None

def copy_worksheet_to_new_workbook(source_wb_path:Path, sheet_name:str, target_wb_path:Path=None):
    # Load source
    source_wb = load_workbook(source_wb_path, data_only=False)
    source_ws = source_wb[sheet_name]

    # Create or load target workbook
    if target_wb_path is None:
        target_wb = Workbook()
        target_ws = target_wb.active
        target_ws.title = sheet_name
    else:
        target_wb = load_workbook(target_wb_path)
        target_ws = target_wb.create_sheet(sheet_name)

    # ────────────────────────────────────────────────
    # Copy cells (value + style + hyperlink + comment)
    # ────────────────────────────────────────────────
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column)
            if cell.has_style:
                new_cell.font      = copy(cell.font)
                new_cell.border    = copy(cell.border)
                new_cell.fill      = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            new_cell.value     = cell.value
            new_cell.hyperlink = cell.hyperlink
            new_cell.comment   = cell.comment

    # ────────────────────────────────────────────────
    # Copy merged cells
    # ────────────────────────────────────────────────
    target_ws.merged_cells = copy(source_ws.merged_cells)

    # ────────────────────────────────────────────────
    # Copy column dimensions
    # ────────────────────────────────────────────────
    for col, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col].width = dim.width
        # You can also copy hidden, bestFit etc. if needed

    # ────────────────────────────────────────────────
    # Copy row dimensions
    # ────────────────────────────────────────────────
    for row, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row].height = dim.height
        target_ws.row_dimensions[row].hidden = dim.hidden

    # ────────────────────────────────────────────────
    # Copy sheet properties (title, tab color, page setup, etc.)
    # ────────────────────────────────────────────────
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins     = copy(source_ws.page_margins)
    target_ws.page_setup       = copy(source_ws.page_setup)
    target_ws.print_options    = copy(source_ws.print_options)
    target_ws.print_title_cols = copy(source_ws.print_title_cols)
    target_ws.print_title_rows = copy(source_ws.print_title_rows)
    # target_ws.sheet_view       = copy(source_ws.sheet_view)

    # Optional: freeze panes, header/footer, etc.
    if source_ws.sheet_view.selection:
        target_ws.sheet_view.selection = copy(source_ws.sheet_view.selection)

    # Save target workbook
    target_wb.save(target_wb_path or "new_workbook.xlsx")

def clear_range(ws: Worksheet, start_row: int, end_row: int,
                start_col: int, end_col: int,
                clear_contents: bool = True,
                clear_formatting: bool = True) -> None:
    """
    Clears a rectangular range of cells in an openpyxl worksheet.

    Parameters:
    - ws: the worksheet object
    - start_row, end_row: 1-based row indices (inclusive)
    - start_col, end_col: 1-based column indices (inclusive)
    - clear_contents: if True, removes values and formulas
    - clear_formatting: if True, resets fill (background color), font, alignment, border, number_format

    Example usage:
        clear_range(ws, start_row=5, end_row=20, start_col=2, end_col=10)
    """
    # Default "no style" objects
    no_fill = PatternFill(fill_type=None)
    default_font = Font()
    default_alignment = Alignment()
    default_border = Border()
    default_number_format = 'General'

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            if clear_contents:
                cell.value = None  # removes text, numbers, dates, formulas, hyperlinks, etc.

            if clear_formatting:
                # Reset the most common formatting properties
                cell.fill = no_fill
                cell.font = default_font
                cell.alignment = default_alignment
                cell.border = default_border
                cell.number_format = default_number_format

                # Optional: reset protection, comment, data validation, etc. if needed
                # cell.protection = Protection(locked=True, hidden=False)
                # cell.comment = None
                # (data validation requires ws.data_validations.remove(...))

    # Note: Conditional formatting, merged cells, tables, charts are NOT cleared by this function.
    # If you need to remove conditional formatting from the range, see:
    # ws.conditional_formatting = {}   # (nuclear option - removes ALL CF from sheet)
    # or more targeted removal via ws.conditional_formatting.remove('A1:Z99')

def sum_csv_files_by_year(
    csv_files: List[Union[str, Path]],
    output_path: Union[str, Path],
    year_column: str = 'Year',
    encoding: str = 'utf-8'
) -> pd.DataFrame:
    """
    Sums numeric columns across CSV files, aligning rows by the 'Year' column (treated as int).
    Keeps 'Year' as a regular column in the output (does not sum it).
    """
    if not csv_files:
        raise ValueError("No CSV files provided")

    csv_paths = [Path(f) for f in csv_files]
    output_path = Path(output_path)

    # ── Load first file ───────────────────────────────────────────────────────
    df_total = pd.read_csv(
        csv_paths[0],
        encoding=encoding,
        # dtype=str,                   # read everything as string first → safer
        sep=r'\s+',
        low_memory=False
    )

    print(f"\nBase file: {csv_paths[0].name}")
    print("Columns found:", list(df_total.columns))
    print("First few rows:\n", df_total.head(3))

    # Try to find the year column (robust matching)
    year_col = None
    for col in df_total.columns:
        if str(col).strip().lower() in ['year', 'yr', year_column.lower()]:
            year_col = col
            break

    if year_col is None:
        raise KeyError(
            f"Could not find a 'Year' column in {csv_paths[0].name}.\n"
            f"Available columns: {list(df_total.columns)}"
        )

    # Convert year to integer (after stripping any weird whitespace)
    df_total[year_col] = pd.to_numeric(df_total[year_col], errors='coerce').astype('Int64')
    df_total = df_total.dropna(subset=[year_col])  # drop rows where year is invalid

    # Set as index (but we'll bring it back later)
    df_total = df_total.set_index(year_col).sort_index()

    print(f"Using '{year_col}' as year column. Unique years: {sorted(df_total.index.unique())}")

    # ── Process remaining files ───────────────────────────────────────────────
    for path in csv_paths[1:]:
        df = pd.read_csv(
            path,
            encoding=encoding,
            # dtype=str,
            sep=r'\s+',
            low_memory=False
        )

        print(f"\nProcessing: {path.name}")
        print("Columns:", list(df.columns))

        # Find year column in this file
        this_year_col = None
        for col in df.columns:
            if str(col).strip().lower() in ['year', 'yr', year_column.lower()]:
                this_year_col = col
                break

        if this_year_col is None:
            print(f"→ Skipping {path.name} — no 'Year' column found")
            continue

        df[this_year_col] = pd.to_numeric(df[this_year_col], errors='coerce').astype('Int64')
        df = df.dropna(subset=[this_year_col])
        df = df.set_index(this_year_col).sort_index()

        # Only add numeric columns that exist in both
        number_total_cols = df_total.select_dtypes(include='number')
        number_cols = df.select_dtypes(include='number')
        numeric_cols = number_cols.columns.intersection(
            number_total_cols.columns
        )

        if len(numeric_cols) == 0:
            print(f"→ No common numeric columns — skipping addition")
            continue

        # Align and add (missing years get 0)
        df_total[numeric_cols] = df_total[numeric_cols].add(
            df[numeric_cols].reindex(df_total.index, fill_value=0),
            fill_value=0
        )

        print(f"→ Added {len(numeric_cols)} numeric columns from {path.name}")

    # Bring Year back as column
    df_total = df_total.reset_index(names=year_column)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_total.to_csv(output_path, index=False, sep=' ', encoding=encoding)

    print(f"\nDone. Saved to: {output_path}")
    print(f"Final shape: {df_total.shape}")
    print(f"Years: {sorted(df_total[year_column].unique())}")

    return df_total

def generate_cul_river_total(path:Path, river:str, out_file:str | None=None,
                             exclude:str | None=None):
    if out_file is None:
        out_path = path / f"{river}_total_cul.csv"
    else:
        out_path = path / out_file

    remove_file(out_path)
    files = find_files(path, f"{river}*")
    if exclude is not None:
        files = [x for x in files if exclude not in x.lower()]
    sum_csv_files_by_year(files, out_path, year_column='Year')

def remove_file(file_path: str | Path) -> bool:
    """
    Safely delete a file at the given path.

    Returns:
        True  → file was successfully deleted
        False → file did not exist (no error raised)

    Raises:
        IsADirectoryError → if the path points to a directory
        PermissionError   → if access denied
        OSError           → other OS-level errors
    """
    path = Path(file_path)

    try:
        path.unlink(missing_ok=True)  # missing_ok=True → no error if file doesn't exist
        print(f"Removed: {path}")
        return True
    except IsADirectoryError:
        print(f"Error: {path} is a directory, not a file.")
        return False
    except PermissionError:
        print(f"Error: Permission denied for {path}")
        return False
    except OSError as e:
        print(f"Error removing {path}: {e}")
        return False

def find_files(directory: str | Path, pattern: str) -> list[str]:
    """
    Find files using pathlib.glob (Python 3.5+)
    """
    base_dir = Path(directory).resolve()
    return sorted([
        str(path)
        for path in base_dir.glob(pattern)
        if path.is_file()
    ])

def evaluate_excel_formulas_on_dataframe(
    df: pd.DataFrame,
    formulas: dict[str, str],          # e.g. {'New Col': '=B2 * 1.1 + C2'}
    output_columns: list[str] | None = None,  # which new/updated columns to keep
    start_row: int = 2,                # data starts at Excel row 2 (row 1 = headers)
    inplace: bool = False
) -> pd.DataFrame:
    """
    Evaluate Excel formulas (A1 style) on a pandas DataFrame by mapping columns to letters.

    Parameters:
    - df: DataFrame with named columns (can have spaces, parentheses, etc.)
    - formulas: dict of {target_column_name: 'excel_formula'}
      Formula uses A1 notation where:
        - Column A = df.columns[0]
        - Column B = df.columns[1]
        - etc.
      Use relative row refs like B2, C2 (pycel will adjust per row)
    - output_columns: optional list of columns to return (new or existing)
    - start_row: Excel row where data starts (usually 2 if row 1 is headers)
    - inplace: if True, modify df directly; else return new copy

    Returns:
    - DataFrame with evaluated columns added/updated
    """
    df = df.copy() if not inplace else df

    # Map column names → Excel column letters (A, B, C, ..., AA, AB, ...)
    col_to_letter = {}
    for i, col_name in enumerate(df.columns):
        letter = ''
        n = i
        while n >= 0:
            letter = string.ascii_uppercase[n % 26] + letter
            n = n // 26 - 1
        col_to_letter[col_name] = letter

    # Reverse map: letter → original column name
    letter_to_col = {v: k for k, v in col_to_letter.items()}

    print("Column mapping:")
    for col, let in col_to_letter.items():
        print(f"  {col:20} → {let}")

    # Create a temporary in-memory Excel-like structure
    # pycel needs a filename, so we use a dummy one (it can work without real file)
    compiler = ExcelCompiler(filename=None)  # dummy

    # We'll evaluate row by row
    results = {target_col: [] for target_col in formulas}

    for row_idx, row in df.iterrows():
        excel_row = row_idx + start_row  # e.g. pandas row 0 → Excel row 2

        # Set cell values for this row
        for col_name, value in row.items():
            letter = col_to_letter[col_name]
            cell = f"{letter}{excel_row}"
            compiler.set_value(cell, value)

        # Evaluate each formula for this row
        for target_col, formula in formulas.items():
            # pycel expects formula like '=B2+C2'
            # We keep it as-is — it uses the current row context
            try:
                value = compiler.evaluate(formula)
            except Exception as e:
                print(f"Error evaluating {formula} at row {excel_row}: {e}")
                value = None  # or pd.NA, 0, etc.

            results[target_col].append(value)

    # Add results to DataFrame
    for target_col, values in results.items():
        df[target_col] = values

    if output_columns:
        keep_cols = [c for c in df.columns if c in output_columns]
        df = df[keep_cols]

    return df

def is_simple_numeric_expression(expr: str) -> bool:
    """Heuristic: true if expression looks safe & fast for numexpr"""
    # Very basic check — can be improved
    dangerous = r'\b(import|exec|eval|open|__|\.|\[|\]|\(|\))'
    if re.search(dangerous, expr):
        return False
    return True

def import_numpy():
    import numpy as np
    return np

formulas = {
    'Delta Sales': '`Sales Amount`.diff()',
    'Delta Profit': '`Gross Profit`.diff().fillna(0)',
    '2-period change': '`Revenue`.diff(periods=2)',
    'YoY Growth %': '(`Revenue`.diff() / `Revenue`.shift(1)).fillna(0) * 100',
    'Cumulative Sum': '`Sales Amount`.cumsum()',
}

def evaluate_formulas_with_column_names(
        df: pd.DataFrame,
        formulas: Dict[str, str],
        inplace: bool = False,
        engine: str = "numexpr",  # or "python" for more complex expressions
        allowed_names: Optional[Dict] = None
) -> pd.DataFrame:
    """
    Evaluate formulas that use column names directly (like Excel named ranges).

    Examples of supported formula styles:
        '= Sales * 1.1'
        '= Profit / Revenue * 100'
        '= IF(Region == "West", Amount * 0.9, Amount)'
        '= Sales + Cost.shift(1).fillna(0)'

    Parameters:
    -----------
    df : pd.DataFrame
        Input data
    formulas : dict
        {new_or_existing_column: formula_string}
        Formula should start with '=' (optional) and use column names directly
    inplace : bool
        Modify df in place or return new copy
    engine : str
        'numexpr' → faster for numeric/vectorized operations
        'python'  → slower but supports more complex expressions
    allowed_names : dict, optional
        Extra variables/functions allowed in eval (e.g. {'today': pd.Timestamp.now()})

    Returns:
    --------
    DataFrame with new/updated columns
    """
    df = df.copy() if not inplace else df

    # Clean formulas: remove leading '=' if present
    cleaned_formulas = {}
    for target_col, formula in formulas.items():
        formula = formula.strip()
        if formula.startswith('='):
            formula = formula[1:].strip()
        cleaned_formulas[target_col] = formula

    # Prepare evaluation context
    ctx = df.copy()  # columns become variables
    if allowed_names:
        ctx.update(allowed_names)

    # Add common pandas functions / aliases if needed
    ctx['pd'] = pd
    ctx['np'] = pd.np if hasattr(pd, 'np') else import_numpy()  # fallback
    ctx['diff'] = lambda x, periods=1: x.diff(periods=periods)
    ctx['cumsum'] = lambda x: x.cumsum()

    for target_col, expr in cleaned_formulas.items():
        try:
            if engine == "numexpr" and is_simple_numeric_expression(expr):
                # Fast path for numeric vectorized ops
                result = ne.evaluate(expr, local_dict=ctx)
            else:
                # General safe eval (more capable but slower)
                result = eval(expr, {"__builtins__": {}}, ctx)

            # Ensure result is Series with correct index
            if not isinstance(result, pd.Series):
                result = pd.Series(result, index=df.index)

            df[target_col] = result

        except Exception as e:
            raise ValueError(
                f"Error evaluating formula for '{target_col}':\n"
                f"  Formula: {expr}\n"
                f"  Error: {str(e)}"
            ) from e

    return df


import pandas as pd
from pathlib import Path
from typing import Union, Optional


def load_excel_sheet(
        file_path: Union[str, Path],
        sheet_name: Optional[Union[str, int]] = 0,
        header: Union[int, list[int], None] = 0,
        skiprows: Optional[Union[int, list[int]]] = None,
        usecols: Optional[Union[str, list, None]] = None,
        dtype: Optional[dict] = None,
        engine: str = "openpyxl",
        **kwargs
) -> pd.DataFrame:
    """
    Load a specific sheet from an Excel (.xlsx) file into a pandas DataFrame.

    Parameters:
    -----------
    file_path : str or Path
        Path to the .xlsx file

    sheet_name : str, int, or None, default 0
        Sheet to load:
        - str: sheet name (case-sensitive)
        - int: sheet index (0-based)
        - None: load all sheets (returns OrderedDict of DataFrames)

    header : int, list of int, or None, default 0
        Row(s) to use as column names

    skiprows : int or list-like, optional
        Number of rows to skip or rows to skip at the start

    usecols : str, list, or None, optional
        Columns to parse (e.g. "A:C", ["Name", "Value"], [0, 2, 3])

    dtype : dict, optional
        Data type for columns (e.g. {"ID": str, "Amount": float})

    engine : str, default "openpyxl"
        Engine to use ("openpyxl" recommended for .xlsx, "xlrd" for .xls)

    **kwargs : dict
        Additional arguments passed to pd.read_excel()

    Returns:
    --------
    pd.DataFrame (or OrderedDict if sheet_name=None)

    Raises:
    -------
    FileNotFoundError
    ValueError (if sheet not found, etc.)
    """
    # Convert to Path object for better handling
    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if not file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
        raise ValueError(f"Unsupported file extension: {file_path.suffix}. Expected .xlsx, .xls or .xlsm")

    try:
        df = pd.read_excel(
            io=str(file_path),
            sheet_name=sheet_name,
            header=header,
            skiprows=skiprows,
            usecols=usecols,
            dtype=dtype,
            engine=engine,
            **kwargs
        )

        # If sheet_name is None, pd.read_excel returns OrderedDict
        if sheet_name is None:
            print(f"Loaded {len(df)} sheets: {list(df.keys())}")

        return df

    except ValueError as e:
        if "No sheet named" in str(e) or "Worksheet" in str(e):
            # Try to give helpful info
            xl = pd.ExcelFile(file_path)
            available = xl.sheet_names
            raise ValueError(
                f"Sheet '{sheet_name}' not found.\n"
                f"Available sheets: {available}"
            ) from e
        raise


# ────────────────────────────────────────────────
# Example usage
# ────────────────────────────────────────────────

if __name__ == "__main__":
    # Basic usage - load by sheet name
    df = load_excel_sheet(
        "data/financials_2025.xlsx",
        sheet_name="Q4 Summary"
    )
    print(df.head())

    # Load by sheet index (third sheet)
    df_index = load_excel_sheet(
        "data/financials_2025.xlsx",
        sheet_name=2,  # 0-based index
        header=1,  # headers start on row 2
        skiprows=3  # skip first 3 rows
    )

    # Load specific columns and force types
    df_clean = load_excel_sheet(
        "data/water_levels.xlsx",
        sheet_name="Lake Mead",
        usecols="A:D",  # columns A to D
        dtype={"Date": str, "Elevation_ft": float, "Volume": float}
    )

    # Load ALL sheets at once
    all_sheets = load_excel_sheet(
        "data/report.xlsx",
        sheet_name=None
    )
    print(all_sheets.keys())  # dict_keys(['Sheet1', 'Summary', 'Raw Data'])
    summary_df = all_sheets["Summary"]