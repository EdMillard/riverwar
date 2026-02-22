"""
Copyright (c) 2025 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from datetime import datetime
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.formula import Tokenizer
import re
from source.water_year_info import WaterYearInfo


class InflowOutflowMonth:
    all_variable_names = {}
    def __init__(self):
        self.month = None
        self.month_num = 0
        self.num_days_in_month = 0
        self.water_year_info = None
        self.year = None
        self.variable_names = None
        self.variable_names_raw = None
        self.column_indexes = {}
        self.column_to_name = {}
        self.name_to_column = {}
        self.units = {}
        self.values = {}
        self.formulas = {}
        self.equations = {}
        self.comments = {}
        self.row_headers = None
        self.daily_row_range = None
        self.excel_log = None
        self.alias_log = None
        self.column_aliases = None
        self.unit_aliases = None

    @staticmethod
    def reset():
        InflowOutflowMonth.all_variable_names = {}

    @staticmethod
    def max_used_column(ws):
        """
        Return the 1-based column index of the *right-most* cell that contains
        a value (int, float, str, datetime, bool, …).
        Formatted-but-empty cells are ignored.
        """
        max_col = 0
        # iter_rows() yields every cell that openpyxl knows about.
        # Empty cells are still present if they have a style, but .value is None.
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_col = max(max_col, cell.column)
        return max_col

    def load_month_from_excel(self, wb, wb_f, logger, excel_log, alias_log, config, column_aliases, unit_aliases,
                              month_name=None, month_index=None):
        self.excel_log = excel_log
        self.alias_log = alias_log
        self.column_aliases = column_aliases
        self.unit_aliases = unit_aliases

        if month_name:
            ws = wb[month_name]
            ws_f = wb_f[month_name]
        elif month_index is not None:
            ws = wb.worksheets[month_index]
            ws_f = wb_f.worksheets[month_index]
        else:
            ws = wb.active
            ws_f = wb_f.active

        logger.log_message(f'  Load {ws.title}')
        excel_log.write(f'\t {ws.title}\n')
        excel_log.flush()
        alias_log.write(f'{ws.title}\n')
        alias_log.flush()

        self.water_year_info, self.month, self.year, self.month_num = InflowOutflowMonth.get_year_info(ws)
        if self.water_year_info is None:
            return False
        elif self.month == 'APR' or self.month == 'Apr':
            self.month = 'APRIL'

        if self.water_year_info.year >= 2010:
            variable_row_range = (5, 10)
        elif self.water_year_info.year > 2002:
            variable_row_range = (5, 8)
        else:
            variable_row_range = (4, 6)

        self.variable_names_raw = self.variable_names_from_columns(logger, ws, variable_row_range, apply_aliases=False)
        self.variable_names = self.variable_names_from_columns(logger, ws, variable_row_range, apply_aliases=True)

        sections = []
        for cell in ws[4]:
            sections.append(cell.value)

        self.row_headers = InflowOutflowMonth.get_column_values(ws, 1, range(1, ws.max_row + 1))
        year_to_date_row = self.row_headers.index('YTD')+1
        self.load_accounting_parameters(ws, config, self.water_year_info, year_to_date_row)

        self.daily_row_range = InflowOutflowMonth.days_of_month_row_range(self.row_headers)
        days_of_month = InflowOutflowMonth.get_column_values(ws, 1, self.daily_row_range)
        self.num_days_in_month = len(days_of_month)

        # use None in reservoir elevation to detect partial months
        column_index = self.column_indexes['RESERV ELEV']
        reservoir_elevation_feet = self.column_values(ws, 'RESERV ELEV', column_index)
        reservoir_elevation_feet = InflowOutflowMonth.remove_trailing_none(reservoir_elevation_feet)
        if not reservoir_elevation_feet:
            return False
        elif len(reservoir_elevation_feet) < self.num_days_in_month:
            logger.log_message(f'    Partial month: {self.month}')
            self.num_days_in_month = len(reservoir_elevation_feet)
            self.daily_row_range = range(self.daily_row_range[0] ,self.daily_row_range[0] + self.num_days_in_month)

        total_row_index_errors = 0
        for column_name in self.variable_names:
            column_index = self.column_indexes[column_name]
            self.values[column_name] = self.column_values(ws, column_name, column_index)
            self.formulas[column_name], self.equations[column_name], row_index_errors = self.column_formulas(ws_f,
                                                                                column_name, column_index)
            total_row_index_errors += row_index_errors
            self.comments[column_name] = self.column_comments(ws_f, column_name, column_index)

        if total_row_index_errors:
            logger.log_message(f'    Row Index Errors: {total_row_index_errors}')

        datetime_array = InflowOutflowMonth.make_datetime_array(self.month_num, self.year, self.num_days_in_month)
        self.values['DATE'] = datetime_array.tolist()
        excel_log.flush()
        return True

    @staticmethod
    def get_year_info(ws):
        month = None
        year = None
        month_num = None
        water_year_info = None
        month_header = []
        for cell in ws[2]:
            month_header.append(cell.value)
        if len(month_header) < 15:
            # probably not a month sheet
            return water_year_info, month, year, month_num

        parts = None
        for col in range(10, 15):
            if month_header[col]:
                parts = month_header[col].strip().split('  ')
                if parts and len(parts) == 1:
                    parts = month_header[col].strip().split(' ')
                break
        if not parts:
            print(f'Month header not found: {month_header}')
        if parts and len(parts) == 2:
            month = parts[0]
            year_string = InflowOutflowMonth.strip_nonnumeric(parts[1])
            if len(year_string) < 4:
                year_string = '19' + year_string
            year = int(year_string)

            # 1987-1988 sheets have hosed years
            if year == 1889:
                year = 1989
            elif year == 1888:
                year = 1988
            elif year == 1887:
                year = 1987
            elif year == 1886:
                year = 1986

            try:
                dt = datetime.strptime(f"{month} {year}", "%B %Y")
            except ValueError:
                dt = datetime.strptime(f"{month} {year}", "%b %Y")
            water_year_info = WaterYearInfo.get_water_year(dt)

            try:
                month_num = datetime.strptime(month, '%B').month
            except ValueError:
                month_num = datetime.strptime(month, '%b').month
        else:
            print(f'Month/Year invalid {month_header}\n')
            return water_year_info, month, year, month_num
        return water_year_info, month, year, month_num

    # FIXME - Hardwired row 8 default wont work before 2004
    def get_param_from_cell(self, ws, variable_name, config, row=8, alias=None)->str:
        value = None
        if variable_name in self.column_indexes:
            column_index = self.column_indexes[variable_name]
            if column_index:
                value = ws[row][column_index - 1].value
                if value:
                    if alias:
                        self.excel_log.write(f'\t\tParameter {variable_name} {alias} = {value}\n')
                        config[alias] = value
                    else:
                        self.excel_log.write(f'\t\tParameter {variable_name} = {value}\n')
                        config[variable_name] = value
                    print(f'\tget_param_from_cell {variable_name} found in row {row} {self.month} = {value}')
        # else:
        #     print(f'\tget_param_from_cell {variable_name} not found in row {row} {self.month}')
        return value

    def load_accounting_parameters(self, ws, config, water_year_info, year_to_date_row):
        if self.month == 'NOVEMBER':
            # Get Uct 31 previous water year values from year to date row
            max_used_column = InflowOutflowMonth.max_used_column(ws)
            for column_index in range(1, max_used_column - 1):
                value = ws[year_to_date_row][column_index].value
                if value:
                    variable_name = self.variable_names[column_index]
                    config[variable_name] = value
                    self.excel_log.write(f'\t\tParameter {variable_name} from \'oct-31\' = {value}\n')
        elif self.month == 'APRIL':
            # Fish pool resets in April
            self.get_param_from_cell(ws, "FISH LEFT", config, row=year_to_date_row, alias='fish_pool_left_apr1_af')
        elif self.month == 'JULY'  or self.month == 'AUGUST' or self.month == 'SEPTEMBER' or self.month == 'OCTOBER':
            if water_year_info.year == 2003:
                row = 7
            else:
                row = 8
            self.get_param_from_cell(ws, 'MVI PROJ', config, row=row)

        if water_year_info.year >= 2003:
            if water_year_info.year == 2003:
                row = 9
            else:
                row = 8
            # "U/S USERS EXCH" started in 2004, target in header probably 2012
            self.get_param_from_cell(ws, "U/S USERS EXCH", config, row=row)
        if water_year_info.year >= 2009:
            self.get_param_from_cell(ws, 'DIST CLASS B', config)
            # FIXME - Hardwired row 9 wont work before 2004
            self.get_param_from_cell(ws, 'UF&R LEASE WATER', config, row=9)

        if 'TO EX' in self.column_indexes:
            self.get_param_from_cell(ws, 'TO EX', config, row=9)

    @staticmethod
    def get_operand_indices(formula):
        # Create a tokenizer for the formula
        tokenizer = Tokenizer(formula)

        # Initialize variables to track position in the formula string
        current_pos = 0
        operand_indices = []

        # Remove the leading '=' if present
        formula_clean = formula.lstrip('=')
        if formula.startswith('='):
            current_pos = 1  # Adjust for the '='

        # Iterate through the tokens
        for token in tokenizer.items:
            # Get the token's value (the substring in the formula)
            token_value = token.value

            # Find the token in the formula starting from current_pos
            # We use the formula_clean to search, but adjust indices for original formula
            start_pos = formula_clean.find(token_value, current_pos - (1 if formula.startswith('=') else 0))
            if start_pos == -1:
                print(f"Warning: Token '{token_value}' not found in formula from position {current_pos}")
                continue

            # Adjust start_pos for the original formula (account for '=' if present)
            start_pos += 1 if formula.startswith('=') else 0
            end_pos = start_pos + len(token_value) - 1

            # Check if the token is an operand
            if token.type == 'OPERAND':
                operand_indices.append({
                    'value': token_value,
                    'stype': token.subtype,
                    'start': start_pos,
                    'end': end_pos
                })

            # Update current position to the end of the current token
            current_pos = end_pos + 1

        return operand_indices

    @staticmethod
    def excel_column_name_to_index(col_name):
        """
        Convert an Excel column name (e.g., 'A', 'AA', 'AB') to a zero-based column index.
        'A' -> 0, 'B' -> 1, ..., 'Z' -> 25, 'AA' -> 26, 'AB' -> 27, etc.
        """
        if not col_name or not col_name.isalpha() or not col_name.isupper():
            raise ValueError("Column name must be a non-empty string of uppercase letters")

        index = 0
        for char in col_name:
            # Shift left by multiplying by 26 (base-26 system)
            index *= 26
            # Add value of current letter (A=1, B=2, ..., Z=26)
            index += ord(char) - ord('A') + 1

        # Subtract 1 to convert from one-based to zero-based index
        return index - 1

    @staticmethod
    def excel_column_name_to_variable(column_name, names):
        column_index = InflowOutflowMonth.excel_column_name_to_index(column_name)
        if 0 <= column_index < len(names):
            name = names[column_index]
        else:
            print('\t\texcel_column_name_to_variable failed', column_name, column_index)
            name = column_name

        return name

    @staticmethod
    def abbreviate_month(month_name):
        """
        Converts a full month name to its three-letter abbreviation.

        Args:
            month_name (str): The full month name (e.g., 'JANUARY', 'January').

        Returns:
            str: The three-letter abbreviation (e.g., 'Jan') or empty string if invalid.
        """
        month_map = {
            'january': 'Jan', 'february': 'Feb', 'march': 'Mar', 'april': 'Apr',
            'may': 'May', 'june': 'Jun', 'july': 'Jul', 'august': 'Aug',
            'september': 'Sep', 'october': 'Oct', 'november': 'Nov', 'december': 'Dec'
        }
        # Convert input to lowercase for case-insensitive matching
        month_name = month_name.lower().strip()
        return month_map.get(month_name, '')

    @staticmethod
    def string_for_transform(variable, formula, col, cell_row, row, month, column_name, daily_row_range, file,
                             python_format=True, filter_first_of_month_equations=True):
        # transformed_value = (f'\'{variable}\'[{row}]')
        row_index_invalid = False
        row_offset = cell_row - int(row)

        if filter_first_of_month_equations:
            if row_offset == 2:
                if cell_row == daily_row_range[0]:
                    # Special case where formula on day 1 of month (cell row 15) is referencing ytd row 13, last value from previous month
                    # Causes scanner to trip on a false formula change first of every month on accumulator formulas
                    row_offset = 1
        if python_format:
            variable_str = f'x[\'{variable}\']'
        else:
            variable_str = f'\'{variable}\''
        if row_offset > 0:
            transformed_value = f'{variable_str}[day-{row_offset}]'
        elif row_offset < 0:
            day = int(cell_row) - daily_row_range[0] - 1
            month = InflowOutflowMonth.abbreviate_month(month)
            row_index_invalid = True
            file.write(f"\t\tRow index invalid \'{col}{row}\' in \'{column_name}\'[{month}-{day}]: {formula}\n")
            file.flush()
            transformed_value = f'{variable_str}[day+{-row_offset}]'
        else:
            transformed_value = f'{variable_str}[day]'
        return transformed_value, row_index_invalid


    @staticmethod
    def transform_operands(formula, operand_indices, names, month, column_name, cell_row, daily_row_range, file):
        row_index_errors = 0

        # Regular expression to split cell references (e.g., 'A1' -> ('A', '1'), 'C3:D4' -> ('C3', 'D4'))
        cell_ref_pattern = re.compile(r'([A-Z]+)(\d+)')
        # cell_ref_pattern = re.compile(r'(\$?[A-Z]+)(\$?\d+)(?::(\$?[A-Z]+)(\$?\d+))?')

        # Create a copy of the operand indices with transformed values
        transformed_operands = []
        for operand in operand_indices:
            transformed_operand = operand.copy()
            if operand['stype'] == 'RANGE':
                value = operand['value']
                # Handle ranges (e.g., 'C3:D4')
                if ':' in value:
                    start_cell, end_cell = value.split(':')
                    transformed_parts = []
                    for cell in [start_cell, end_cell]:
                        match = cell_ref_pattern.match(cell)
                        if match:
                            col, row = match.groups()
                            # transformed_parts.append(f'COL[{row}]')
                            variable = InflowOutflowMonth.excel_column_name_to_variable(col, names)
                            s, row_index_invalid = InflowOutflowMonth.string_for_transform(variable, formula, col,
                                                                cell_row, row, month, column_name, daily_row_range, file)
                            if row_index_invalid:
                                row_index_errors += 1
                            transformed_parts.append(s)
                        else:
                            # print('\t\t transform_operands failed:', cell)
                            transformed_parts.append(cell)  # Fallback if not a valid cell reference
                    transformed_value = ':'.join(transformed_parts)
                else:
                    # Single cell reference (e.g., 'A1')
                    match = cell_ref_pattern.match(value)
                    if match:
                        col, row = match.groups()
                        variable = InflowOutflowMonth.excel_column_name_to_variable(col, names)
                        '''
                        groups = match.groups()
                        if groups[2] is None:
                            col, row = groups[0].replace('$', ''), groups[1].replace('$', '')
                            variable = InflowOutflowMonth.excel_column_name_to_variable(col, names)
                        else:
                            col1, row1, col2, row2 = (
                                groups[0].replace('$', ''),
                                groups[1].replace('$', ''),
                                groups[2].replace('$', ''),
                                groups[3].replace('$', '')
                            )
                            variable1 = InflowOutflowMonth.excel_column_name_to_variable(col1, names)
                            variable2 = InflowOutflowMonth.excel_column_name_to_variable(col2, names)
                            variable = f'{variable1}:{variable2}'
                        '''
                        transformed_value, row_index_invalid = InflowOutflowMonth.string_for_transform(variable,
                                                                        formula, col, cell_row,
                                                                        row, month, column_name, daily_row_range, file)
                        if row_index_invalid:
                            row_index_errors += 1
                    else:
                        transformed_value = value  # Fallback if not a valid cell reference
                transformed_operand['transformed_value'] = transformed_value
            else:
                # Non-range operands (e.g., numbers, text) remain unchanged
                transformed_operand['transformed_value'] = operand['value']
            transformed_operands.append(transformed_operand)

        # Sort operands by end index in descending order to process from end to start
        transformed_operands.sort(key=lambda x: x['end'], reverse=True)

        # Start with the original formula
        transformed_formula = formula

        # Apply replacements from end to start
        for operand in transformed_operands:
            if operand['stype'] == 'RANGE':
                start, end = operand['start'], operand['end']
                transformed_formula = (
                        transformed_formula[:start] +
                        operand['transformed_value'] +
                        transformed_formula[end + 1:]
                )

        return transformed_formula, transformed_operands, row_index_errors

    def column_values(self, ws, column_name, column_index):
        if column_index >= 0:
            # if cell.data_type == 'f':
            #    return cell.value
            return [ws.cell(row=i, column=column_index).value for i in self.daily_row_range]
        else:
            print("Column not found: ", column_name)
            return []

    def column_formulas(self, ws, column_name, column_index):
        formulas = []
        equations = []
        total_row_index_errors = 0

        if column_index >= 0:
            for row in self.daily_row_range:
                cell = ws.cell(row=row, column=column_index)
                if cell.data_type == 'f':
                    formula = cell.value
                    formulas.append(formula)
                    operands = InflowOutflowMonth.get_operand_indices(formula)

                    equation, t_operands, row_index_errors = InflowOutflowMonth.transform_operands(formula,
                                                                    operands, self.variable_names, self.month, column_name,
                                                                    row, self.daily_row_range, self.excel_log)
                    total_row_index_errors += row_index_errors
                    equations.append(equation)
                else:
                    formulas.append(None)
                    equations.append(None)
        else:
            self.excel_log.fwrite(f'Column not found: {column_name}\n')
        is_all_none = all(x is None for x in formulas)
        self.excel_log.flush()
        if is_all_none:
            return None, None, total_row_index_errors
        else:
            return formulas, equations, total_row_index_errors

    def column_comments(self, ws, column_name, column_index, print_comments=False):
        comments = []

        if column_index >= 0:
            for row in self.daily_row_range:
                cell = ws.cell(row=row, column=column_index)
                if cell.comment is not None:
                    try:
                        if print_comments:
                            print("\tComment: ", cell.comment.text)
                        comments.append(cell.comment.text)
                    except AttributeError as e:
                        print(f"Error accessing comment text for cell: {e}")
                        comments.append(None)
                else:
                    comments.append(None)
        else:
            print("Column not found: ", column_name)
        is_all_none = all(x is None for x in comments)
        if is_all_none:
            return None
        else:
            return comments

    @staticmethod
    def get_variable_name(ws, variable_row_range, column_index, column_aliases,  alias_log,
                          variable_name_prev, units):
        variable_name = ''
        for row in range(*variable_row_range):
            value = ws.cell(row=row, column=column_index).value
            if isinstance(value, str):
                value = value.strip()
                variable_name += value
                variable_name += ' '
            elif len(variable_name):
                break

        variable_name = variable_name.strip()

        if variable_name == 'MCPHEE DAILY EVAP':
            if units == 'AF':
                variable_name = 'MCPHEE DAILY EVAP AF'
            elif units == 'IN':
                variable_name = 'MCPHEE DAILY EVAP IN'
            else:
                variable_name = 'MCPHEE DAILY EVAP ' + units
                print(f'{variable_name} unexpected units {units}')
        elif variable_name == 'PROJ':
            if variable_name_prev == 'CORTEZ NON - PROJ':
                variable_name = 'CORTEZ PROJ'
            else:
                variable_name = 'D/S SENIOR PROJ'

        if column_aliases is not None:
            if variable_name in column_aliases:
                alias = column_aliases[variable_name]
                alias_log.write(f'\t\'{variable_name}\': \'{alias}\'\n')
                variable_name = alias

        # strip month off DATE
        if variable_name.startswith('DATE'):
            variable_name = 'DATE'

        return variable_name

    def variable_names_from_columns(self, logger, ws, variable_row_range, apply_aliases=True, ignore_hidden=False):
        variable_names = []
        hidden_max = 0
        variable_name_prev = ''
        max_used_column = InflowOutflowMonth.max_used_column(ws)
        # print(f'Variable Names {self.month}')
        for column_index in range(ws.min_column, max_used_column+1):
            # Handle hidden columns
            column_letter = get_column_letter(column_index)
            col_dim = ws.column_dimensions[column_letter]

            # FIXME - Units appeared in 2009 in row 8, row 10 2010 and later
            units = ''
            if self.water_year_info.year >= 2009:
                if self.water_year_info.year == 2009:
                    if 7 <= self.month_num <= 10:
                        unit_row = 9
                        variable_row_range = (5, 9)
                    else:
                        unit_row = 8
                else:
                    unit_row = 10
                units = ws.cell(row=unit_row, column=column_index).value
                if units is None or not isinstance(units, str) or not len(units):
                    units = ws.cell(row=unit_row+1, column=column_index).value
                if units is not None and isinstance(units, str) and len(units):
                    units = units.strip()
                    if units == 'Outflow':
                        units = ws.cell(row=unit_row+1, column=column_index).value

            if apply_aliases:
                alias_table = self.column_aliases
            else:
                alias_table = None
            variable_name = InflowOutflowMonth.get_variable_name(ws, variable_row_range, column_index,
                                                                 alias_table, self.alias_log, variable_name_prev, units)
            variable_name = self.patch_variable_name(logger, ws, variable_name, units, column_index,  column_letter)

            variable_name_prev = variable_name

            # print(f'  {column_letter} variable: {variable_name} {units}')
            if apply_aliases:
                if not len(variable_name):
                    # print('\t\tNo column name', column_index, column_letter)
                    variable_name = column_letter


            if len(variable_name) > 0:
                if apply_aliases:
                    # This is used to track columns that don't appear in all months
                    if variable_name not in InflowOutflowMonth.all_variable_names:
                        InflowOutflowMonth.all_variable_names[variable_name] = True

                    if variable_name not in self.column_indexes:
                        self.column_indexes[variable_name] = column_index
                        # print(f"Added '{variable_name}' to dictionary")
                    else:
                        self.column_indexes[column_letter] = column_index
                        if variable_name != 'DATE':
                            logger.log_message(f"\tWarning: {variable_name} already in dictionary using {column_letter}")
                    self.column_to_name[column_letter] = variable_name
                    self.name_to_column[variable_name] = column_letter
                    if units is not None and isinstance(units, str) and len(units):
                        if units == 'UNIT':
                            units = 'DATE'
                        self.units[variable_name] = units
                        if units in self.unit_aliases:
                            units = self.unit_aliases[units]
                            # print(f"\tunit alias to: {units}")
                    else:
                        pass
                    if ignore_hidden:
                        if col_dim.hidden:
                            hidden_max = col_dim.max
                            continue
                        elif hidden_max and column_index > hidden_max:
                            hidden_max = 0
                        elif hidden_max:
                            continue
                if variable_name:
                    variable_names.append(variable_name)

        return variable_names

    def patch_variable_name(self, logger, ws, variable_name, units, column_index, column_letter):
        # Painful year specific variable name patches
        #
        if self.water_year_info.year == 2009:
            variable_name = InflowOutflowMonth.patch_variable_name_2009(logger, ws, variable_name, units, column_index,  column_letter)
        elif self.water_year_info.year == 2002:
            variable_name = InflowOutflowMonth.patch_variable_name_2002(logger, ws, variable_name, units, column_index,  column_letter)
        else:
            if not len(variable_name):
                if self.water_year_info.year <= 1989 and column_letter == 'AD':
                    variable_name = 'MVI CALL DAILY'
                # else:
                #    logger.log_message(f"\tWarning: no variable name column: {column_letter}")
        return variable_name

    @staticmethod
    def patch_variable_name_2002(logger, ws, variable_name, units, column_index, column_letter):
        if not variable_name and column_letter == 'AE':
            variable_name = 'MVI STORABLE'
        return variable_name

    @staticmethod
    def patch_variable_name_2009(logger, ws, variable_name, units, column_index, column_letter):
        if variable_name == 'TOTAL MVI IRRIG.'  or variable_name == 'MVI TOTAL IRRIG' or variable_name == 'TOTAL MVI IRRIG. N. PROJ':
            variable_name = 'MVI TOTAL IRRIG'
            if units == 'Ac. Ft.' or units == 'AC FT':
                variable_name += ' AF'

        if (variable_name == 'TOTAL MVI' or variable_name == 'TOTAl MVI'
                or variable_name == 'MVI TOTAL WATER USED' or variable_name == 'TOTAL MVI WATER USED'):
            variable_name = 'MVI TOTAL WATER USED'
            units2 = ws.cell(row=10, column=column_index).value
            if units == 'Ac. Ft.' or units == 'AC FT':
                variable_name += ' AF'
            elif units2 is not None and units2 == 'Ac. Ft.' or units2 == 'AC FT':
                variable_name += ' AF'

        if not variable_name and column_letter == 'AG':
            variable_name = 'D/S SENIOR PROJ'
        elif variable_name == 'D/S SENIOR':
            variable_name = 'D/S SENIOR MVIC'

        return variable_name

    @staticmethod
    def pop_after_last_match(list1, list2):
        # Convert list2 to a set for O(1) lookup
        set2 = set(list2)

        # Find the last index where any name from list2 appears in list1
        last_index = -1
        for i, name in enumerate(list1):
            if name in set2:
                last_index = i

        # If no match is found, return the original list (or empty list if preferred)
        if last_index == -1:
            print('\tpop_after_last_match no match found')
            return list1[:]

        # Return the list up to (and including) the last matched name
        return list1[:last_index + 1]

    @staticmethod
    def concatenate_month(year_val, month_to_append, num_days_in_month):
        for column_name in InflowOutflowMonth.all_variable_names:
                month_val = None
                if column_name in month_to_append:
                    month_val = month_to_append[column_name]
                if not month_val:
                    month_val = [None] * num_days_in_month
                if column_name in year_val:
                    data = year_val[column_name]
                    data.extend(month_val)
                else:
                    data = month_val
                year_val[column_name] = data

    @staticmethod
    def strip_nonnumeric(s):
        return ''.join(char for char in s if char.isdigit())

    @staticmethod
    def get_column_values(ws, column, row_range):
        values = [ws.cell(row=i, column=column).value for i in row_range]
        return values

    @staticmethod
    def days_of_month_row_range(day_of_month):
        row_of_first_day = None
        row_of_last_day = None
        row = 1
        for day in day_of_month:
            if day == 1:
                row_of_first_day = row
            elif row_of_first_day and not day or row == len(day_of_month)-1:
                row_of_last_day = row
                break
            row += 1
        if row_of_first_day and row_of_last_day:
            return range(row_of_first_day, row_of_last_day)
        else:
            print('days_of_month invalid')
            return None

    @staticmethod
    def remove_trailing_none(lst):
        # Find the index where all following values are None
        for i in range(len(lst) - 1, -1, -1):  # Iterate backwards
            if lst[i] is not None:
                # Return list up to i+1, preserving non-trailing None values
                return lst[:i + 1]
        # If all values are None, return empty list
        return []

    @staticmethod
    def make_datetime_array(month_num, year, num_days):
        # Create list of datetime64 objects for each day
        dates = [np.datetime64(datetime(year, month_num, day)) for day in range(1, num_days + 1)]
        # Convert to numpy array
        return np.array(dates)