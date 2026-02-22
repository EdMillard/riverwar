"""
Copyright (c) 2025 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from datetime import datetime
import difflib
import math
import numpy as np
import pandas as pd
import plyvel
from openpyxl.styles import Alignment, Font
import openpyxl
import os
import re
import textwrap
import json
from pathlib import Path
from pandas import Timestamp
import pytz
from dwcd.dolores_inputs import DoloresInputs
from dwcd.inflow_outflow import InflowOutflowMonth
from source.water_year_info import WaterYearInfo
from dwcd.verify import verify
from typing import List, Callable, Any
from api.times_series import TimeSeries
from api.pool import PoolQueue, FillRun, FillType

class WaterYear:
    af_to_cfs = 1.9835
    mcphee_dead_pool_af = 152000
    gallons_to_cfs_string = '646272'
    gallons_to_af_string = '325851'

    global_file_names = {}

    run_verification = True
    load_usgs_data = False
    load_cdss_data = False
    load_usbr_data = False

    def __init__(self, ui_abstraction):
        self.pools:dict = {}
        self.logger = ui_abstraction
        self.plotter = ui_abstraction
        self.file_names = {}
        self.water_year_info:WaterYearInfo|None = None
        self.inflow_outflow_months = []
        self.y = {}
        self.n = {}
        self.calc = {}
        self.custom_data = {}
        self.config = {}
        self.units = {}
        self.formulas = {}
        self.equations = {}
        self.comments = {}
        self.inputs_dwcd = {}
        self.inputs_usbr = {}
        self.inputs_usgs = {}
        self.inputs_cdss = {}
        self.source_names_api:list[str] = []
        self.unknown_inputs = {}
        self.mcphee_elevation_to_capacity = None
        self.mcphee_elevation_to_surface_area = None
        self.generate_excel = False
        self.column_aliases = {}
        self.unit_aliases = {}
        self.sources = {}
        self.drain_runs = {}
        self.drain_queue:PoolQueue = PoolQueue([])
        self.fill_runs = {}
        self.fill_queue:PoolQueue = PoolQueue([])
        self.evap_runs = {}
        self.evap_queue:PoolQueue = PoolQueue([])
        self.active_runs = {}
        self.tolerances = {}
        self.verification_errors = []
        self.cache_base_path:Path|None = None

    def load(self, name, force_reload=False):
        print(f'Load {name}  Force reload: {force_reload}')
        data_path = Path('data/Dolores')
        file_name = data_path.joinpath(name)

        parent = file_name.parent.name
        cache_path = Path('cache/Dolores') / parent
        cache_path.mkdir(parents=True, exist_ok=True)
        if not file_name.exists():
            file_name = cache_path.joinpath(name)
            if not file_name.exists():
                self.logger.log_message(f'File not found: {name}')
                return False

        WaterYear.global_file_names['dolores_all_data_db'] = cache_path / 'dolores_all_data_db'

        WaterYear.global_file_names['variable_names_db'] = cache_path / 'variable_names_db'
        WaterYear.global_file_names['variable_names_raw_db'] = cache_path / 'variable_names_raw_db'

        WaterYear.global_file_names['variable_names_history_log'] = cache_path / 'variable_names_history.log'
        WaterYear.global_file_names['variable_names_raw_history_log'] = cache_path / 'variable_names_raw_history.log'

        cache_file_name = cache_path.joinpath(name)
        self.cache_base_path = cache_base_path = Path(cache_path.joinpath(cache_file_name.stem))
        cache_base_path.mkdir(parents=True, exist_ok=True)

        self.file_names['excel_csv'] = cache_base_path / 'excel_data.csv'
        self.file_names['inputs_csv'] = cache_base_path / 'inputs.csv'
        self.file_names['formulas_csv'] = cache_base_path / 'formulas.csv'
        self.file_names['equations_csv'] = cache_base_path / 'equations.csv'
        self.file_names['comments_csv'] =  cache_base_path / 'comments.csv'
        self.file_names['custom_data_csv'] =  cache_base_path / 'custom_data.csv'

        self.file_names['config_json'] = cache_base_path / 'config.json'
        self.file_names['units_json'] = cache_base_path / 'units.json'

        self.file_names['alias_log'] = cache_base_path / 'aliases_YEAR.log'
        self.file_names['comment_log'] = cache_base_path / 'comments_YEAR.log'
        self.file_names['equation_log'] = cache_base_path / 'equations_YEAR.log'
        self.file_names['water_year_log'] = cache_base_path / 'events_YEAR.log'
        self.file_names['message_log'] = cache_base_path / 'messages_YEAR.log'
        self.file_names['variable_name_log'] = cache_base_path / 'variables_YEAR.log'
        self.file_names['verify_log'] = cache_base_path / 'verify_YEAR.log'

        self.file_names['excel_error_log'] = cache_base_path / 'excel_errors_YEAR.log'
        self.file_names['message_excel_log'] = cache_base_path / 'excel_messages_YEAR.log'

        self.file_names['xls'] = cache_base_path / 'data.xls'
        self.file_names['report_doc'] = cache_base_path / 'dolores_report.docx'

        # Load from Excel if any of these files are missing or are older than the xls
        #
        loaded_from_excel = False
        if (force_reload
            or not self.file_names['excel_csv'].exists()
            or WaterYear.is_file_newer(file_name, self.file_names['excel_csv'])
            or not self.file_names['formulas_csv'].exists()
            or not self.file_names['inputs_csv'].exists()
            or not self.file_names['custom_data_csv'].exists()
            or not self.file_names['equations_csv'].exists()
            or not self.file_names['comments_csv'].exists()
            or not self.file_names['config_json'].exists()
            or not self.file_names['units_json'].exists()
        ):
            loaded_from_excel = True
            if force_reload:
                self.logger.log_message(f' Reloading excel file force_reload=True')
            else:
                self.logger.log_message(f' Reloading excel file, cache empty or outdated')
            self.load_from_excel(file_name)

        # Fast load from cached data files previously generated from the xls file
        #
        print('Variable active ranges:')
        df = pd.read_csv(self.file_names['excel_csv'])
        df['DATE'] = pd.to_datetime(df['DATE'])
        self.y = df.to_dict(orient='list')

        dates_tmp = self.y.pop('DATE')
        for key, lst in self.y.items():
            active_run = WaterYear.find_active_run(lst)
            if active_run is not None:
                if active_run[0] != 0 or active_run[1] != len(lst):
                    active_date_run = (WaterYear.timestamp_to_month_day(dates_tmp[active_run[0]]),
                                       WaterYear.timestamp_to_month_day(dates_tmp[active_run[1]]))
                    self.active_runs[key] = active_date_run
                    print(f'\tactive   {str(active_date_run):<21} {key}')
            else:
                print(f"\tinactive {' ':<21} {key}")
            try:
                self.n[key] = np.asarray(lst, dtype=float)
            except (ValueError, TypeError) as e:
                indices = WaterYear.collect_error_indices(lst, float, type(e))
                self.logger.log_message(f'Excel data conversion error {key}: {indices}')
                for index in indices:
                    date = dates_tmp[index]
                    date_str = WaterYearInfo.format_to_month_day(date)
                    self.logger.log_message(f'\t{date_str}[{index}] = \'{lst[index]}\'')
                self.n[key] = np.fromiter((WaterYear.to_float_or_zero(x) for x in lst), dtype=float)

        WaterYear.convert_array_dictionary_nan_to_zero(self.n)
        self.n = {'DATE': np.asarray(dates_tmp, dtype='datetime64'), **self.n}
        self.y = {'DATE': dates_tmp, **self.y}

        dates = self.y['DATE']
        self.water_year_info = WaterYearInfo.get_water_year(dates[0])
        mdt = pytz.timezone('America/Denver')
        current_datetime = datetime.now(mdt)
        current_water_year_info = WaterYearInfo.get_water_year(current_datetime)
        if self.water_year_info.year == current_water_year_info.year:
            self.water_year_info.is_current_water_year = True

        df_f = pd.read_csv(self.file_names['formulas_csv']).convert_dtypes()
        self.formulas = df_f.to_dict(orient='list')

        df_f = pd.read_csv(self.file_names['equations_csv']).convert_dtypes()
        self.equations = df_f.to_dict(orient='list')

        df_f = pd.read_csv(self.file_names['comments_csv']).convert_dtypes()
        self.comments = df_f.to_dict(orient='list')

        with self.file_names['config_json'].open("r") as file:
            self.config = json.load(file)

        with self.file_names['units_json'].open("r") as file:
            self.units = json.load(file)

        # Load lookup tables to McPhee elevation to capacity and surface area
        #
        self.mcphee_elevation_to_capacity = pd.read_csv('data/Dolores/mcphee_elevation_to_capacity.csv')
        self.mcphee_elevation_to_surface_area = pd.read_csv('data/Dolores/mcphee_elevation_to_surface_area.csv')

        # Optionally load some input channels from USBR RISE, USGS and DWR CDSS
        #
        self.unknown_inputs = WaterYear.remove_keys_from_dictionary(self.y, self.formulas)
        self.source_names_api.clear()
        if WaterYear.load_usbr_data:
            source_name_api = 'usbr'
            self.source_names_api.append(source_name_api)
            self.inputs_usbr = DoloresInputs.load_usbr(self.logger, self.water_year_info, dates[0], dates[-1])
            for input_usbr in self.inputs_usbr:
                self.logger.log_message(f'  {source_name_api} \'{input_usbr}\'')
            self.unknown_inputs = WaterYear.remove_keys_from_dictionary(self.unknown_inputs, self.inputs_usbr)
        if WaterYear.load_usgs_data:
            source_name_api = 'usgs'
            self.source_names_api.append(source_name_api)
            self.inputs_usgs = DoloresInputs.load_usgs(self.logger, self.water_year_info, dates[0], dates[-1])
            for input_usgs in self.inputs_usgs:
                self.logger.log_message(f'  {source_name_api} \'{input_usgs}\'')
            self.unknown_inputs = WaterYear.remove_keys_from_dictionary(self.unknown_inputs, self.inputs_usgs)
        if WaterYear.load_cdss_data:
            source_name_api = 'cdss'
            self.source_names_api.append(source_name_api)
            self.inputs_cdss = DoloresInputs.load_cdss(self.logger, self.water_year_info, units=self.units)
            for inputs_cdss in self.inputs_cdss:
                self.logger.log_message(f'  {source_name_api} \'{inputs_cdss}\'')
            self.unknown_inputs = WaterYear.remove_keys_from_dictionary(self.unknown_inputs, self.inputs_cdss)

        if loaded_from_excel:
            # FIXME It would be better if this was in load_from_excel but it mangles MWC, CORTEX and UTE when done there
            # Build equation log
            #
            path = self.file_names['equation_log']
            year_string = str(self.water_year_info.year)
            path = Path(str(path).replace('YEAR', year_string))
            with open(path, "w") as file:
                variable_names = list(self.y.keys())
                variable_names_inputs, variable_names_from_equations = (
                    WaterYear.equations_for_vars(variable_names, self.equations, self.formulas, self.y, self.y['DATE'],
                                                file=file))

            # Save input channels to CSV file
            #
            variable_name_inputs_with_equations = ['UTE F&R', 'MWC', 'CORTEZ NON - PROJ','CORTEZ PROJ', 'UTE PROJ']
            variable_names_inputs = WaterYear.remove_names(variable_names_inputs, variable_name_inputs_with_equations)
            self.inputs_dwcd = WaterYear.extract_to_new_dict(self.n, variable_names_inputs, remove_from_original=False)
            WaterYear.extract_tunnel_flows(self.logger, self.y, self.inputs_dwcd, dates, self.equations, self.units)
            # variable_names_inputs = list(self.inputs_dwcd.keys())

            df_inputs = pd.DataFrame(self.inputs_dwcd)
            df_inputs.to_csv(self.file_names['inputs_csv'], index=False)

            df_custom_data = pd.DataFrame(self.custom_data)
            df_custom_data.to_csv(self.file_names['custom_data_csv'], index=False)

            path = self.file_names['message_excel_log']
            year_string = str(self.water_year_info.year)
            path = Path(str(path).replace('YEAR', year_string))
            self.logger.progress_bar.save(path)
        else:
            df_i = pd.read_csv(self.file_names['inputs_csv'])
            tmp = df_i.to_dict(orient='list')
            self.inputs_dwcd = {key: np.asarray(lst, dtype=float) for key, lst in tmp.items()}

            df_c = pd.read_csv(self.file_names['custom_data_csv'])
            tmp = df_c.to_dict(orient='list')
            self.custom_data = {key: np.asarray(lst, dtype=float) for key, lst in tmp.items()}

        WaterYear.convert_array_dictionary_nan_to_zero(self.inputs_dwcd)
        self.inputs_dwcd['DATE'] = self.n['DATE']

        return True

    def load_from_excel(self, file_name):
        wb = openpyxl.load_workbook(file_name, data_only=True)
        wb_f = openpyxl.load_workbook(file_name, data_only=False)
        ws = wb['NOV']
        self.water_year_info, month, year, month_num = InflowOutflowMonth.get_year_info(ws)
        if self.water_year_info is None:
            self.logger.log_message(f'Load year info failed {file_name}')
            return False

        year_string = str(self.water_year_info.year)

        path = self.file_names['excel_error_log']
        path = Path(str(path).replace('YEAR', year_string))
        excel_log = open(path, "w")
        excel_log.write(f'Load Excel: {file_name.name}\n')
        excel_log.flush()

        path = self.file_names['alias_log']
        path = Path(str(path).replace('YEAR', year_string))
        alias_log = open(path, "w")

        self.build_alias_dictionary()
        InflowOutflowMonth.reset()
        month_names = ['NOV', 'DEC', 'JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT']
        for month_name in month_names:
            month = InflowOutflowMonth()
            if month.load_month_from_excel(wb, wb_f, self.logger, excel_log, alias_log, self.config,
                                           self.column_aliases, self.unit_aliases, month_name=month_name):
                self.inflow_outflow_months.append(month)
            else:
                self.logger.log_message(f'    Empty month {month_name}')
                break

        # We have a bunch of column arrays in each month, we are going to concatenate them all together
        # to make arrays for the water year
        #
        for month in self.inflow_outflow_months:
            InflowOutflowMonth.concatenate_month(self.y, month.values, month.num_days_in_month)
            InflowOutflowMonth.concatenate_month(self.formulas, month.formulas, month.num_days_in_month)
            InflowOutflowMonth.concatenate_month(self.equations, month.equations, month.num_days_in_month)
            InflowOutflowMonth.concatenate_month(self.comments, month.comments, month.num_days_in_month)
            WaterYear.merge_dictionaries(self.units, month.units, description=month.month)

        self.write_comment_log(self.y['DATE'], year_string)
        self.write_variable_name_archaeology(year_string)

        # Automatically generate configuration from Excel, if Python is standalone(without Excel) user has to set these
        self.auto_config_custom_data()

        # Find columns with no formulas, these are probably input channels
        #
        for key, f in self.formulas.items():
            is_all_none = all(x is None for x in f)
            if is_all_none:
                self.formulas[key] = None

        self.formulas = WaterYear.remove_none_keys_from_dictionary(self.formulas)

        self.unknown_inputs = WaterYear.remove_keys_from_dictionary(self.y, self.formulas)

        # Cache loaded data to csv and json for fast loader
        #
        self.logger.log_message(f' Save excel data to cache')

        df = pd.DataFrame(self.y)
        df.to_csv(self.file_names['excel_csv'], index=False)

        df = pd.DataFrame(self.custom_data)
        df.to_csv(self.file_names['custom_data_csv'], index=False)

        df_f = pd.DataFrame(self.formulas)
        df_f.to_csv(self.file_names['formulas_csv'], index=False)

        df_eq = pd.DataFrame(self.equations)
        df_eq.to_csv(self.file_names['equations_csv'], index=False)

        df_comm = pd.DataFrame(self.comments)
        df_comm.to_csv(self.file_names['comments_csv'], index=False)

        with self.file_names['config_json'].open("w") as file:
            json.dump(self.config, file, indent=4)

        with self.file_names['units_json'].open("w") as file:
            json.dump(self.units, file, indent=4)

        self.logger.log_message(f'Load Excel complete {file_name.name}\n')

        db_path = Path(WaterYear.global_file_names['dolores_all_data_db'])
        db = plyvel.DB(str(db_path), create_if_missing=True)
        db.put(str(self.water_year_info.year).encode('utf-8'),
               json.dumps(self.y, cls=WaterYear.DateTimeEncoder).encode('utf-8'))
        # WaterYear.print_db(db)
        db.close()

        excel_log.close()
        alias_log.close()
        return True

    def build_pool_queue_var(self, variable_name:str, fill_values, drain_values, spill_values, user_configs,
                             fill_drain_variable_names:list[str], evap_variable_names:list[str], evap_values, dates,
                             allow_transfers: bool = False, fill_max_end_date:str=None,
                             transfer_to_names:list[str]|None=None, do_paper_drains:bool=False):
        pool_queue = PoolQueue([])
        fill_runs, drain_runs, evap_runs = pool_queue.build_pool_runs(self.logger, variable_name,
                                                                     fill_values, drain_values, spill_values, user_configs,
                                                                     fill_drain_variable_names, evap_variable_names, evap_values,
                                                                     self.equations, self.y, dates,
                                                                     allow_transfers=allow_transfers,
                                                                     fill_max_end_date=fill_max_end_date,
                                                                     transfer_to_names=transfer_to_names,
                                                                     do_paper_drains=do_paper_drains)
        if fill_runs:
            self.fill_runs[variable_name] = fill_runs
        if drain_runs:
            self.drain_runs[variable_name] = drain_runs
        if evap_runs:
            self.evap_runs[variable_name] = evap_runs

    def auto_config_custom_data(self):
        # Overridden
        pass

    def get_pool_runs_all(self):
        pool_queue = PoolQueue([])
        for runs in self.fill_runs.values():
            pool_queue.extend(runs)
        for runs in self.drain_runs.values():
            pool_queue.extend(runs)
        for runs in self.evap_runs.values():
            pool_queue.extend(runs)

        pool_queue.sort(key=lambda x: x.start_day)
        return pool_queue

    def get_pool_runs_for_variable(self, variable_name:str):
        pool_queue = PoolQueue([])
        if variable_name in self.fill_runs:
            pool_queue.extend(self.fill_runs[variable_name])
        if variable_name in self.drain_runs:
            pool_queue.extend(self.drain_runs[variable_name])
        if variable_name in self.evap_runs:
            pool_queue.extend(self.evap_runs[variable_name])
        pool_queue.sort(key=lambda x: x.start_day)
        return pool_queue

    @staticmethod
    def collect_error_indices(
            items: List[str],
            func: Callable[[str], Any],
            exception_type: type = Exception
    ) -> List[int]:
        """
        Apply `func` to each string in `items` and return a list of indices
        where `func` raised an exception of type `exception_type` (or subclass).

        Parameters
        ----------
        items          : list of strings to process
        func           : the function you want to apply (e.g. int, float, my_parser, ...)
        exception_type : the exception class you want to catch (default: catch everything)

        Returns
        -------
        List[int]: indices where an error occurred
        """
        error_indices = []

        for i, s in enumerate(items):
            try:
                func(s)
            except exception_type:
                error_indices.append(i)

        return error_indices

    @staticmethod
    def variable_name_archaeology_all_years(db_path:Path, file_path:Path):
        db = plyvel.DB(str(db_path), create_if_missing=True)
        variable_names_by_year = {}
        for key in db.iterator(reverse=True, include_value=False):
            year = key.decode('utf-8')
            variable_names_by_year[year] = WaterYear.load_variable_name_archaeology_year(db, year)

        variable_names_by_month = {}
        for year, variable_names_year in variable_names_by_year.items():
            for month, variable_names_month in variable_names_year.items():
                if month in variable_names_by_month:
                    by_month = variable_names_by_month[month]
                    by_month.append((year, variable_names_month))
                else:
                    variable_names_by_month[month] = [(year, variable_names_month)]
            pass

        with (open(file_path, "w") as file):
            for month, values in variable_names_by_month.items():
                file.write(f'{month}\n')
                for value in values:
                    file.write(f'\t{value[0]} {value[1]}\n')

    @staticmethod
    def load_variable_name_archaeology_year(db, year):
        values = db.get(year.encode('utf-8'))
        json_str = values.decode('utf-8')
        variable_names = json.loads(json_str)
        return variable_names

    def write_comment_log(self, dates, year_string):
        path = self.file_names['comment_log']
        path = Path(str(path).replace('YEAR', year_string))
        f = open(path, "w")
        for variable_name, comments in self.comments.items():
            f.write(f"'{variable_name}'\n")
            comment:str|None
            for day, comment in enumerate(comments):
                if comment is not None:
                    comment = comment.replace('\n', ' ')
                    date = dates[day]
                    date_str = WaterYearInfo.format_to_month_day(date)
                    f.write(f"\t'{date_str}' {comment}\n")

    def write_variable_name_archaeology(self, year_string):

        # Variable names before aliasing
        #
        path = self.file_names['variable_name_log']
        path = Path(str(path).replace('YEAR', year_string))
        with (open(path, "w") as f):
            f.write('Variable names before aliasing:\n')
            prev_month_variables = None
            variable_names_raw_dict = {}
            last_month = False
            for month in self.inflow_outflow_months:
                variable_names_raw_dict[month.month] = month.variable_names_raw
                month_variables = str(month.variable_names_raw)
                if month.month == self.inflow_outflow_months[-1].month:
                    last_month = True
                prev_month_variables = WaterYear.print_variable_diffs(f, month.month, month_variables,
                                                                      prev_month_variables, last_month=last_month)


            # Variable names after aliasing
            #
            f.write('Variable names after aliasing:\n')
            prev_month_variables = None
            variable_names_dict = {}
            for month in self.inflow_outflow_months:
                variable_names = []
                for variable_name_raw in month.variable_names_raw:
                    if variable_name_raw in self.column_aliases:
                        variable_names.append(self.column_aliases[variable_name_raw])
                    else:
                        variable_names.append(variable_name_raw)
                variable_names_dict[month.month] = variable_names
                if month.month == self.inflow_outflow_months[-1].month:
                    last_month = True
                prev_month_variables = WaterYear.print_variable_diffs(f, month.month, str(variable_names),
                                                                      prev_month_variables, last_month=last_month)

        db_path = Path(WaterYear.global_file_names['variable_names_raw_db'])
        db = plyvel.DB(str(db_path), create_if_missing=True)
        db.put(str(self.water_year_info.year).encode('utf-8'), json.dumps(variable_names_raw_dict).encode('utf-8'))
        db.close()

        db_path = Path(WaterYear.global_file_names['variable_names_db'])
        db = plyvel.DB(str(db_path), create_if_missing=True)
        db.put(str(self.water_year_info.year).encode('utf-8'), json.dumps(variable_names_dict).encode('utf-8'))
        db.close()

    @staticmethod
    def print_variable_diffs(file, description, month_variables, prev_month_variables, last_month=False):
        if prev_month_variables is not None:
            diffs = WaterYear.diff_strings(prev_month_variables, month_variables)
            if len(diffs) > 1:
                for diff in diffs:
                    if diff.startswith('-'):
                        pass
                    else:
                        file.write(f'  {description:<9} {diff.rstrip()}\n')
            elif last_month:
                file.write(f'  {description:<9}   {month_variables}\n')
        else:
            file.write(f'  {description:<9}   {month_variables}\n')
        return month_variables

    @staticmethod
    def print_db(db):
        for key, value in db:
            year = key.decode('utf-8')
            print(f'{year}')
            variables = json.loads(value.decode('utf-8'))
            for variable_name, values in variables.items():
                print(f"\t{variable_name: <30} {values}")

    class DateTimeEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, (np.datetime64, pd.Timestamp, datetime)):
                if isinstance(obj, np.datetime64):
                    dt = pd.to_datetime(obj).to_pydatetime()
                elif isinstance(obj, pd.Timestamp):
                    dt = obj.to_pydatetime()
                else:
                    dt = obj
                return dt.strftime("%Y-%m-%d")
                # month = dt.strftime("%b")
                # day = dt.day
                # return f"{month}-{day}"
                # return str(obj)
            # if isinstance(obj, np.ndarray):
            #     return obj.tolist()
            return super().default(obj)

    def build_alias_dictionary(self):
        self.unit_aliases['AC  FT'] = 'AC FT'
        self.unit_aliases['AC. Ft.'] = 'AC FT'

        self.column_aliases['MWC NON - PROJ'] = 'MWC'

        self.column_aliases['TOTAL MVI WATER USED'] = 'MVI TOTAL WATER USED'
        self.column_aliases['MVI TOTAL'] = 'MVI TOTAL WATER USED'
        self.column_aliases['TOTAL MVI'] = 'MVI TOTAL WATER USED'
        self.column_aliases['TOTAL MVI IRRIG. WATER USED'] = 'MVI TOTAL IRRIG'

        self.column_aliases['MVI STORED in MCP Apr - Jun'] = 'MVI STORABLE'
        self.column_aliases["MVI STORABLE in MCP Apr - Jun"] = 'MVI STORABLE'
        self.column_aliases["MVI STORED Apr - Jun"] = 'MVI STORABLE'
        self.column_aliases["MVI STORABLE in MCP"] = 'MVI STORABLE'

        self.column_aliases['MVI CALL'] = 'MVI CALL STOR'

        self.column_aliases['CTX & UTE'] = 'CTZ & UTE'


        if self.water_year_info.year == 2003:
            self.column_aliases['MVI CALL & PROJ.'] = 'MVI CALL TO EX U/S EX'
            self.column_aliases['MVI CALL & PROJ'] = 'MVI CALL TO EX U/S EX'
            self.column_aliases['MVI CALL &PROJ.'] = 'MVI CALL TO EX U/S EX'
            self.column_aliases['MVI CALL + TOT EX'] = 'MVI CALL TO EX U/S EX'
            self.column_aliases['MVI CALL + TO EX'] = 'MVI CALL TO EX U/S EX'
        elif self.water_year_info.year == 2004 or self.water_year_info.year == 2005:
            self.column_aliases['MVI CALL + TOT EX'] = 'MVI CALL TO EX U/S EX'
            self.column_aliases['MVI CALL + TO EX'] = 'MVI CALL TO EX U/S EX'
        if self.water_year_info.year == 2005 or self.water_year_info.year == 2006:
            self.column_aliases['MVI PROJ & TO EX &'] = 'MVI PROJ TO EX U/S EX'

        if self.water_year_info.year == 2003:
            self.column_aliases['MVI CALL EVAP acft'] = 'MVI STORED EVAP'
            self.column_aliases['MVI STORED EVAP acft'] = 'MVI STORED EVAP'
            self.column_aliases['MVI CALL Apr - Jun'] = 'MVI STORABLE'
            self.column_aliases['MVI CALL Apr-Jun'] = 'MVI STORABLE'
            self.column_aliases['MVI CALL Apr-Jun'] = 'MVI STORABLE'

        if self.water_year_info.year == 2002:
            self.column_aliases['MVI CALL APR/OCT'] = 'MVI STORABLE'
            self.column_aliases['MVI CALL Apr - Jun'] = 'MVI STORABLE'
            self.column_aliases['MVIC DAILY'] = 'MVI STORABLE'
            self.column_aliases['MVIC Daily Call'] = 'MVI STORABLE'

        if self.water_year_info.year <= 1989:
            self.column_aliases['MVI CALL Apr - Jun'] = 'MVI STORABLE'
            self.column_aliases['MVIC Daily Call'] = 'MVI STORABLE'
            self.column_aliases['MVI CALL APR/OCT'] = 'MVI STORABLE'
            self.column_aliases['MVIC DAILY'] = 'MVI STORABLE'
            self.column_aliases['MVI CALL DAILY'] = 'MVI STORABLE'

        self.column_aliases['MVI DIVER'] = 'LONE PINE'

        self.column_aliases["DOL TUN TOT DIV"] = "DOLORES TUNNEL"
        self.column_aliases["DOLORES TUNNEL DOL TUN TOT DIV"] = "DOLORES TUNNEL"
        self.column_aliases['DOLORES TUNNEL TOT DIV'] = "DOLORES TUNNEL"
        self.column_aliases['TUNNEL TOT DIV'] = "DOLORES TUNNEL"

        self.column_aliases["NAR FILL"] = "NARR FILL"
        self.column_aliases['NAR FILLL'] = "NARR FILL"
        self.column_aliases["NAR EXTND\"D FILL"] = "NARR EXT'D FILL"

        self.column_aliases['NAR CALL AC.FT.'] = "NARR CALL"
        self.column_aliases['NAR CALL AC FT'] = "NARR CALL"
        self.column_aliases['NAR. CALL AC.FT.'] = "NARR CALL"
        self.column_aliases['NAR CALL'] = "NARR CALL"
        self.column_aliases['NAR. CALL'] = "NARR CALL"

        self.column_aliases["DIST. CLASS \"B\""] = 'DIST CLASS B'

        # self.column_aliases["UFR"] = "UF&R LEASE WATER" # Empty duplicate hidden column in Jun 2023
        self.column_aliases["UF&R LEASED MVIC SHARES"] = "UF&R LEASE WATER"
        self.column_aliases['UTES FROM DWCD'] = "UF&R LEASE WATER"
        self.column_aliases['UTES FROM MVIC'] = "UF&R LEASE WATER"

        self.column_aliases['BELOW MCPHEE'] = 'BELOW MCP'
        self.column_aliases['BELOW MC  PHEE'] = 'BELOW MCP'

        self.column_aliases['U LATERAL'] = 'U LAT'

        # Needs research, column headers are missing in May and June 1986
        # self.column_aliases['MVI CALL Apr - Jun'] = 'MVI APR- JUNE (DIV+ CALL)'

        # Units mixed into names
        self.column_aliases["MVI STOCK CFS"] = "MVI STOCK"
        self.column_aliases['MVI STOCK AC.FT.'] = "MVI STOCK"

        self.column_aliases['ACTIVE CAP AC FT'] = 'ACTIVE CAP'

        # Name shortening
        #
        self.column_aliases['DWCD FLOWS AVAILABLE FOR DWCD 585 CFS Right'] = "DWCD 505 FLOWS AVAIL"
        self.column_aliases['DWCD FLOWS AVAILABLE FOR DWCD 505 CFS'] = "DWCD 505 FLOWS AVAIL"
        self.column_aliases['DOL INFs OVER 795+SNRs 15 MIN Resolution'] = 'DOL INFs OVER 795'
        # 2022 'DWCD 585 CFS DIRECT DIV'

        self.column_aliases['MPP BELOW MCP from ctrl room'] = 'BELOW MCP MPP'
        self.column_aliases['BELOW MCP MPP from ctrl room'] = 'BELOW MCP MPP'

        # Totten and Upstream Exchange
        # "TO EX & U/S EX" first split into TO EX and U/S EX in 2022
        self.column_aliases["U/S EX IN MCP"] = "U/S EX"
        if self.water_year_info.year >= 2022:
            self.column_aliases["TO EX & U/S EX"] = "TO EX"
            self.column_aliases["TOTTEN EXCHANGE CREDIT"] = "TO EX"
            self.column_aliases["TOTTEN CREDIT"] = "TO EX"
            self.column_aliases["TOTT EX"] = "TO EX"
            self.column_aliases['TO EX &'] = "TO EX"
            self.column_aliases["TOTTEN CREDIT & U/S EX"] = "TO EX"

        self.column_aliases['UPSTREAM USERS EXCHANGE'] = "U/S USERS EXCH"

        if self.water_year_info.year <= 2003:
            # First appeared 1991, fish pool in cfs
            self.column_aliases['REMAIN RELEASE'] = 'FISH LEFT'
            # 1990 'FISH LOAN' which is 'BELOW MCP' discharge
            # 1989 and before has only 'BELOW MC PHEE'

        # 2004 Name shortening
        #
        if self.water_year_info.year <= 2009:
            self.column_aliases['TOTAl MVI'] = 'MVI TOTAL WATER USED'
        self.column_aliases['TOTAL TUNNEL Outflow'] = 'TOTAL TUNNEL Misc.'

        if self.water_year_info.year < 2004:
            self.column_aliases['LOST CANYON'] = 'LOST CAN'
            self.column_aliases['RECORDED INFLOW'] = 'RECORD INFLOW'
            self.column_aliases['RECORDED OUTFLOW'] = 'RECORD OTF'
            self.column_aliases['RECORD OUTFLOW'] = 'RECORD OTF'
            self.column_aliases['CALC OUTFLOW'] = 'CALC OTF'
            self.column_aliases['TOTAL OUTFLOW'] = 'TOTAL OTF'

            self.column_aliases['MVI/CTZ'] = 'MVI/CTZ/MWC'
            self.column_aliases['MVI,CTZ'] = 'MVI/CTZ/MWC'
            #self.column_aliases['MVI/CTZ/MWC'] = 'MVI HIGHLINE'

        if self.water_year_info.year >= 2018 or self.water_year_info.year <= 2022:
            self.column_aliases["RESERV"] = 'RESERV TOT CAP'
        if self.water_year_info.year <= 2008:
            self.column_aliases["RESERV AC  FT"] = 'RESERV TOT CAP'
        if self.water_year_info.year <= 2002:
            self.column_aliases["WETLAND"] = 'WET LAND'
        if self.water_year_info.year <= 2003:
            self.column_aliases["SWAMP"] = 'WET LAND'
        if self.water_year_info.year >= 2007 or self.water_year_info.year <= 2008:
            self.column_aliases["F&W"] = 'WET LAND'

        self.column_aliases["SEELEY"] = 'SEELEY QUEST'
        self.column_aliases["QEP"] = 'SEELEY QUEST'
        self.column_aliases["QUEST"] = 'SEELEY QUEST'
        self.column_aliases["F-BAY SPILL"] = 'F-BAY BAL'

    @staticmethod
    def convert_array_dictionary_nan_to_zero(array_dict):
        for key, array in array_dict.items():
            has_nan = np.any(np.isnan(array))
            if has_nan:
                array_dict[key] = np.nan_to_num(array)

    @staticmethod
    def to_float_or_zero(s):
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0  # Invalid strings → 0.0

    @staticmethod
    def timestamp_to_month_day(ts: datetime) -> str:
        """
        Convert a datetime (Timestamp) object to 'Mon-D' format.
        """
        if not isinstance(ts, datetime):
            raise TypeError("Argument must be a datetime object")

        month_abbr = ts.strftime("%b")  # 'Jan', 'Feb', ..., 'Dec'
        day = ts.day
        return f"{month_abbr}-{day}"

    @staticmethod
    def find_active_run(lst):
        """
            Given a list with elements that are float('nan') or numbers, returns a tuple
            (first_index, last_index) containing the indices of the first and last
            non-NaN elements. NaN values in between non-NaN elements are ignored.

            If the list has no NaN values at all, returns None.
            If all elements are NaN, returns (None, None).

            Args:
                lst: List containing float('nan') and/or numbers (int, float, etc.)

            Returns:
                tuple(int, int) or None
            """
        if not lst:
            return None

        # Check if list contains any NaN
        has_nan = any(math.isnan(x) for x in lst if isinstance(x, float))
        if not has_nan:
            return 0, len(lst)

        first = None
        last = None

        for i, val in enumerate(lst):
            # Only consider non-NaN numeric values
            if isinstance(val, (int, float)) and not math.isnan(val):
                if first is None:
                    first = i
                last = i  # Update last every time we see a valid number

        if first is not None and last is not None:
            return first, last
        else:
            return None

    @staticmethod
    def find_string_or_data_runs(lst:list, data, default_equations, break_run_on_zero=True):
        runs = []
        n = len(lst)
        data_out:list|None = [None] * n
        day:int = 0

        has_custom_data = False
        while day < n:
            # Skip over None values
            while day < n:
                if lst[day] is None and (data[day] is None) or (data[day] == 0 and break_run_on_zero):
                    day += 1
                else:
                    break
            if day == n:
                break

            # Found start of a string run
            start = day
            while day < n:
                e = lst[day]
                if e is None and (data[day] is None) or (data[day] == 0 and break_run_on_zero):
                    break
                else:
                    if e is None or '[' not in e:
                        if data[day] is not None and data[day] != 0:
                            data_out[day] = float(data[day])
                            has_custom_data = True
                    elif e is not None and e.endswith('*0'):
                        data_out[day] = float(data[day])
                        has_custom_data = True
                    elif default_equations is not None:
                        if '[' not in e:
                            data_out[day] = float(data[day])
                            has_custom_data = True
                        else:
                            matched = False
                            for default_equation in default_equations:
                                if e == default_equation:
                                    matched = True
                                    break

                            if not matched:
                                data_out[day] = float(data[day])
                                has_custom_data = True
                    day += 1

            # now at the end (exclusive) of the run
            end = day - 1
            runs.append((start, end))

        if not has_custom_data:
            data_out = None
        return runs, data_out

    @staticmethod
    def get_flow_runs(config, logger, variable_name):
        runs = None
        user_configs = config.get('user')
        if user_configs is not None:
            runs = user_configs.get(variable_name)
            # if runs is not None and runs:
                # print(f'  Flow Runs   \'{variable_name}\'\n\t{runs}')
                # logger.log_message(f'  Flow Runs   \'{variable_name}\'\n\t{runs}')

        return runs

    @staticmethod
    def get_pool_runs(config, logger, variable_name):
        fill_runs = None
        drain_runs = None
        evap_runs = None
        custom_runs = None
        user_configs = config.get('user')
        if user_configs is not None and variable_name in user_configs:
            fill_runs, drain_runs, evap_runs, custom_runs = user_configs.get(variable_name)
            message = f'  Pool Runs \'{variable_name}\''
            if fill_runs is not None and fill_runs:
                message += f'\n\t    Fill\t={fill_runs}'
            if drain_runs is not None and drain_runs:
                message += f'\n\t    Drain\t={drain_runs}'
            if evap_runs is not None and evap_runs:
                message += f'\n\t    Evap\t={evap_runs}'
            if custom_runs is not None and custom_runs:
                message += f'\n\t    Custom\t={custom_runs}'

            #logger.log_message(message)
            print(message)

        return fill_runs, drain_runs, evap_runs, custom_runs

    @staticmethod
    def get_date_runs_as_day_runs(date_runs, dates):
        day_runs = []
        for date_run in date_runs:
            start_day = WaterYear.day_for_date(dates, date_run[0])
            end_day = WaterYear.day_for_date(dates, date_run[1])
            day_run = [start_day, end_day] + list(date_run[2:])
            day_runs.append(day_run)
        return day_runs

    @staticmethod
    def get_pool_runs_as_day_runs(config, logger, variable_name, dates):
        fill_day_runs = []
        drain_day_runs = []
        evap_day_runs = []
        custom_day_runs = []
        user_configs = config.get('user')
        if user_configs is not None and variable_name in user_configs:
            fill_runs, drain_runs, evap_runs, custom_runs = user_configs.get(variable_name)
            fill_day_runs = WaterYear.get_date_runs_as_day_runs(fill_runs, dates)
            drain_day_runs = WaterYear.get_date_runs_as_day_runs(drain_runs, dates)
            evap_day_runs = WaterYear.get_date_runs_as_day_runs(evap_runs, dates)
            if custom_runs is not None:
                custom_day_runs = WaterYear.get_date_runs_as_day_runs(custom_runs, dates)

            message = f'  Pool Runs \'{variable_name}\''
            if fill_runs is not None and fill_runs:
                message += f'\n\t    Fill\t={fill_runs}'
            if drain_runs is not None and drain_runs:
                message += f'\n\t    Drain\t={drain_runs}'
            if evap_runs is not None and evap_runs:
                message += f'\n\t    Evap\t={evap_runs}'
            if custom_runs is not None and custom_runs:
                message += f'\n\t    Custom\t={custom_runs}'

            #logger.log_message(message)
            print(message)

        return fill_day_runs, drain_day_runs, evap_day_runs, custom_day_runs

    @staticmethod
    def apply_custom_data(custom_data, variable_name, variable):
        if custom_data is not None:
            var_custom_data = custom_data.get(variable_name)
            if var_custom_data is not None:
                arr = np.array(var_custom_data, dtype=float)  # Converts None -> np.nan
                mask = ~np.isnan(arr)  # True where value was not None
                variable[mask] = arr[mask]

    @staticmethod
    def day_runs_to_date_string_runs(runs, datetimes):
        """
        Convert index-based runs to date strings in 'Mon-D' format.

        Args:
            runs: List of tuples (start_idx, end_idx) from string runs
            datetimes: List of datetime objects corresponding to original indices

        Returns:
            List of tuples ('start_date_str', 'end_date_str') in 'Nov-1' format
        """
        result = []
        for run in runs:
            start_idx = run[0]
            start_dt = datetimes[start_idx]
            end_idx = run[1]
            end_dt = datetimes[end_idx]

            # Format as 'Nov-1', 'Dec-25', etc.
            start_str = start_dt.strftime('%b-%-d').replace(' 0', ' ')  # Handle single-digit days
            end_str = end_dt.strftime('%b-%-d').replace(' 0', ' ')

            date_run = (start_str, end_str) + run[2:]
            result.append(date_run)

        return result

    @staticmethod
    def build_mvi_stock_run_detail(variable_name, user_configs, equations, dates):
        f = equations.get(variable_name)
        if f:
            previous = []
            runs = []
            start_day = -1
            for day, equation in enumerate(f):
                result = []
                if equation is not None:
                    if 'x[\'DOLORES TUNNEL\']' in equation:
                        result.append('MVI HIGHLINE')
                    if '+x[\'LONE PINE\']' in equation or '=x[\'LONE PINE\']' in equation:
                        result.append('LONE PINE')
                    if '+x[\'U LAT\']' in equation or '=x[\'U LAT\']' in equation:
                        result.append('U LAT')
                    if '+x[\'MVI TOTAL WATER USED\'' in equation or '=x[\'MVI TOTAL WATER USED\'' in equation:
                        result.append('MVI TOTAL WATER USED')
                if previous != result:
                    if previous:
                        runs.append((start_day, day-1, previous))
                    if result:
                        start_day = day
                    previous = result
            if runs:
                date_runs = WaterYear.day_runs_to_date_string_runs(runs, dates)
                all_runs = user_configs.get(variable_name, []) + date_runs
                user_configs[variable_name] = all_runs
                return all_runs
        return []

    @staticmethod
    def build_upstream_exchange_fill_detail(variable_name, user_configs, equations, data, dates, ground_hog_discharges):
        f = equations.get('U/S EX')
        d = data.get('U/S USERS EXCH')
        if f and d:
            previous_value = None
            runs = []
            start_day = -1
            default_value = 8
            for day, us_users_exch in enumerate(d):
                date_str = WaterYearInfo.format_to_month_day(dates[day])
                if date_str == 'Jul-19':
                    pass

                if us_users_exch is not None and f[day] is not None:
                    if ground_hog_discharges is None:
                        if previous_value is None:
                            us_users_exch = default_value
                        else:
                            us_users_exch = previous_value
                        print(f'build_upstream_exchange {dates[day]} \'U/S USERS EXCH\'={us_users_exch}')
                    else:
                        ground_hog_discharge = ground_hog_discharges[day]
                        if ground_hog_discharges[day] is not None:
                            if us_users_exch == ground_hog_discharges[day]:
                                if previous_value is None:
                                    us_users_exch = default_value
                                else:
                                    us_users_exch = previous_value
                                if ground_hog_discharge == 0:
                                    if previous_value:
                                        runs.append((start_day, day - 1, previous_value))
                                        break
                                    else:
                                        continue
                            elif us_users_exch > ground_hog_discharges[day]:
                                pass
                    if us_users_exch != previous_value:
                        if previous_value:
                            runs.append((start_day, day - 1, previous_value))
                        if us_users_exch is not None and us_users_exch != 0:
                            start_day = day
                        previous_value = us_users_exch
                else:
                    if previous_value:
                        runs.append((start_day, day - 1, previous_value))
                        previous_value = 0
            if runs:
                date_runs = WaterYear.day_runs_to_date_string_runs(runs, dates)
                all_runs = user_configs.get(variable_name, []) + date_runs
                user_configs[variable_name] = all_runs
                return all_runs
        return []

    @staticmethod
    def build_custom_data(variable_name, equations, data, custom_data, strings):
        has_custom_data = False

        f = equations.get(variable_name)
        d = data.get(variable_name)
        if f and d:
            data_out:List|None = [None] * len(f)
            for day, equation in enumerate(f):
                if equation is not None:
                    for string in strings:
                        if string not in equation:
                            data_out[day] = float(d[day])
                            has_custom_data = True
                            break
                elif d[day] is not None:
                    data_out[day] = float(d[day])
                    has_custom_data = True

            if has_custom_data:
                custom_data[variable_name] = data_out
        return has_custom_data

    @staticmethod
    def build_flow_runs(variable_name, config, custom_data, equations, exl, dates, default_equations=None):
        date_runs = None
        data_out = None
        f = equations.get(variable_name)
        data = exl.get(variable_name)
        if f is not None and data is not None:
            day_runs, data_out = WaterYear.find_string_or_data_runs(f, data, default_equations)
            date_runs = WaterYear.day_runs_to_date_string_runs(day_runs, dates)
            if date_runs:
                config[variable_name] = date_runs
            else:
                pass
            if data_out is not None:
                if custom_data is not None:
                    custom_data[variable_name] = data_out
        return date_runs, data_out

    @staticmethod
    def pre_process(logger, exl, inp, dates):
        # Create empty arrays for variables that may have been empty in Excel and pitched
        # FIXME - Need to settle the right place to deal with these, maybe in Excel loade
        missing_inp_variables = [
            'DWR EST PAN EVAP',     # Missing before 2021
            'WET LAND',             # Has formulas in 2019
            'SEELEY QUEST',
            'PAN EVAP',             # Missing before 2017
            'D/S SENIOR MVIC',      # Missing before 2011
            'D/S SENIOR PROJ',      # Missing before 2011
            'DV  CR CANAL',         # Has Formula 1998, 1997...
            'U LAT',                # Has Formula 1996
            'LONE PINE',            # Has Formula 1995
        ]
        for missing_variable in missing_inp_variables:
            if missing_variable not in inp:
                if missing_variable in exl:
                    inp[missing_variable] = exl[missing_variable]
                    # logger.log_message(f'  \'{missing_variable}\'')
                else:
                    inp[missing_variable] = WaterYear.output(len(dates))
                    logger.log_message(f'  \'{missing_variable}\' missing')

        missing_exl_variables = [
            'SPILL',                        # Empty column in no spill years
            'U/S EX',                       # Mixed in with TO EX before 2022
            'CUM NARR IN MCPHEE',           # Missing before 2019
            'MCPHEE DAILY EVAP AF',         # Missing before 2021
            'CORTEZ NON - PROJ',            # Missing before 2011, CTZ appeared August, 2003
            'CORTEZ PROJ',                  # Missing before 2011
            'UTE PROJ',                     # Missing before 2011
            'MWC',                          # Missing before 2003
            'NARR FILL',                    # Missing before 2010, may be 'NARR CALL'
            # 'MVI CALL STOR',                # Missing before 2010, may be 'MVI CALL'
            'MVI TOTAL WATER USED',
            'DIST CLASS B',                 # Missing before 2009
            'MVI STORED EVAP',              # Missing before 2009
            # 'TO EX',                        # Missing before 2005 MVI PROJ, TO EX and US EX all in one column
            'MVI STORABLE',                 # Missing before 2004
        ]
        for missing_variable in missing_exl_variables:
            if missing_variable not in exl:
                if missing_variable in inp:
                    exl[missing_variable] = inp[missing_variable]
                    # logger.log_message(f'  \'{missing_variable}\' exl was in inputs?')
                else:
                    exl[missing_variable] = WaterYear.output(len(dates))
                    logger.log_message(f'  \'{missing_variable}\' missing')

    def post_process(self):
        unimplemented = list(WaterYear.remove_keys_from_dict(self.n, self.inputs_dwcd, self.calc))
        results = WaterYear.check_arrays_zero_or_nan(self.n, unimplemented)

        really_unimplemented = []
        for key, is_zero in results.items():
            if not is_zero:
                if len(key) == 2 and key[0] == 'B':
                    # Columns with no variable, probably have debug stuff in them
                    pass
                else:
                    really_unimplemented.append(key)
            else:
                pass

        if len(really_unimplemented):
            self.logger.log_message(f'Unimplemented: ')
            for key in really_unimplemented:
                self.logger.log_message(f'  \'{key}\'')

        self.logger.log_message(f'Python generated')
        for variable_name in self.calc:
            if variable_name not in self.y:
                self.logger.log_message(f"  '{variable_name}' python generated")

        if self.generate_excel:
            self.logger.log_message(f'Generate Excel')
            self.export_to_excel()

        if WaterYear.run_verification:
            path = self.file_names['verify_log']
            year_string = str(self.water_year_info.year)
            path = Path(str(path).replace('YEAR', year_string))
            with (open(path, "w") as file):
                self.logger.log_message(f'Starting Verification')
                verification_errors = 0
                for key, value in self.calc.items():
                    if key != 'DATE':
                        unit = self.units.get(key)
                        variables, text, pass_fail = verify(file, self.logger, self.n, self.calc, self.custom_data, self.tolerances, key,
                                        self.formulas, self.equations, self.comments, 'Excel', 'Python')
                        if variables:
                            if pass_fail == 'FAIL':
                                verification_errors += 1
                            self.verification_errors.append((variables, text))
                        file.flush()
                if verification_errors:
                    self.logger.log_message(f' Verified Accounting {len(self.calc)} channels {verification_errors} failed')
                else:
                    self.logger.log_message(f' Verified Accounting {len(self.calc)} channels Success!!!')

        # Add logs to message log for quick click access
        for name, path in  self.file_names.items():
            if '.log' in path.name:
                year_string = str(self.water_year_info.year)
                path = Path(str(path).replace('YEAR', year_string))
                self.logger.log_message(f'  \'{path.name}\'')

    @staticmethod
    def flexible_diff(times_series:list[TimeSeries], tolerance=0.001):
        text = ''
        if len(times_series) <= 1:
            return None, text

        dt1, y1 = times_series[0].get_data()
        source1 = times_series[0].source_name

        # if len(times_series) > 2:
        #     dt2, y2 = times_series[-1].get_data()
        #    source2 = times_series[-1].source_name
        # else:
        dt2, y2 = times_series[1].get_data()
        source2 = times_series[1].source_name

        num_days = len(y1)
        if len(y2) > num_days:
            num_days = len(y2)
        diffs = WaterYear.output(num_days)
        fmt = WaterYearInfo.format_float
        for day in range(0, len(y1)):
            if day == 222:
                pass
            if dt1[day].astype('datetime64[D]') == dt2[day].astype('datetime64[D]'):
                if math.isclose(y1[day], y2[day],  abs_tol=tolerance):
                    diffs[day] = 0
                else:
                    diffs[day] = y2[day] - y1[day]
                    date = dt1[day]
                    date_str = WaterYearInfo.format_to_month_day(date)

                    text += f'\'{date_str}\': diff={diffs[day]:{fmt}} {source1}={y1[day]:{fmt}} {source2}={y2[day]:{fmt}}\n'
            else:
                print(f'sample missing {dt1[day]} {dt2[day]}')
                if dt1[day] > dt2[day]:
                    pass
                else:
                    pass

        return diffs, text

    @staticmethod
    def get_config(config, kind, variable_name):
        if variable_name in config :
            return config[variable_name]
        else:
            msg = f'  {kind}/{variable_name} not found in config'
            logger = config.get('logger')
            if logger:
                logger.log_message(msg)
            else:
                print(msg)
            return 0

    @staticmethod
    def extract_tunnel_flows(logger, exl, inp, dates, equations, units):
        # Ute Farm and Ranch, this is synonymous with DWR CDSS Towaoc canal flow, with 1.05 adjustment removed
        logger.log_message(' Analyze Dolores Tunnel Users')
        equations_ute_fr = equations.get('UTE F&R')
        flows = WaterYear.extract_tunnel_flow(logger, 'UTE F&R', equations_ute_fr, exl, dates)
        if flows is not None:
            inp['UTE F&R CFS'] = flows
            units['UTE F&R CFS'] = 'CFS'
        else:
            inp['UTE F&R CFS'] = WaterYear.output(len(dates))

        # City of Cortez Non Project water
        # 'CORTEZ NON - PROJ'
        # 	 2024-11-01 to 2025-04-01
        # 		  0.0
        # 	 2025-04-01 to 2025-04-02
        # 		+ =IF(((1.9*1.04)-exl['UTE PROJ'][day])>=4.2,4.2,(1.9*1.04)-exl['UTE PROJ'][day])
        # 	 2025-09-01 to 2025-09-02 to 2025-10-01
        # 		+ =2429300/325851/1.9835*1.02
        # 	 2025-10-01 to 2025-10-02 to 2025-10-14
        # 		+ =IF(((4.2*1.04)-exl['UTE PROJ'][day])>=4.2,4.2,(4.2*1.04)-exl['UTE PROJ'][day])
        #
        equations_non_proj = equations.get('CORTEZ NON - PROJ')
        flows = WaterYear.extract_tunnel_flow(logger, 'CORTEZ NON - PROJ', equations_non_proj, exl, dates)
        if flows is not None:
            inp['CORTEZ NON - PROJ CFS'] = flows
            units['CORTEZ NON - PROJ CFS'] = 'CFS'
        else:
            inp['CORTEZ NON - PROJ CFS'] = WaterYear.output(len(dates))

        # City of Cortez Project water
        # 'CORTEZ PROJ'
        # 2024-11-01 to 2024-11-02 to 2025-10-14
        # 		  =IF(((2.5*1.04)-exl['UTE PROJ'][day])>=4.2,4.2,(2.5*1.04)-exl['UTE PROJ'][day])
        #
        equations_proj = equations.get('CORTEZ PROJ')
        flows = WaterYear.extract_tunnel_flow(logger, 'CORTEZ PROJ', equations_proj, exl, dates)
        if flows is not None:
            inp['CORTEZ PROJ CFS'] = flows
            units['CORTEZ PROJ CFS'] = 'CFS'
        else:
            inp['CORTEZ PROJ CFS'] = WaterYear.output(len(dates))

        # Montezuma Water Company project water
        #
        equations_mwc = equations.get('MWC')
        flows = WaterYear.extract_mwc_flow(logger, 'MWC', equations_mwc)
        if flows is not None:
            inp['MWC CFS'] = flows
            units['MWC CFS'] = 'CFS'
        else:
            inp['MWC CFS'] = WaterYear.output(len(dates))

        # Ute Project water, I think this is Towaoc Domestic water via Cortez treatment plant
        #
        equations_non_proj = equations.get('UTE PROJ')
        flows = WaterYear.extract_tunnel_flow(logger, 'UTE PROJ', equations_non_proj, exl, dates)

        if flows is not None:
            inp['UTE PROJ CFS'] = flows
            units['UTE PROJ CFS'] = 'CFS'
        else:
            inp['UTE PROJ CFS'] = WaterYear.output(len(dates))

    @staticmethod
    def extract_mwc_flow(logger, variable_name, equations):
        if equations is None:
            logger.log_message(f'  \'{variable_name}\' has no equations')
            return None
        num_days = len(equations)
        inputs = WaterYear.output(num_days)
        for day in range(0, num_days):
            equation = equations[day]
            try:
                if isinstance(equation, str) and len(equation):
                    result = WaterYear.find_first_number(equation)
                    if result:
                        if WaterYear.gallons_to_af_string in equation:
                            af = WaterYear.gallons_to_acre_feet(float(result[0]))
                            inputs[day] = af / WaterYear.af_to_cfs
                        elif WaterYear.gallons_to_cfs_string in equation:
                            cfs = WaterYear.gallons_to_acre_feet(float(result[0])) / WaterYear.af_to_cfs
                            inputs[day] = cfs
                        else:
                            inputs[day] = float(result[0])
                    else:
                        logger.log_message(f'  \'{variable_name}\' extract_mwc_flow failed: {equation}')
            except (IndexError, ValueError, TypeError) as e:
                logger.log_message(f'  \'{variable_name}\' Exception on equation: {equation} {e}')
        return inputs

    @staticmethod
    def extract_tunnel_flow(logger, variable_name, equations, exl, dates):
        if equations is None:
            logger.log_message(f'  \'{variable_name}\' has no equations')
            return None
        # print(f'Extract tunnel flow: {variable_name}')
        num_days = len(equations)
        inputs = WaterYear.output(num_days)
        for day in range(0, num_days):
            equation = equations[day]
            try:
                if isinstance(equation, str) and len(equation):
                    if '+' in equation and 'CORTEZ' in variable_name:
                        pass
                    result = WaterYear.find_floats_multiplied_by_104(equation)
                    if result:
                        if len(result) == 2:
                            if result[0][0] == result[1][0]:
                                # print(f'\t{day} {dates[day]} {result[0][0]}  {equation}')
                                inputs[day] =float(result[0][0])
                            else:
                                logger.log_message(f'  \'{variable_name}\' value mismatch: {day} {dates[day]} {equation}')
                        elif len(result) == 1:
                            inputs[day] = float(result[0][0])
                        else:
                            logger.log_message(f'  \'{variable_name}\' unexpected equation: {day} {dates[day]}  {equation}')
                    else:
                            #result = WaterYear.find_floats_multiplied_by_106(equation)
                            result = WaterYear.extract_constant_before_106(equation)
                            if result and isinstance(result, (int, float)):
                                # print(f'\t{day} {dates[day]} {result} {equation}')
                                inputs[day] = result
                            else:
                                result = WaterYear.find_first_number(equation)
                                if result:
                                    # print(f'\t{day} {dates[day]} {result[0]} {equation}')
                                    if WaterYear.gallons_to_af_string in equation:
                                        af = WaterYear.gallons_to_acre_feet(float(result[0]))
                                        inputs[day] = af / WaterYear.af_to_cfs
                                    elif WaterYear.gallons_to_cfs_string in equation:
                                        cfs = WaterYear.gallons_to_acre_feet(float(result[0])) / WaterYear.af_to_cfs
                                        inputs[day] = cfs
                                    else:
                                        inputs[day] = float(result[0])
                                    # if variable_name == 'UTE F&R':
                                    #     inputs[day] = float(result[0])
                                    #    af = WaterYear.gallons_to_acre_feet(float(result[0]))
                                    #   inputs[day] = af / WaterYear.af_to_cfs
                                else:
                                    coefficient_1, op, coefficient_2 = WaterYear.parse_first_parens_expression(equation)
                                    if coefficient_1 is not None and coefficient_2 is not None:
                                        if op == '+':
                                            inputs[day] = float(coefficient_1) + float(coefficient_2)
                                            # print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {inputs[day]}={coefficient_1}+{coefficient_2} {equation}')
                                        elif op == '*':
                                            if coefficient_2 == '1.04' or coefficient_2 == '1.06':
                                                inputs[day] = float(coefficient_1)
                                                print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {coefficient_1} {equation}')
                                            elif coefficient_2 == '1.05' and variable_name == 'UTE F&R':
                                                inputs[day] = float(coefficient_1)
                                                print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {coefficient_1} {equation}')
                                            else:
                                                inputs[day] = float(coefficient_1) * float(coefficient_2)
                                                print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {inputs[day]:6.2f}={coefficient_1}*{coefficient_2} {equation}')
                                        elif op == '/':
                                            logger.log_message(
                                                f'  \'{variable_name}\' unexpected divide:  {day} {dates[day]} {equation}')
                                        else:
                                            print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {coefficient_1} {op} {coefficient_2} {equation}')
                                    elif coefficient_1 is not None:
                                        if '/' in coefficient_1:
                                            parts = coefficient_1.split('/')
                                            if len(parts) == 2:
                                                coefficient_1 = parts[0]
                                                coefficient_2 = parts[1]
                                                if coefficient_2 == WaterYear.gallons_to_af_string:
                                                    af = WaterYear.gallons_to_acre_feet(float(coefficient_1))
                                                    inputs[day] = af / WaterYear.af_to_cfs
                                                elif coefficient_2 == WaterYear.gallons_to_cfs_string:
                                                    cfs = WaterYear.gallons_to_acre_feet(float(coefficient_1)) / WaterYear.af_to_cfs
                                                    inputs[day] = cfs
                                                else:
                                                    print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {coefficient_1} {coefficient_2} {equation}')
                                                    inputs[day] = float(coefficient_1)
                                            else:
                                                print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {equation}')
                                        else:
                                            print(f'  \'{variable_name}\' extract tunnel  {day} {dates[day]} {coefficient_1} {equation}')
                                            inputs[day] = float(coefficient_1)
                                    else:
                                        result = WaterYear.find_gallons_divisor(equation, WaterYear.gallons_to_cfs_string)
                                        if result is None:
                                            result = WaterYear.find_gallons_divisor(equation, WaterYear.gallons_to_af_string)
                                        if result is not None:
                                            inputs[day] = result
                                        else:
                                            logger.log_message(f'  \'{variable_name}\' unexpected equation:  {day} {dates[day]} {equation}')
                else:
                    if equation is None:
                        fallback = exl[variable_name][day]
                        if not math.isnan(fallback):
                            if variable_name == 'UTE PROJ':
                                fallback = float(fallback) / 1.06 / 1.04
                                # print(f'\t\'{variable_name}\' use data as fallback {day} {dates[day]} {fallback:8.2f}')
                            elif variable_name == 'CORTEZ NON - PROJ':
                                fallback = float(fallback) / 1.04
                            # else:
                                # print(f'  \'{variable_name}\' use data as fallback: {day} {dates[day]} {fallback:8.2f}')
                            inputs[day] = float(fallback)
                        else:
                            inputs[day] = 0.0
                    else:
                        logger.log_message(f'  \'{variable_name}\' unexpected equation: {day} {dates[day]} {equation}')
                        inputs[day] = 0.0
            except (IndexError, ValueError, TypeError) as e:
                logger.log_message(f'  \'{variable_name}\' Exception on equation: {equation} {e}')
        return inputs

    @staticmethod
    def find_gallons_divisor(equation, gallons_divisor_string):
        if gallons_divisor_string in equation:
            coefficient = equation.split(gallons_divisor_string, 1)[0]
            if coefficient:
                coefficient = coefficient.strip('=(/')
                try:
                    if '+' in coefficient:
                        pass
                    result = float(coefficient) / float(gallons_divisor_string)
                    if gallons_divisor_string == WaterYear.gallons_to_af_string:
                        result /= WaterYear.af_to_cfs
                    return result
                except ValueError as e:
                    print(e)
            else:
                return None
        else:
            return None

    def export_to_excel(self):
        df_data = pd.DataFrame(self.y)
        df_data = WaterYear.insert_units_row(df_data, self.units)

        df_python = pd.DataFrame(self.calc)
        df_python = WaterYear.insert_units_row(df_python, self.units)

        inputs_with_date = {'Date': self.calc['DATE'], **self.inputs_dwcd}
        df_inputs = pd.DataFrame(inputs_with_date)
        df_inputs = WaterYear.insert_units_row(df_inputs, self.units)

        with pd.ExcelWriter(self.file_names['xls'], engine='openpyxl') as writer:
            df_data.to_excel(writer, sheet_name='Excel', index=False)
            ws = writer.sheets['Excel']
            ws.freeze_panes = ws['B2']
            # Date column format
            #
            for row in range(2, len(df_data) + 2):
                cell = ws.cell(row=row, column=1)
                cell.number_format = 'dd-mmm'
            WaterYear.format_sheet(ws, df_data)

            df_python.to_excel(writer, sheet_name='Python', index=False)
            ws = writer.sheets['Python']
            ws.freeze_panes = ws['B2']
            for row in range(2, len(df_python) + 2):
                cell = ws.cell(row=row, column=1)
                cell.number_format = 'dd-mmm'
            WaterYear.format_sheet(ws, df_python)

            df_inputs.to_excel(writer, sheet_name='Inputs', index=False)
            ws = writer.sheets['Inputs']
            ws.freeze_panes = ws['B2']
            for row in range(2, len(df_python) + 2):
                cell = ws.cell(row=row, column=1)
                cell.number_format = 'dd-mmm'
            WaterYear.format_sheet(ws, df_inputs)

    @staticmethod
    def insert_units_row(df, variable_name_to_units):
        variable_names = df.columns.tolist()
        units = []
        for variable_name in variable_names:
            unit = variable_name_to_units.get(variable_name)
            if unit:
                units.append(unit)
            else:
                units.append('')
        df.loc[-1] = units  # Add units row at index -1
        df.index = df.index + 1  # Shift index
        df = df.sort_index()  # Sort to move units row to position 0
        return df.reset_index(drop=True)

    @staticmethod
    def format_sheet(ws, df):
        # Set font for everything
        font = Font(name='Arial Narrow', size=10)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font

        # Set default float format to for all columns except date
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.number_format = '0.00'

        # Variable name ahd units Row
        #
        ws.row_dimensions[1].height = 25
        for col in range(1, len(df.columns) + 1):
            for row in range(1, 2 + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

    @staticmethod
    def get_capacities(elevations, df, fallback_capacity=5000.0):
        """
        Find capacities for an array of elevations by identifying the row with the closest smaller elevation
        and using the column corresponding to the second decimal digit.

        Parameters:
        - elevations (np.ndarray): Array of elevation values (e.g., [6666.56, 6666.23, ...]).
        - df (pd.DataFrame): DataFrame with 'Elevation' and capacity columns ('0.00' to '0.09').
        - fallback_capacity (float, optional): Value to return for None/NaN elevations.

        Returns:
        - np.ndarray: Array of corresponding capacities.
        """
        # Ensure elevations is a NumPy array
        elevations = np.asarray(elevations)

        # Initialize output array with fallback capacity for NaN/None cases
        capacities = np.full_like(elevations, fallback_capacity, dtype=float)

        # Process each elevation
        for i, elevation in enumerate(elevations):
            # Skip if elevation is NaN or None
            if elevation is None or np.isnan(elevation):
                continue

            # Get second decimal digit
            elev_str = f"{elevation:.2f}"
            try:
                second_decimal = int(elev_str.split('.')[1][-1])
            except (IndexError, ValueError):
                second_decimal = 0  # Default to 0 if decimal part is invalid

            # Map second decimal to column name
            col_name = f"0.0{second_decimal}"

            # Ensure column exists
            if col_name not in df.columns:
                continue

            # Find the row with the closest smaller elevation
            try:
                smaller = df[df['Elevation'] <= elevation]
                if smaller.empty:
                    continue
                row_index = smaller['Elevation'].idxmax()
                capacities[i] = df.loc[row_index, col_name]
            except:
                continue  # Leave fallback capacity if lookup fails

        return capacities

    @staticmethod
    def get_surface_areas(elevations, df, fallback_area=0.0):
        """
        Find surface areas for an array of elevations using a lower-resolution lookup table.
        Uses the row with the closest smaller elevation and the column matching the second decimal digit.

        Parameters:
        - elevations (np.ndarray): Array of elevation values (e.g., [6666.56, 6666.23, ...]).
        - df (pd.DataFrame): DataFrame with 'Elevation' (10-ft increments) and surface area
          columns ('0.0' to '9.0').
        - fallback_area (float, optional): Value to return for None/NaN or invalid elevations.

        Returns:
        - np.ndarray: Array of corresponding surface areas.
        """
        # Ensure elevations is a NumPy array
        elevations = np.asarray(elevations)

        # Initialize output array with fallback area
        areas = np.full_like(elevations, fallback_area, dtype=float)

        # Process each elevation
        for i, elevation in enumerate(elevations):
            # Skip if elevation is NaN or None
            if elevation is None or np.isnan(elevation):
                continue

            # Get foot digit
            fraction = elevation % 1.0  # type: ignore
            if fraction == 0.5:
                rounded = np.ceil(elevation)
            else:
                rounded = np.round(elevation)
            elev_str = f"{rounded:.0f}"
            try:
                foot = int(elev_str[-1])  # Last digit of feet
            except (IndexError, ValueError):
                foot = 0

            # Map second decimal to column name (0to 9)
            col_name = f"{foot}"

            # Ensure column exists
            if col_name not in df.columns:
                continue

            # Find the row with the closest smaller elevation
            try:
                smaller = df[df['Elevation'] <= rounded]
                if smaller.empty:
                    continue
                row_index = smaller['Elevation'].idxmax()
                areas[i] = area = df.loc[row_index, col_name]
                areas[i] = area
            except:
                continue  # Leave fallback area if lookup fails

        return areas

    @staticmethod
    def get_surface_area_interpolated(elevations, df, fallback_area=0.0):
        """
        Find interpolated surface areas for an array of elevations using bilinear interpolation
        on a low-resolution lookup table (10-ft elevation increments, 1-ft column offsets).

        Parameters:
        - elevations (np.ndarray): Array of elevation values (e.g., [6666.56, 6666.23, ...]).
        - df (pd.DataFrame): DataFrame with 'Elevation' (10-ft increments) and surface area
          columns ('0.0' to '9.0').
        - fallback_area (float, optional): Value for None/NaN or invalid elevations.

        Returns:
        - np.ndarray: Array of interpolated surface areas.
        """
        # Ensure elevations is a NumPy array
        elevations = np.asarray(elevations)

        # Initialize output array with fallback area
        areas = np.full_like(elevations, fallback_area, dtype=float)

        # Process each elevation
        for i, elevation in enumerate(elevations):
            # Skip if elevation is NaN or None
            if elevation is None or np.isnan(elevation):
                continue

            # Find bracketing elevations
            smaller = df[df['Elevation'] <= elevation]
            larger = df[df['Elevation'] > elevation]

            if smaller.empty:
                continue  # Elevation below table range
            if larger.empty:
                # Use the last row if elevation is at or above max
                elev1 = elev2 = df['Elevation'].iloc[-1]
                row1 = row2 = df.index[-1]
            else:
                elev1 = smaller['Elevation'].iloc[-1]  # Closest smaller
                elev2 = larger['Elevation'].iloc[0]  # Closest larger
                row1 = smaller.index[-1]
                row2 = larger.index[0]

            # Get decimal part for column interpolation
            decimal_part = elevation - np.floor(elevation)  # type: ignore
            col1_idx = int(np.floor(decimal_part * 10))  # e.g., 5 for 0.56
            col2_idx = col1_idx + 1 if col1_idx < 9 else col1_idx  # e.g., 6 or stay at 9

            # Map to column names
            col1 = f"{col1_idx}.0"
            col2 = f"{col2_idx}.0"

            # Check if columns exist
            if col1 not in df.columns or col2 not in df.columns:
                continue

            # Get the four surface area values for bilinear interpolation
            f11 = df.loc[row1, col1]  # Smaller elevation, smaller column
            f12 = df.loc[row1, col2]  # Smaller elevation, larger column
            f21 = df.loc[row2, col1]  # Larger elevation, smaller column
            f22 = df.loc[row2, col2]  # Larger elevation, larger column

            # Interpolation weights
            x = elevation  # Input elevation
            x1, x2 = elev1, elev2  # Row elevations (10-ft increments)
            y = decimal_part * 10  # Column offset (0.0 to 9.0)
            y1, y2 = col1_idx, col2_idx  # Column indices

            # Avoid division by zero (if elev1 == elev2 or col1 == col2)
            if x2 == x1:
                tx = 0.5  # Arbitrary midpoint if same row
            else:
                tx = (x - x1) / (x2 - x1)  # Row interpolation factor

            if y2 == y1:
                ty = 0.5  # Arbitrary midpoint if same column
            else:
                ty = (y - y1) / (y2 - y1)  # Column interpolation factor

            # Bilinear interpolation
            area = (f11 * (1 - tx) * (1 - ty) +
                    f21 * tx * (1 - ty) +
                    f12 * (1 - tx) * ty +
                    f22 * tx * ty)

            areas[i] = area

        return areas

    @staticmethod
    def merge_dictionaries(dict1:dict, dict2:dict, description:str=''):
        """
        Merges dict2 into dict1. Adds new key-value pairs, checks for matching values,
        and prints warnings for differing values.

        Args:
            dict1 (dict): The first dictionary (modified in-place).
            dict2 (dict): The second dictionary to merge into dict1.
            description (str): Description for debug print

        Returns:
            dict: The modified dict1.
        """
        for key, value2 in dict2.items():
            if key not in dict1:
                dict1[key] = value2
            elif dict1[key] != value2:
                print(f"\tMerge has different values: \t'{key}' {dict1[key]} {description}: {value2}")
        return dict1

    @staticmethod
    def get_strings_from_file(filename, prefix):
        """
        Reads a text file and returns a list of strings that start with the given prefix,
        with the prefix removed from each string.

        Args:
            filename (str): Path to the text file
            prefix (str): The prefix to match at the start of each line

        Returns:
            list: List of strings that started with the prefix, with prefix removed

        Raises:
            FileNotFoundError: If the specified file doesn't exist
            IOError: If there's an error reading the file
        """
        result = []
        try:
            with open(filename, 'r') as file:
                for line in file:
                    # Remove leading/trailing whitespace and check if line starts with prefix
                    line = line.strip()
                    if line.startswith(prefix):
                        # Add the string with prefix removed
                        result.append(line[len(prefix):].strip())
            return result
        except FileNotFoundError:
            raise FileNotFoundError(f"The file '{filename}' was not found")
        except IOError as e:
            raise IOError(f"Error reading file '{filename}': {str(e)}")

    @staticmethod
    def report_excel_row_index_errors(report, file_name):
        if report:
            row_index_errors = WaterYear.get_strings_from_file(file_name, 'Row index invalid')
            if row_index_errors:
                report.page_break()
                report.header('Excel Errors', 1)
                report.header('Row Indexing Errors', 2)
                string = ''
                for text in row_index_errors:
                    match = re.match(r"'([A-Za-z]\d+)'|[A-Za-z]\d+", text)
                    if match:
                        # Return the captured group (if quoted) or the full match
                        cell = match.group(1) if match.group(1) else match.group(0)
                        string += text.replace(cell, f'**{cell}**')
                        string += '\n'
                    else:
                        pass
                report.add_markdown_paragraph(string, font='Courier New')

    @staticmethod
    def output(length):
        return np.zeros(length, dtype=np.float64)

    @staticmethod
    def get_out(variable_name:str, out:dict, length:int)->np.ndarray:
        if variable_name in out:
            return out[variable_name]
        else:
            out[variable_name] = np.zeros(length, dtype=np.float64)
            return out[variable_name]

    @staticmethod
    def comments_for_vars(variable_names, comments, dates, print_comments=False):
        results = []
        variables_comments = []

        for variable_name in variable_names:
            arr = comments.get(variable_name)
            if arr and len(arr):
                variables_comments.append(arr)
            else:
                print('\tWARN comments_for_vars variable not found: ', variable_name)
                return None

        num_days = len(variables_comments[0])

        for day in range(0, num_days):
            for variable_comments, variable_name in zip(variables_comments, variable_names):
                if variable_comments[day] and type(variable_comments[day]) is str:
                    if print_comments:
                        print(f'\t{dates[day]} \'{variable_name}\'')
                        print(textwrap.indent(variable_comments[day], '\t\t'))
        return results

    @staticmethod
    def sort_variables_by_type(variable_names, equation_dict, formula_dict, value_dict):
        variable_names_from_equations = []
        variable_names_inputs = []
        variable_names_empty = []

        for variable_name in variable_names:
            equations = equation_dict.get(variable_name)
            if not equations:
                pass
            formulas = formula_dict.get(variable_name)
            if not formulas:
                pass
            values = value_dict.get(variable_name)
            if not isinstance(values, list) or not values:
                pass

            has_formulas = False
            has_data = False
            if formulas and equations and isinstance(equations, list) and len(equations):
                num_days = len(equations)
                for day in range(0, num_days):
                    equation = equations[day]
                    formula = formulas[day]
                    if isinstance(equation, str) and formula.startswith("="):
                        if WaterYear.is_formula_constant_only(formula):
                            has_formulas = True
                        elif isinstance(equation, (int, float)):
                            if not math.isnan(equation):
                                has_data = True
                        else:
                            has_data = True
                    elif isinstance(equation, (int, float)):
                        if not math.isnan(equation):
                            has_data = True
                    # elif equation and math.isnan(equation):
                        # pass
            elif values:
                num_days = len(values)
                for day in range(0, num_days):
                    value = values[day]
                    if isinstance(value, (int, float)):
                        if not math.isnan(value):
                            has_data = True
                            break
                    elif WaterYear.is_valid_number(value):
                        has_data = True
                        break

            if has_formulas:
                variable_names_from_equations.append(variable_name)
            elif has_data:
                variable_names_inputs.append(variable_name)
            else:
                variable_names_empty.append(variable_name)

        return variable_names_from_equations, variable_names_inputs, variable_names_empty

    @staticmethod
    def compile_wildcard_regex(wildcard):
        """
        Convert "*.xls;*.xlsx" → regex that matches any of those extensions (case-insensitive)
        Returns compiled regex or None if wildcard == "*"
        """
        if not wildcard or wildcard.strip() == "*":
            return None  # match all

        patterns = []
        for part in wildcard.split(";"):
            part = part.strip()
            if not part:
                continue
            # Convert *.ext → \.ext$  (literal dot, end of string)
            esc = re.escape(part.replace("*", ""))  # escape . in case of weird chars
            esc = esc.replace(r"\.", ".")  # unescape the dot
            pattern = esc + "$"  # must end with extension
            patterns.append(pattern)

        if not patterns:
            return None

        combined = "|".join(patterns)
        return re.compile(combined, re.IGNORECASE)

    @staticmethod
    def is_valid_number(s: str) -> bool:
        """Return True if string is a valid int or float."""
        if not isinstance(s, str):
            return False
        s = s.strip()  # Ignore leading/trailing whitespace
        if not s:  # Empty string?
            return False
        try:
            float(s)  # float() also accepts int-like strings
            return True
        except ValueError:
            return False
    @staticmethod
    def get_file_list(path, wildcard):
        files = []
        try:
            entries = os.listdir(path)
            for entry in entries:
                full_path = os.path.join(path, entry)
                if os.path.isfile(full_path):
                    # Wildcard filtering
                    # if self.wildcard == "*" or any(entry.lower().endswith(ext.strip().lower().lstrip("*."))
                    #                               for ext in self.wildcard.split(";") if ext.strip()):
                    wildcard_regex = WaterYear.compile_wildcard_regex(wildcard)
                    if wildcard_regex and not wildcard_regex.search(entry):
                        continue
                    size = os.path.getsize(full_path)
                    mtime = os.path.getmtime(full_path)
                    # dt = wx.DateTime.FromTimeT(int(mtime))  # Fixed line
                    # modified = dt.Format("%Y-%m-%d %H:%M")
                    files.append((entry, size, mtime, full_path))

            # Sort reverse alphabetical (case-insensitive)
            files.sort(key=lambda x: x[0].lower(), reverse=True)
        except Exception as e:
            print(f'get_file_list error {e}')
        return files

    @staticmethod
    def is_file_newer(file1_path, file2_path):
        """
        Compare the last modified dates of two files and return True if file1 is newer than file2.

        Args:
            file1_path (str or Path): Path to the first file.
            file2_path (str or Path): Path to the second file.

        Returns:
            bool: True if file1 is newer than file2, False otherwise.

        Raises:
            FileNotFoundError: If either file does not exist.
            OSError: If there's an error accessing file metadata.
        """
        try:
            file1 = Path(file1_path)
            file2 = Path(file2_path)

            # Check if both files exist
            if not file1.is_file():
                raise FileNotFoundError(f"File not found: {file1}")
            if not file2.is_file():
                raise FileNotFoundError(f"File not found: {file2}")

            # Get modification times
            file1_mtime = file1.stat().st_mtime
            file2_mtime = file2.stat().st_mtime

            # Return True if file1 is newer
            return file1_mtime > file2_mtime

        except FileNotFoundError as e:
            raise FileNotFoundError(str(e))
        except OSError as e:
            raise OSError(f"Error accessing file metadata: {str(e)}")

    @staticmethod
    def parse_first_parens_expression(s: str):
        """
        Parses the first top-level ( ... ) and returns:
            - (num1, op, num2)  if two numbers connected by +, -, *, /, ×, ÷
            - (num1, None, None) if only one number/expression
            - (None, None, None) if no valid match
        Supports fractions like 123/456, scientific notation, commas in large numbers
        """
        # Match first top-level parentheses (ignores nested ones)
        match = re.search(r'\(([^\(\)]*)\)', s)
        if not match:
            return None, None, None

        content = match.group(1)

        # Remove comma thousands separators for parsing
        clean_content = re.sub(r'(\d),(\d)', r'\1\2', content)

        # Regex for a full "number op number" expression
        # Captures: (number1)  (op)  (number2)
        pattern2 = r'''
            ^\s*                                     # Start + optional whitespace
            ([+-]?(?:\d{1,3}(?:_\d{3})*(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)   # num1: 1_234, 1.23, .5, 1e-10
            \s*                                      # optional spaces
            ([+/×÷])                                 # ← ONLY these are operators: +, /, ×, ÷
            \s*                                      # optional spaces
            ([+-]?(?:\d{1,3}(?:_\d{3})*(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)   # num2
            \s*$                                     # optional trailing space + end
        '''
        pattern = r'''
            ^\s*                                      # start + optional space
            ([+-]?(?:\d{1,3}(?:_\d{3})*(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?|\d+/\d+)\s*   # num1 (incl fractions)
            ([+\-*/×÷])\s*                            # operator
            ([+-]?(?:\d{1,3}(?:_\d{3})*(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?|\d+/\d+)\s*   # num2
            $                                         # end
        '''

        m = re.search(pattern, clean_content, re.VERBOSE)
        if m:
            num1, op, num2 = m.groups()
            # Restore commas for display if needed, or keep clean
            return num1.strip(), op.strip(), num2.strip()

        # If no operator: check for single number/fraction
        single_num = re.search(r'^[\s\d\./eE+-]+$', clean_content)
        if single_num:
            num_str = content.strip()  # keep original formatting
            op = None
            parts = []
            num_str2 = None
            if '+' in num_str:
                parts = num_str.split('+')
                op = '+'
            elif '*' in num_str:
                parts = num_str.split('*')
                op = '*'
            elif '/' in num_str:
                parts = num_str.split('/')
                op = '/'
            if len(parts) > 1:
                num_str = parts[0].strip()
                num_str2 = parts[1].strip()

            return num_str, op, num_str2

        return None, None, None

    @staticmethod
    def get_two_numbers_from_parens(s: str):
        """
        Extracts up to two numeric expressions from the first top-level parentheses group.
        Returns:
            - (num1, None)   if one number/fraction found
            - (num1, num2)   if two numbers/fractions found
            - (None, None)   if no match or invalid
        Supports: 1.23, 123/456, -0.5e-10, etc.
        """
        # Match the first top-level ( ... ) — handles nested parens correctly
        match = re.search(r'\(([^\(\)]*)\)', s)
        if not match:
            return None, None

        content = match.group(1).strip()

        # Define what counts as a "number" — integer, float, fraction, scientific notation
        number_pattern = r'-?(?:\d+\.\d*|\.\d+|\d+)(?:[eE][+-]?\d+)?'
        fraction_pattern = r'-?\d+\s*\/\s*\d+'  # e.g. 182003/645272 or  1 /  2

        # Find all number-like tokens: fractions first (to avoid splitting them), then regular numbers
        tokens = re.findall(f'{fraction_pattern}|{number_pattern}', content)

        # Remove empty strings and clean up
        tokens = [t.strip() for t in tokens if t.strip()]

        if not tokens:
            return None, None
        elif len(tokens) == 1:
            return tokens[0], None
        else:
            return tokens[0], tokens[1]

    @staticmethod
    def find_first_number(formula):
        """
        Find the first number (integer or float) in an Excel formula.
        Returns its value and start/end indices.

        Parameters:
        - formula (str): The Excel formula string.

        Returns:
        - tuple: (value, start_index, end_index) or None if no number found.
        """
        # Pattern for first number (integer or float) after optional =
        pattern = r'^=?\d+\.?\d*'

        match = re.match(pattern, formula)
        if match:
            value = float(match.group().lstrip('='))  # Remove = and convert to float
            start = match.start() + (1 if formula.startswith('=') else 0)  # Adjust for =
            end = match.end()
            return value, start, end
        return None

    @staticmethod
    def find_floats_multiplied_by_104(formula):
        """
        Find numbers (X.Y or X format) multiplied by 1.04 in an Excel formula.
        Returns their values and start/end indices.

        Parameters:
        - formula (str): The Excel formula string.

        Returns:
        - list: List of tuples (value, start_index, end_index) for each matching number.
        """
        # Regex pattern for X.Y or X followed by *1.04
        # \b(\d+\.\d|\d)\b(?=\*1\.04)
        # - \d+\.\d matches X.Y (e.g., 3.2, 12.3)
        # - \d matches X (e.g., 2)
        # - | is OR operator
        # - (?=\*1\.04) is positive lookahead for *1.04
        pattern = r'\b(\d+\.\d|\d)\b(?=\*1\.04)'

        # Find all matches
        matches = [(float(match.group()), match.start(), match.end())
                   for match in re.finditer(pattern, formula)]

        return matches

    @staticmethod
    def find_floats_multiplied_by_106(formula):
        """
        Extract float constant multiplied by 1.06 in Excel formula.

        Parameters:
        - formula (str): Excel formula string

        Returns:
        - float or None: The constant value, or None if not found
        """
        # Pattern: (\d+\.\d+|[-+]?\d+) followed by *1\.06
        # Matches: 0.64, 1.23, 5, -2.5, etc.
        pattern = r'(\d+\.\d+|[-+]?\d+)\*1\.06'

        match = re.search(pattern, formula)
        if match:
            return float(match.group(1))
        return None

    @staticmethod
    def extract_constant_before_106(formula):
        """Extract SINGLE constant before *1.06) - FIXED"""
        pattern = r'(\d+\.\d+|\d+)\*1\.06\)'
        match = re.search(pattern, formula)
        return float(match.group(1)) if match else None

    @staticmethod
    def gallons_to_acre_feet(gallons):
        """
        Convert gallons to acre-feet.

        Parameters:
        - gallons (float): Volume in US gallons

        Returns:
        - float: Volume in acre-feet
        """
        return gallons / 325851

    @staticmethod
    def acre_feet_to_gallons(acre_feet):
        """
        Convert acre-feet to gallons.
        """
        return acre_feet * 325851

    @staticmethod
    def equations_for_vars(variable_names, equation_dict, formula_dict, value_dict, dates, file=None):
        variable_names_from_equations, variable_names_inputs, variable_names_empty =(
            WaterYear.sort_variables_by_type(variable_names, equation_dict, formula_dict, value_dict))

        # FIXME - These names are no longer here to remove, input channels have CFS on name and no equations now
        # variable_names_inputs_with_equations = ['CORTEZ NON - PROJ','CORTEZ PROJ']
        # variable_names_inputs_with_equations = ['UTE F&R', 'MWC', 'CORTEZ NON - PROJ','CORTEZ PROJ', 'UTE PROJ']
        variable_names_inputs_with_equations = []
        print('Equation processing, inputs, inputs with equations, etc.')
        variable_names_from_equations = WaterYear.remove_names(variable_names_from_equations, variable_names_inputs_with_equations)
        for variable_name in variable_names_inputs:
            if file:
                file.write(f'Input \'{variable_name}\'\n')
            else:
                print(f'Input \'{variable_name}\'')

        for variable_name in variable_names_inputs_with_equations:
            if file:
                file.write(f'Input with Formula \'{variable_name}\'\n')
            else:
                print(f'Input with Formula \'{variable_name}\'')

        for variable_name in variable_names_from_equations:
            if file:
                file.write(f'Output \'{variable_name}\'\n')
            else:
                print(f'\'Output {variable_name}\'')

            prev = ''
            prev_date = ''
            printed_date = ''

            equations = equation_dict.get(variable_name)
            # formulas = formula_dict.get(variable_name)
            values = value_dict.get(variable_name)
            num_days = len(equations)
            string = ''
            prev_diffs = None
            for day in range(0, num_days):
                if equations[day]:
                    if type(equations[day]) is str:
                        string = equations[day]
                    elif type(equations[day]) is float:
                        string = str(equations[day])
                    else:
                        string = str(equations[day])

                    if string == 'nan':
                        if variable_name == 'DATE':
                            if isinstance(values[day], Timestamp):
                                day = values[day].day
                                string = str(day)
                            elif isinstance(values[day], str):
                                string = str(values[day])
                                string = str(datetime.strptime(string, '%Y-%m-%d').day)
                            else:
                                pass
                        else:
                            string = str(values[day])

                    if prev != string:
                        date_for_day = dates[day]
                        if prev_date:
                            printed_date = date_for_day
                            prev_diffs = WaterYear.print_equation_diffs(prev, string, prev_date, date_for_day, prev_diffs, file)

                        prev = string
                        prev_date = date_for_day

            if printed_date != dates[num_days-1]:
                WaterYear.print_equation_diffs(prev, string, prev_date, dates[num_days-1],  prev_diffs, file)
        return variable_names_inputs, variable_names_from_equations

    @staticmethod
    def check_arrays_zero_or_nan(data_dict, key_list):
        """
        Check if NumPy arrays for given keys in a dictionary are all zeros or NaN.
        Returns a dictionary with keys and boolean results.
        """
        results = {}
        for key in key_list:
            if key in data_dict:
                array = data_dict[key]
                # Check if array is all zeros or NaN
                is_zero_or_nan = np.all((array == 0) | np.isnan(array))
                results[key] = is_zero_or_nan
            else:
                results[key] = False  # Key not found in dictionary
        return results

    @staticmethod
    def extract_to_new_dict(original_dict, name_list, remove_from_original=False):
        new_dict = {name: original_dict.get(name) for name in name_list if name in original_dict}

        if remove_from_original:
            for name in name_list:
                original_dict.pop(name, None)  # Remove key if it exists, None prevents KeyError

        return new_dict

    @staticmethod
    def convert_date(date_obj):
        result = None
        try:
            if not isinstance(date_obj, datetime):
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
            else:
                pass
            result = f"{date_obj.strftime('%b')}-{date_obj.day}"
        except Exception as e:
            print(f"convert_date exception: {date_obj} {e}")
        return result

    @staticmethod
    def print_equation_diffs(string1, string2, start_date, end_date, prev_diffs, file=None):
        start = WaterYear.convert_date(start_date)
        end = WaterYear.convert_date(end_date)
        diffs = WaterYear.diff_strings(string1, string2)
        if file:
            if prev_diffs:
                if prev_diffs[0].startswith('?'):
                    diff = prev_diffs.pop(0).rstrip()
                    file.write(f'\t\t{diff}\n')
                file.write(f'\t {start} to {end}\n')
                for diff in prev_diffs:
                    if diff.startswith('-'):
                        pass
                    else:
                        diff = diff.rstrip()
                        file.write(f'\t\t{diff}\n')
            else:
                file.write(f'\t {start} to {end}\n')
                diff = diffs.pop(0).rstrip()
                diff= ' ' + diff[1:]
                file.write(f'\t\t{diff}\n')
        else:
            print(f'\t   {start} to {end}')
            print(textwrap.indent(string1, '\t\t'))
            # print(diffs)
        return diffs

    @staticmethod
    def diff_strings(str1, str2):
        """
        Compute the difference between two strings.

        Args:
            str1 (str): First string
            str2 (str): Second string

        Returns:
            list: List of strings representing the diff
        """
        # Ensure inputs are strings
        str1 = str(str1) if str1 is not None else ""
        str2 = str(str2) if str2 is not None else ""

        # Use difflib to compute differences
        differ = difflib.Differ()
        diff = list(differ.compare(str1.splitlines(), str2.splitlines()))

        return diff

    @staticmethod
    def remove_names(main_list, names_to_remove):
        """
        Remove all occurrences of names_to_remove from main_list.

        Args:
            main_list (list): The list to remove names from
            names_to_remove (list): The list of names to remove

        Returns:
            list: New list with names removed
        """
        # Create a new list to avoid modifying the original
        result_list = main_list.copy()

        # Remove each name if it exists in the list
        for name in names_to_remove:
            # Handle NaN comparison safely
            if isinstance(name, (int, float)) and math.isnan(name):
                result_list = [x for x in result_list if not (isinstance(x, (int, float)) and math.isnan(x))]
            else:
                if name in result_list:
                    result_list.remove(name)
                else:
                    print(f'\tremove_names: {name} not found')

        return result_list
    @staticmethod
    def is_formula_constant_only(formula):
        """
        Check if an Excel formula contains only constants and no cell references.

        Args:
            formula (str): The formula string (e.g., '=SUM(1,2,3)', '=A1+B2')

        Returns:
            bool: True if the formula uses only constants, False if it references cells/ranges.
        """
        if not formula or not formula.startswith("="):
            return False  # Not a formula

        # Remove the leading '='
        formula = formula.lstrip("=")

        # Regular expression to match cell references and ranges
        # Matches: A1, $A$1, A1:B2, Sheet1!A1, named ranges (e.g., MyRange)
        cell_ref_pattern = re.compile(
            r"""
            [A-Za-z]+!\$?[A-Za-z]+\$?\d+ |              # Sheet1!A1 or Sheet1!$A$1
                \$?[A-Za-z]+\$?\d+ |                    # A1, $A$1, A$1, $A1
                \$?[A-Za-z]+\$?\d+:\$?[A-Za-z]+\$?\d+ | # A1:B2, $A$1:$B$2
                [A-Za-z_][A-Za-z0-9_]*
            """,
            re.VERBOSE
        )

        # Find all matches in the formula
        matches = cell_ref_pattern.findall(formula)

        # If any matches are found, the formula references cells or ranges
        if len(matches):
            return True
        else:
            return False

    @staticmethod
    def correction(variables, month, day, variable_name, new_value):
        if variable_name in variables:
            exl = variables.get(variable_name)
            idx = WaterYear.index_for_month_day(variables.get('DATE'), month, day)
            if idx:
                exl[idx] = new_value
        else:
            if variables in variable_name:
                print(f'correction failed {variable_name}')

    @staticmethod
    def remove_keys_from_dictionary(dict1, dict2):
        dict1 = {k: v for k, v in dict1.items() if k not in dict2}
        return dict1


    @staticmethod
    def remove_keys_from_dict(dict1, dict2, dict3):
        # Create a copy of dict1 to avoid modifying it directly
        result = dict1.copy()
        # Remove keys from dict1 that are in dict2 or dict3
        for key in dict2.keys() | dict3.keys():
            result.pop(key, None)  # Use pop with default None to avoid KeyError
        return result


    @staticmethod
    def remove_none_keys_from_dictionary(my_dict):
        return {k: v for k, v in my_dict.items() if v is not None}

    @staticmethod
    def index_for_month_day(datetime_arr, target_month, target_day):
        dt_objects = [np.datetime64(dt).astype(object) for dt in datetime_arr]
        months = np.array([dt.month for dt in dt_objects])
        days = np.array([dt.day for dt in dt_objects])
        match_idx = np.where((months == target_month) & (days == target_day))[0]
        if len(match_idx) == 0:
            idx = None
        else:
            idx = match_idx[0]
        return idx

    @staticmethod
    def day_for_date(datetime_arr, date_str):
        if date_str == 'Feb-29':
            dt = datetime.strptime('Feb-28', '%b-%d')
            day_of_month = 29
        else:
            dt = datetime.strptime(date_str, '%b-%d')
            day_of_month = dt.day
        month = dt.month
        day = WaterYear.index_for_month_day(datetime_arr, month, day_of_month)
        if day is None:
            day = len(datetime_arr)
        return day

    @staticmethod
    def add(d, keys, start_date=None, end_date=None):
        """
        Return a single NumPy array with the element-wise sum of arrays for the specified keys.
        Print a warning and skip if a key is not found.

        Parameters:
        d (dict): Dictionary with NumPy arrays as values
        keys (list): List of keys to process

        Returns:
        np.ndarray: Element-wise sum of the arrays (floats)
        """
        valid_arrays = []
        for key in keys:
            if key in d:
                arr = np.asarray(d[key], dtype=float)
                valid_arrays.append(arr)
            else:
                print(f"Warning: Key '{key}' not found in dictionary")

        if not valid_arrays:
            return np.array([])

        # Check if all arrays have the same length
        lengths = [len(arr) for arr in valid_arrays]
        if len(set(lengths)) > 1:
            raise ValueError("All arrays must have the same length")

        if start_date and end_date:
            arrays = np.stack(valid_arrays)
            result = np.sum(arrays[:, start_date:end_date], axis=0)
        else:
            result = np.sum(valid_arrays, axis=0)
        return result

    @staticmethod
    def set_array_to_value(datetime_arr, target_month, target_day, value_arr, set_value):
        idx = WaterYear.index_for_month_day(datetime_arr, target_month, target_day)
        if not idx:
            # No match found, set entire array to set_value
            value_arr[:] = set_value
        else:
            # Set elements from start to idx (inclusive) to set_value
            value_arr[:idx + 1] = set_value

        return value_arr

    @staticmethod
    def positive_values(arr):
        """
        Return a NumPy array with positive values unchanged and negative values replaced with 0.0.

        Parameters:
        arr (np.ndarray): Input array

        Returns:
        np.ndarray: Array with negative values set to 0.0
        """
        arr = np.asarray(arr, dtype=float)  # Ensure input is a float array
        return np.where(arr > 0, arr, 0.0)

    @staticmethod
    def negative_as_positive(arr):
        """
        Return a NumPy array with negative values flipped to positive and positive values set to 0.0.

        Parameters:
        arr (np.ndarray): Input array

        Returns:
        np.ndarray: Array with negative values flipped and positive values set to 0.0
        """
        arr = np.asarray(arr, dtype=float)  # Ensure input is a float array
        return np.where(arr < 0, -arr, 0.0)

    @staticmethod
    def compute_int_diffs(logger, variable_name, arr, first_ref):
        if np.isnan(arr).any() or np.isinf(arr).any():
            # print("compute_int_diffs: Has NaN:", np.isnan(arr).any())
            bad_mask = ~np.isfinite(arr) | np.isnan(arr)
            indices = np.where(bad_mask)
            logger.log_message(f'\tcompute_int_diffs: Bad values in {variable_name}: {indices}')
            # if bad_mask.any():
            #    print("\tBad values:", arr[bad_mask])
        arr = np.asarray(arr, dtype=int)

        # Initialize output array with the same length as input
        diffs = np.zeros(len(arr), dtype=int)

        # Compute first difference using first_ref
        diffs[0] = arr[0] - first_ref

        if len(arr) > 1:
            diffs[1:len(arr)] = arr[1:] - arr[:-1]
        return diffs

    @staticmethod
    def compute_float_diffs(logger, variable_name, arr, first_ref):
        if np.isnan(arr).any() or np.isinf(arr).any():
            # print("compute_int_diffs: Has NaN:", np.isnan(arr).any())
            bad_mask = ~np.isfinite(arr) | np.isnan(arr)
            indices = np.where(bad_mask)
            logger.log_message(f'\tcompute_float_diffs: Bad values in {variable_name}: {indices}')
            # if bad_mask.any():
            #    print("\tBad values:", arr[bad_mask])
        arr = np.asarray(arr, dtype=float)

        # Initialize output array with the same length as input
        diffs = np.zeros(len(arr), dtype=float)

        # Compute first difference using first_ref
        diffs[0] = arr[0] - first_ref

        if len(arr) > 1:
            diffs[1:len(arr)] = arr[1:] - arr[:-1]
        return diffs
