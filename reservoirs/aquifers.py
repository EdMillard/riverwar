"""
Copyright (c) 2026 Ed Millard

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute copies of the Software, and
to permit persons to whom the Software is furnished to do so, subject to the
following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""
from pathlib import Path
from reservoirs.reservoir import Reservoir
import colorado.lb as lb
import openpyxl
from sheet import sheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from typing import List, Dict, Union
import pandas as pd
import csv

# Bar graph
# https://www.azwater.gov/recharge/accounting
# By year query, Colorado River only
# https://app.azwater.gov/querycenter/query.aspx?qrysessionid=DED85701C874F812E053A564850A4538
# By user query
# https://app.azwater.gov/querycenter/query.aspx?qrysessionid=DF64BF23ECA02B33E053A564850A6170

class LTSC:
    def __init__(self, path:Path, name:str, headers:List[str]):
        self.name = name
        self.df = sheet.read_csv(path / f'{name}.csv', sep='\s+')
        totals = {}
        for col in headers:
            totals[col] = Reservoir.get_float_value(self.df, 'Total', col)

        self.cut_to_aquifer = totals['Cut_to_Aquifer']
        self.water_delivered = totals['Water_Delivered']
        self.annual_recovery = totals['Annual_Recovery']
        self.credits_recovered = totals['LTS_Credits_Recovered']
        self.credits_extinguished = totals['LTS_Credits_Extinguished']
        self.et_losses = totals['Evaporation_Transpiration_Losses']
        self.other_losses = totals['Other_Losses']

        stored = self.water_delivered - self.annual_recovery - self.credits_recovered
        stored_adjusted = stored - self.cut_to_aquifer  # Unclear if we can count this as stored water, no accounting
        stored_adjusted = stored_adjusted - self.et_losses - self.other_losses # - self.credits_extinguished
        self.stored = stored_adjusted

class Aquifers(Reservoir):
    def __init__(self):
        headers:List[str] = [lb.AQUIFER, lb.AQUIFER_INFLOW, lb.AQUIFER_RELEASE]
        super().__init__('AZ Aquifers', headers)
        self.start_year = 1989
        self.end_year = 2023
        self.years: List[int] = list(range(self.start_year, self.end_year+1))
        self.out_headers = ['Water_Delivered', 'Annual_Recovery', 'Evaporation_Transpiration_Losses', 'Cut_to_Aquifer',
                       'Other_Losses', 'LTS_Credits_Recovered', 'LTS_Credits_Extinguished', 'Other_Adjustment']

        self.path: Path = Path('data/ADWR/AMA')
        # self.recharge_summary_data_from_excel(self.path, self.years)

        ltsc_total = LTSC(self.path, 'total', self.out_headers)
        self.active_capacity_af = ltsc_total.stored + ltsc_total.cut_to_aquifer

        ltscs:List[LTSC] = [LTSC(self.path, 'phx', self.out_headers),
                            LTSC(self.path, 'pin', self.out_headers),
                            LTSC(self.path, 'tuc', self.out_headers)]

        ama_stored = 0
        for ltsc in ltscs:
            ama_stored += ltsc.stored
            if ltsc.name == 'phx':
                color = lb.PHX_COLOR
            elif ltsc.name == 'pin':
                color = lb.PINAL_COLOR
            elif ltsc.name == 'tuc':
                color = lb.TUCSON_COLOR
            else:
                color = '#0'
            self.reserved_parts.append((ltsc.name, ltsc.stored, color))

        # Elevations
        #
        # Must be called first
        self.dead_pool_feet = 0
        self.dead_pool_af = 0

        self.full_feet = 0
        self.full_af = 0

        # Critical
        self.power_head_target_feet = 0
        self.power_head_target_af = 0

        self.power_head_min_feet = 0
        self.power_head_min_af = 0

        self.turbine_intake_feet = 0
        self.turbine_intake_af = 0
        self.critical_elevations_feet = [("Safe Power Head", self.power_head_min_feet, self.power_head_min_af, Reservoir.non_power_pool_color),
                                         ("Min Power Head", self.power_head_target_feet, self.power_head_target_af, Reservoir.low_power_pool_color)]

        # Current
        #
        self.elevation_feet = 0
        # self.active_capacity_af = 9000000

        # Inflow

        self.inflow_actual_af = 0
        if self.inflow_actual_af:
            self.inflow_parts = [("Actual", self.inflow_actual_af, Reservoir.inflow_actual_color),
                                  ("Projected", 0, Reservoir.inflow_projected_color)]

        # Outflow
        self.outflow_actual_af = 0
        self.release_af = 0
        self.outflow_projected_af = self.release_af -  self.outflow_actual_af
        if self.outflow_actual_af or self.outflow_projected_af:
            self.outflow_parts = [("Actual", self.outflow_actual_af, Reservoir.outflow_actual_color),
                              ("Projected", self.outflow_projected_af, Reservoir.outflow_projected_color)]

    @staticmethod
    def add_total_row(df: pd.DataFrame, decimals: int = 2) -> pd.DataFrame:
        """
        Adds a 'Total' row with clean rounded totals.
        """
        df = df.copy()

        # Convert columns after Year to numeric
        for col in df.columns[1:]:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Sum and round to avoid floating point noise
        totals = df.iloc[:, 1:].sum().round(decimals)

        # Create Total row
        total_row = pd.Series(
            ['Total'] + totals.tolist(),
            index=df.columns,
            name='Total'
        )

        # Append
        df = pd.concat([df, total_row.to_frame().T], ignore_index=True)

        return df

    def get_ama_node(self, ama_name: str, headers: List[str], nodes: Dict) -> pd.DataFrame:
        node_name = f"{ama_name}"
        df: Union[pd.DataFrame, None] = nodes.get(node_name, None)
        if df is None:
            df: pd.DataFrame = sheet.create_df(self.start_year, self.end_year, headers, zero=True)
            nodes[node_name] = df
        else:
            pass

        return df

    @staticmethod
    def save_clean_csv(df: pd.DataFrame, out_csv_path, decimals: int = 2):
        df = df.copy()

        # === CRITICAL: Force numeric + round BEFORE writing ===
        for col in df.columns[1:]:  # skip Year column
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            df[col] = df[col].round(decimals)  # round in memory

        # Now write with float_format
        df.to_csv(
            str(out_csv_path),
            index=False,
            quoting=csv.QUOTE_NONE,
            escapechar='\\',
            sep=' ',  # space separated
            float_format=f'%.{decimals}f'
        )

    def recharge_summary_data_from_excel(self, out_path: Path, years: List[int]):
        wb: Workbook = openpyxl.load_workbook('excel/ADWR_Data_Warehouse_Recharge_Summary_Data_2023.xlsx', data_only=True)
        ws: Worksheet = wb['Sheet1']

        sheet.ensure_directory(out_path)

        header_row: int = 2

        headers: List[str] = []
        for column_index in range(ws.min_column, 20):
            header: str = ws.cell(row=header_row, column=column_index).value
            headers.append(header)

        self.df: pd.DataFrame = sheet.create_df(self.start_year, self.end_year, self.out_headers, zero=True)

        nodes: dict[str, pd.DataFrame] = {}
        df: Union[pd.DataFrame, None] = None
        year: int = 0
        ama_name: str | None = None
        category: str | None = None
        parent_water_type: str | None = None
        specific_water_type: str | None = None
        recharge_method: str | None = None
        recharge_element: str | None = None
        water_delivered_total:float = 0
        annual_recovery_total:float = 0
        et_total:float = 0
        cut_total:float = 0
        other_losses_total:float = 0
        lts_credits_recovered_total:float = 0
        lts_credits_extinguished_total:float = 0
        other_adjustment_total:float = 0
        finished:bool = False
        max_row = ws.max_row
        for row in ws.iter_rows(min_row=header_row+2):
            column_index = 0
            for cell in row:
                if column_index < len(headers):
                    header = headers[column_index]
                    if header is None:
                        pass
                    elif header == 'Year':
                        if cell.value == 2023:
                            pass
                        if cell.value is not None:
                            year = int(cell.value)
                        else:
                            finished = True
                            break
                    elif header == 'AMA':
                        ama_name = cell.value.lower()
                        df = self.get_ama_node(ama_name, self.out_headers, nodes)
                    elif header == 'Category':
                        category = cell.value
                    elif header == 'Parent Water Type or Element':
                        parent_water_type = cell.value
                    elif header == 'Specific Water Type':
                        specific_water_type = cell.value
                    elif header == 'Recharge Method':
                        recharge_method = cell.value
                    elif header == 'Recharge Element':
                        recharge_element = cell.value
                    elif header == 'Quantity':
                        quantity:float = float(cell.value)
                        if parent_water_type == 'CAP':
                            column_name = recharge_element.replace(" ", "_")
                            if recharge_element == 'Water Delivered':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                water_delivered_total += quantity
                            elif recharge_element == 'Annual Recovery':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                annual_recovery_total += quantity
                            elif recharge_element == 'Evaporation Transpiration Losses':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                et_total += quantity
                            elif recharge_element == 'Cut to Aquifer':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                cut_total += quantity
                            elif recharge_element == 'LTS Credits Recovered':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                lts_credits_recovered_total += quantity
                            elif recharge_element == 'LTS Credits Extinguished':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                lts_credits_extinguished_total += quantity
                            elif recharge_element == 'Other Losses':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                other_losses_total += quantity
                            elif recharge_element == 'Other Adjustment':
                                df.loc[df['Year'] == year, column_name] += quantity
                                self.df.loc[self.df['Year'] == year, column_name] += quantity
                                other_adjustment_total += quantity
                            else:
                                pass
                column_index += 1
            if finished:
                break

        stored = water_delivered_total - annual_recovery_total
        stored_adjusted = stored - et_total - cut_total - other_losses_total - lts_credits_extinguished_total - lts_credits_recovered_total
        self.active_capacity_af = stored_adjusted + cut_total

        decimals = 2
        for key, df in nodes.items():
            all_rows_are_zero = df.drop(columns='Year', errors='ignore').eq(0).all(axis=1).all()
            if not all_rows_are_zero:
                df_total = Aquifers.add_total_row(df)
                out_csv_path = out_path / f'{key}.csv'
                Aquifers.save_clean_csv(df_total, out_csv_path, decimals=decimals)

        df_total = Aquifers.add_total_row(self.df)
        out_csv_path = out_path / f'total.csv'
        Aquifers.save_clean_csv(df_total, out_csv_path, decimals=decimals)

